# -*- coding: utf-8 -*-
"""
textbooks.py — บัญชีหนังสือเรียน/แบบฝึกหัด (เงินอุดหนุนค่าหนังสือเรียน)
ทะเบียนหนังสือ (รับเข้า) + ใบเบิก (จ่ายออกให้ชั้นเรียน) + นำเข้า/ส่งออก Excel
ตารางใหม่สร้างอัตโนมัติด้วย init_school_db (ไม่ต้อง ALTER)
"""
import io
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from sqlalchemy.orm import Session

from app.database import get_db, get_data_dir
from app.models import TextBook, TextbookBerk, TextbookBerkItem
from app.thai_utils import (current_academic_year, parse_be_date, be_date_input,
                            thai_date)
from app.templating import templates
from app.routers.pages import get_school, _to_int, _to_float

router = APIRouter()

THAI_FONT = "TH Sarabun New"


def _years(db) -> list:
    ys = {r[0] for r in db.query(TextBook.year).distinct() if r[0]}
    ys |= {r[0] for r in db.query(TextbookBerk.year).distinct() if r[0]}
    cur = current_academic_year()
    ys |= set(range(cur - 2, cur + 2))
    return sorted(ys, reverse=True)


def _issued_map(db, year: int) -> dict:
    """book_id -> จำนวนที่เบิกออกไปแล้วรวม (ปีการศึกษานั้น)"""
    out: dict = {}
    berks = db.query(TextbookBerk).filter_by(year=year).all()
    for b in berks:
        for it in b.items:
            if it.book_id:
                out[it.book_id] = out.get(it.book_id, 0) + (it.qty or 0)
    return out


# ---------------- ทะเบียนหนังสือเรียน ----------------
@router.get("/textbooks", response_class=HTMLResponse)
def textbooks_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    yr = year or current_academic_year()
    books = (db.query(TextBook).filter_by(year=yr)
             .order_by(TextBook.level, TextBook.subject, TextBook.title).all())
    issued = _issued_map(db, yr)
    rows = [{"b": b, "issued": issued.get(b.id, 0),
             "left": (b.qty_received or 0) - issued.get(b.id, 0)} for b in books]
    return templates.TemplateResponse("textbooks.html", {
        "request": request, "school": get_school(db), "rows": rows,
        "year": yr, "years": _years(db),
        "total_qty": sum(b.qty_received or 0 for b in books),
        "total_amount": sum(b.amount for b in books),
        "saved": request.query_params.get("saved"),
    })


@router.post("/textbooks/add")
def textbook_add(db: Session = Depends(get_db), year: str = Form(""), level: str = Form(""),
                 subject: str = Form(""), title: str = Form(""), publisher: str = Form(""),
                 unit_price: str = Form("0"), qty_received: str = Form("0"), note: str = Form("")):
    yr = _to_int(year, current_academic_year())
    if (title or "").strip():
        db.add(TextBook(year=yr, level=level.strip(), subject=subject.strip(),
                        title=title.strip(), publisher=publisher.strip(),
                        unit_price=_to_float(unit_price, 0.0),
                        qty_received=_to_int(qty_received, 0), note=note.strip()))
        db.commit()
    return RedirectResponse(f"/textbooks?year={yr}&saved=1", status_code=303)


@router.post("/textbooks/{bid}/update")
def textbook_update(bid: int, db: Session = Depends(get_db), level: str = Form(""),
                    subject: str = Form(""), title: str = Form(""), publisher: str = Form(""),
                    unit_price: str = Form("0"), qty_received: str = Form("0"), note: str = Form("")):
    b = db.get(TextBook, bid)
    if b:
        b.level = level.strip(); b.subject = subject.strip()
        b.title = title.strip() or b.title; b.publisher = publisher.strip()
        b.unit_price = _to_float(unit_price, 0.0); b.qty_received = _to_int(qty_received, 0)
        b.note = note.strip()
        db.commit()
    yr = b.year if b else current_academic_year()
    return RedirectResponse(f"/textbooks?year={yr}&saved=1", status_code=303)


@router.post("/textbooks/{bid}/delete")
def textbook_delete(bid: int, db: Session = Depends(get_db)):
    b = db.get(TextBook, bid)
    yr = b.year if b else current_academic_year()
    if b:
        db.delete(b); db.commit()
    return RedirectResponse(f"/textbooks?year={yr}", status_code=303)


# ---------------- ใบเบิกหนังสือเรียน ----------------
@router.get("/textbooks/berk", response_class=HTMLResponse)
def berk_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    yr = year or current_academic_year()
    books = (db.query(TextBook).filter_by(year=yr)
             .order_by(TextBook.level, TextBook.title).all())
    berks = (db.query(TextbookBerk).filter_by(year=yr)
             .order_by(TextbookBerk.date, TextbookBerk.id).all())
    from app.models import Person
    persons = db.query(Person).order_by(Person.name).all()
    return templates.TemplateResponse("textbook_berk.html", {
        "request": request, "school": get_school(db), "books": books, "berks": berks,
        "year": yr, "years": _years(db), "today_be": be_date_input(datetime.now()),
        "persons": persons,
    })


@router.post("/textbooks/berk/add")
async def berk_add(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    yr = _to_int(form.get("year"), current_academic_year())
    berk = TextbookBerk(year=yr, berk_no=(form.get("berk_no") or "").strip(),
                        date=parse_be_date(form.get("date") or "") or datetime.now(),
                        recipient=(form.get("recipient") or "").strip(),
                        note=(form.get("note") or "").strip())
    db.add(berk)
    db.flush()
    book_ids = form.getlist("book_id")
    qtys = form.getlist("qty")
    for i, bid in enumerate(book_ids):
        bid_i = _to_int(bid, 0)
        q = _to_int(qtys[i] if i < len(qtys) else 0, 0)
        if bid_i and q > 0:
            db.add(TextbookBerkItem(berk_id=berk.id, book_id=bid_i, qty=q))
    db.commit()
    return RedirectResponse(f"/textbooks/berk?year={yr}", status_code=303)


@router.post("/textbooks/berk/{berk_id}/delete")
def berk_delete(berk_id: int, db: Session = Depends(get_db)):
    b = db.get(TextbookBerk, berk_id)
    yr = b.year if b else current_academic_year()
    if b:
        db.delete(b); db.commit()
    return RedirectResponse(f"/textbooks/berk?year={yr}", status_code=303)


@router.get("/textbooks/berk/{berk_id}/print", response_class=HTMLResponse)
def berk_print(berk_id: int, request: Request, db: Session = Depends(get_db)):
    b = db.get(TextbookBerk, berk_id)
    if not b:
        return RedirectResponse("/textbooks/berk", status_code=303)
    return templates.TemplateResponse("textbook_berk_print.html", {
        "request": request, "school": get_school(db), "b": b,
        "total": sum((it.qty or 0) * (it.book.unit_price if it.book else 0) for it in b.items),
    })


# ---------------- นำเข้า/ส่งออก Excel ----------------
@router.get("/textbooks/template.xlsx")
def berk_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "หนังสือเรียน"
    headers = ["ระดับชั้น", "กลุ่มสาระ/วิชา", "ชื่อหนังสือ", "สำนักพิมพ์",
               "ราคาต่อเล่ม", "จำนวนรับเข้า(เล่ม)", "หมายเหตุ"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name=THAI_FONT, bold=True, size=14)
        ws.column_dimensions[chr(64 + c)].width = 20
    ws.append(["ป.1", "ภาษาไทย", "ภาษาพาที ป.1", "สสวท.", 85, 40, ""])
    out = get_data_dir() / "documents"; out.mkdir(exist_ok=True)
    path = out / "แบบฟอร์มนำเข้าหนังสือเรียน.xlsx"
    wb.save(str(path))
    return FileResponse(path, filename=path.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/textbooks/import")
async def textbook_import(db: Session = Depends(get_db), year: str = Form(""),
                          file: UploadFile = File(...)):
    from openpyxl import load_workbook
    yr = _to_int(year, current_academic_year())
    data = await file.read()
    n = 0
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] and not (len(row) > 2 and row[2]):
                continue
            title = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if not title:
                continue
            db.add(TextBook(
                year=yr, level=str(row[0] or "").strip(),
                subject=str(row[1] or "").strip() if len(row) > 1 else "",
                title=title, publisher=str(row[3] or "").strip() if len(row) > 3 else "",
                unit_price=_to_float(row[4] if len(row) > 4 else 0, 0.0),
                qty_received=_to_int(row[5] if len(row) > 5 else 0, 0),
                note=str(row[6] or "").strip() if len(row) > 6 else ""))
            n += 1
        db.commit()
    except Exception:
        return RedirectResponse(f"/textbooks?year={yr}", status_code=303)
    return RedirectResponse(f"/textbooks?year={yr}&saved={n}", status_code=303)


@router.get("/textbooks/export.xlsx")
def textbook_export(db: Session = Depends(get_db), year: int | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    yr = year or current_academic_year()
    books = (db.query(TextBook).filter_by(year=yr)
             .order_by(TextBook.level, TextBook.subject, TextBook.title).all())
    issued = _issued_map(db, yr)
    school = get_school(db)
    wb = Workbook(); ws = wb.active; ws.title = f"หนังสือเรียน {yr}"[:31]
    thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append([f"ทะเบียนคุมหนังสือเรียน/แบบฝึกหัด  ปีการศึกษา {yr}"])
    ws.append([(school.name if school else "")])
    ws.append([])
    headers = ["ลำดับ", "ระดับชั้น", "กลุ่มสาระ/วิชา", "ชื่อหนังสือ", "สำนักพิมพ์",
               "ราคา/เล่ม", "รับเข้า", "เบิกออก", "คงเหลือ", "มูลค่ารับเข้า (บาท)"]
    ws.append(headers)
    n_col = len(headers)
    for c in range(1, n_col + 1):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_col)
        cell = ws.cell(4, c)
        cell.font = Font(name=THAI_FONT, bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(1, 1).font = Font(name=THAI_FONT, bold=True, size=16)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.cell(2, 1).font = Font(name=THAI_FONT, size=14)
    ws.cell(2, 1).alignment = Alignment(horizontal="center")
    for i, b in enumerate(books, start=1):
        iss = issued.get(b.id, 0)
        ws.append([i, b.level, b.subject, b.title, b.publisher, b.unit_price or 0,
                   b.qty_received or 0, iss, (b.qty_received or 0) - iss, b.amount])
    widths = [7, 12, 18, 30, 18, 11, 10, 10, 10, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font = Font(name=THAI_FONT, size=13)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[5].number_format = "#,##0.00"
        row[9].number_format = "#,##0.00"
    out = get_data_dir() / "documents"; out.mkdir(exist_ok=True)
    path = out / f"ทะเบียนหนังสือเรียน_ปีการศึกษา{yr}.xlsx"
    wb.save(str(path))
    return FileResponse(path, filename=path.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
