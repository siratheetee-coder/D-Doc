# -*- coding: utf-8 -*-
"""
admin.py — งานธุรการ
หน้าหลักธุรการ + ทะเบียนหนังสือรับ/ส่ง + บันทึกข้อความ + คำสั่งโรงเรียน (ออก Word) + นำเข้า Excel
เลขบันทึกข้อความ/คำสั่ง ใช้ชุดเลขกลางร่วมกับทุกงาน (suggest_doc_no/commit_doc_no)
"""
import re
import uuid
import urllib.request
from urllib.parse import urlparse
from pathlib import Path
from datetime import datetime

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from sqlalchemy.orm import Session

from app.database import get_db, get_data_dir
from app.models import (
    School, IncomingLetter, OutgoingLetter, OfficeMemo, SchoolOrder, Person, Department,
    OfficialLetter, CertificateBatch,
)
from app.services import file_upload
from app.services.doc_number import (
    suggest_next, commit_number, suggest_doc_no, commit_doc_no, check_doc_no, format_doc_no,
    parse_seq,
)
from app.services.office_doc import render_memo, render_order, render_official_letter
from app.services.admin_io import build_admin_template, import_admin_workbook, export_admin_register
from app.services.pdf_extract import extract_letter_fields
from app.thai_utils import current_fiscal_year, parse_be_date, be_date_input
from app.templating import templates
from app.routers.pages import get_school, _to_int, _to_float, serve_generated

router = APIRouter()

_DOCX = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF = "application/pdf"


def _ai_key() -> str:
    """AI key กลาง (หลังบ้าน) เฉพาะสมาชิก — คืน '' ถ้าไม่ใช่สมาชิก/ไม่ได้ตั้ง key"""
    from app.tenancy import current_school_id
    from app.accounts import ai_key_for
    return ai_key_for(current_school_id.get())

# ใช้ฟังก์ชันกลางจาก file_upload (alias ชื่อเดิมเพื่อความเข้ากันได้)
from app.services.file_upload import (
    SAFE_FILE_NAME as _SAFE_FILE_NAME, uploads_dir as _uploads_dir,
    detect_ext as _detect_ext, save_upload as _save_upload, fetch_file as _fetch_file,
)


def _render_letter_confirm(request, db, kind: str, pending_file: str, fields: dict):
    """หน้ายืนยันลงทะเบียน (เติมค่าจากการอ่านไฟล์ให้แล้ว) — รองรับ incoming/outgoing/memo"""
    fy = current_fiscal_year()
    ctx = {
        "request": request, "kind": kind, "fiscal_year": fy,
        "pending_file": pending_file, "fields": fields,
        "letter_date_be": be_date_input(fields.get("letter_date")),
        "is_pdf": pending_file.lower().endswith(".pdf"),
    }
    if kind == "incoming":
        ctx["sug_recv"] = suggest_next(db, "incoming", fy)
    elif kind == "outgoing":
        school = get_school(db)
        seq = suggest_next(db, "outgoing", fy)
        ctx["sug_no"] = f"{school.doc_prefix or 'ศธ'} {seq}/{fy}"
    elif kind == "memo":
        ctx["school"] = get_school(db)
        ctx["sug_memo"] = suggest_doc_no(db, "memo", fy)
    else:  # order
        ctx["school"] = get_school(db)
        ctx["sug_order"] = suggest_doc_no(db, "command", fy)
    return templates.TemplateResponse("letter_confirm.html", ctx)


# ---------------- Dashboard ----------------
@router.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request, db: Session = Depends(get_db)):
    fy = current_fiscal_year()
    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request, "school": get_school(db), "fiscal_year": fy,
        "n_in": db.query(IncomingLetter).filter_by(fiscal_year=fy).count(),
        "n_out": db.query(OutgoingLetter).filter_by(fiscal_year=fy).count(),
        "n_memo": db.query(OfficeMemo).filter_by(fiscal_year=fy).count(),
        "n_order": db.query(SchoolOrder).filter_by(fiscal_year=fy).count(),
        "recent_memos": db.query(OfficeMemo).order_by(OfficeMemo.id.desc()).limit(5).all(),
        "recent_orders": db.query(SchoolOrder).order_by(SchoolOrder.id.desc()).limit(5).all(),
    })


# ---------------- ทะเบียนหนังสือรับ ----------------
@router.get("/admin/incoming", response_class=HTMLResponse)
def incoming_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    fy = year or current_fiscal_year()
    rows = (db.query(IncomingLetter).filter_by(fiscal_year=fy)
            .order_by(IncomingLetter.recv_no).all())
    years = sorted({r[0] for r in db.query(IncomingLetter.fiscal_year).distinct()} | {fy}, reverse=True)
    return templates.TemplateResponse("incoming.html", {
        "request": request, "rows": rows, "fiscal_year": fy, "years": years,
        "sug_recv": suggest_next(db, "incoming", fy),
    })


@router.post("/admin/incoming")
def incoming_add(db: Session = Depends(get_db), fiscal_year: int = Form(...),
                 recv_no: str = Form(""), recv_date: str = Form(""), letter_no: str = Form(""),
                 letter_date: str = Form(""), from_org: str = Form(""), to_person: str = Form(""),
                 subject: str = Form(""), action_note: str = Form(""), pending_file: str = Form("")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    rno = _to_int(recv_no, 0) or suggest_next(db, "incoming", fy)
    fp = pending_file.strip() if _SAFE_FILE_NAME.match(pending_file.strip()) else ""
    db.add(IncomingLetter(
        fiscal_year=fy, recv_no=rno, recv_date=parse_be_date(recv_date),
        letter_no=letter_no.strip(), letter_date=parse_be_date(letter_date),
        from_org=from_org.strip(), to_person=to_person.strip(),
        subject=subject.strip(), action_note=action_note.strip(), file_path=fp,
    ))
    commit_number(db, "incoming", fy, rno)
    db.commit()
    return RedirectResponse("/admin/incoming", status_code=303)


@router.post("/admin/incoming/upload")
async def incoming_upload(request: Request, db: Session = Depends(get_db), file: UploadFile = File(...)):
    data = await file.read()
    ext = _detect_ext(data, file.filename or "")
    if not ext:
        return RedirectResponse("/admin/incoming?err=notpdf", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "incoming", name, fields)


@router.post("/admin/incoming/fetch")
async def incoming_fetch(request: Request, db: Session = Depends(get_db), url: str = Form("")):
    data, ext, err = _fetch_file(url)
    if err:
        return RedirectResponse(f"/admin/incoming?err={err}", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "incoming", name, fields)


@router.post("/admin/incoming/{rid}/delete")
def incoming_delete(rid: int, db: Session = Depends(get_db)):
    r = db.get(IncomingLetter, rid)
    if r:
        db.delete(r); db.commit()
    return RedirectResponse("/admin/incoming", status_code=303)


# ---------------- ทะเบียนหนังสือส่ง ----------------
@router.get("/admin/outgoing", response_class=HTMLResponse)
def outgoing_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    fy = year or current_fiscal_year()
    rows = (db.query(OutgoingLetter).filter_by(fiscal_year=fy)
            .order_by(OutgoingLetter.send_seq).all())
    years = sorted({r[0] for r in db.query(OutgoingLetter.fiscal_year).distinct()} | {fy}, reverse=True)
    school = get_school(db)
    seq = suggest_next(db, "outgoing", fy)
    sug_no = f"{school.doc_prefix or 'ศธ'} {seq}/{fy}"
    return templates.TemplateResponse("outgoing.html", {
        "request": request, "rows": rows, "fiscal_year": fy, "years": years, "sug_no": sug_no,
    })


@router.post("/admin/outgoing")
def outgoing_add(db: Session = Depends(get_db), fiscal_year: int = Form(...),
                 send_no: str = Form(""), date: str = Form(""), to_org: str = Form(""),
                 subject: str = Form(""), pending_file: str = Form("")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    m = re.search(r"(\d+)", send_no or "")     # เลขลำดับ = กลุ่มตัวเลขแรก (ข้ามอักษรนำ ศธ)
    seq = (int(m.group(1)) if m else 0) or suggest_next(db, "outgoing", fy)
    fp = pending_file.strip() if _SAFE_FILE_NAME.match(pending_file.strip()) else ""
    db.add(OutgoingLetter(
        fiscal_year=fy, send_seq=seq, send_no=send_no.strip(),
        date=parse_be_date(date), to_org=to_org.strip(), subject=subject.strip(), file_path=fp,
    ))
    commit_number(db, "outgoing", fy, seq)
    db.commit()
    return RedirectResponse("/admin/outgoing", status_code=303)


@router.post("/admin/outgoing/upload")
async def outgoing_upload(request: Request, db: Session = Depends(get_db), file: UploadFile = File(...)):
    data = await file.read()
    ext = _detect_ext(data, file.filename or "")
    if not ext:
        return RedirectResponse("/admin/outgoing?err=notpdf", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "outgoing", name, fields)


@router.post("/admin/outgoing/fetch")
async def outgoing_fetch(request: Request, db: Session = Depends(get_db), url: str = Form("")):
    data, ext, err = _fetch_file(url)
    if err:
        return RedirectResponse(f"/admin/outgoing?err={err}", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "outgoing", name, fields)


# ---------------- เสิร์ฟไฟล์ PDF/Word ที่อัปโหลด/แนบ ----------------
@router.get("/admin/uploaded/{name}")
def serve_uploaded(name: str):
    if not _SAFE_FILE_NAME.match(name):     # กัน path traversal (รับ .pdf/.docx)
        return RedirectResponse("/admin", status_code=303)
    path = _uploads_dir() / name
    if not path.exists():
        return RedirectResponse("/admin", status_code=303)
    low = name.lower()
    if low.endswith(".pdf"):
        # PDF: inline ให้แสดงใน iframe/แท็บ ไม่บังคับดาวน์โหลด
        return FileResponse(str(path), media_type=_PDF, content_disposition_type="inline")
    img_mt = {".png": "image/png", ".jpg": "image/jpeg", ".webp": "image/webp"}
    for ext, mt in img_mt.items():
        if low.endswith(ext):                # รูป: inline ใช้พรีวิวเกียรติบัตร
            return FileResponse(str(path), media_type=mt, content_disposition_type="inline")
    # Word: ให้เปิด/ดาวน์โหลด (เบราว์เซอร์เปิด docx ใน iframe ไม่ได้)
    return FileResponse(str(path), media_type=_DOCX, filename=name)


@router.post("/admin/outgoing/{rid}/delete")
def outgoing_delete(rid: int, db: Session = Depends(get_db)):
    r = db.get(OutgoingLetter, rid)
    if r:
        db.delete(r); db.commit()
    return RedirectResponse("/admin/outgoing", status_code=303)


# ---------------- บันทึกข้อความ ----------------
@router.get("/admin/memos", response_class=HTMLResponse)
def memos_page(request: Request, db: Session = Depends(get_db)):
    fy = current_fiscal_year()
    rows = db.query(OfficeMemo).order_by(OfficeMemo.id.desc()).all()
    school = get_school(db)
    return templates.TemplateResponse("memos.html", {
        "request": request, "rows": rows, "fiscal_year": fy, "school": school,
        "sug_memo": suggest_doc_no(db, "memo", fy),
        "persons": db.query(Person).order_by(Person.name).all(),
        "departments": db.query(Department).order_by(Department.name).all(),
    })


@router.post("/admin/memos")
def memo_create(db: Session = Depends(get_db), fiscal_year: str = Form(""), memo_no: str = Form(""),
                date: str = Form(""), from_dept: str = Form(""), to_person: str = Form(""),
                subject: str = Form(""), body: str = Form(""),
                signer_name: str = Form(""), signer_position: str = Form(""),
                pending_file: str = Form("")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    no = memo_no.strip() or suggest_doc_no(db, "memo", fy)
    from app.services.doc_number import parse_seq
    fp = pending_file.strip() if _SAFE_FILE_NAME.match(pending_file.strip()) else ""
    m = OfficeMemo(fiscal_year=fy, memo_no=no, seq=parse_seq(no), date=parse_be_date(date),
                   from_dept=from_dept.strip(), to_person=to_person.strip(),
                   subject=subject.strip(), body=body, signer_name=signer_name.strip(),
                   signer_position=signer_position.strip(), file_path=fp)
    db.add(m); db.flush()
    commit_doc_no(db, "memo", fy, no, source="admin", ref_id=m.id, subject=m.subject)
    db.commit(); db.refresh(m)
    return RedirectResponse(f"/admin/memos/{m.id}", status_code=303)


@router.post("/admin/memos/upload")
async def memo_upload(request: Request, db: Session = Depends(get_db), file: UploadFile = File(...)):
    data = await file.read()
    ext = _detect_ext(data, file.filename or "")
    if not ext:
        return RedirectResponse("/admin/memos?err=notpdf", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "memo", name, fields)


@router.post("/admin/memos/fetch")
async def memo_fetch(request: Request, db: Session = Depends(get_db), url: str = Form("")):
    data, ext, err = _fetch_file(url)
    if err:
        return RedirectResponse(f"/admin/memos?err={err}", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "memo", name, fields)


@router.get("/admin/memos/{mid}", response_class=HTMLResponse)
def memo_detail(mid: int, request: Request, db: Session = Depends(get_db)):
    m = db.get(OfficeMemo, mid)
    if not m:
        return RedirectResponse("/admin/memos", status_code=303)
    return templates.TemplateResponse("memo_detail.html", {
        "request": request, "m": m, "school": get_school(db),
        "persons": db.query(Person).order_by(Person.name).all(),
        "departments": db.query(Department).order_by(Department.name).all(),
    })


@router.post("/admin/memos/{mid}/update")
def memo_update(mid: int, db: Session = Depends(get_db), memo_no: str = Form(""),
                date: str = Form(""), from_dept: str = Form(""), to_person: str = Form(""),
                subject: str = Form(""), body: str = Form(""),
                signer_name: str = Form(""), signer_position: str = Form("")):
    m = db.get(OfficeMemo, mid)
    if m:
        from app.services.doc_number import parse_seq
        m.memo_no = memo_no.strip(); m.seq = parse_seq(memo_no)
        m.date = parse_be_date(date); m.from_dept = from_dept.strip()
        m.to_person = to_person.strip(); m.subject = subject.strip(); m.body = body
        m.signer_name = signer_name.strip(); m.signer_position = signer_position.strip()
        commit_doc_no(db, "memo", m.fiscal_year, m.memo_no, source="admin", ref_id=m.id, subject=m.subject)
        db.commit()
    return RedirectResponse(f"/admin/memos/{mid}?saved=1", status_code=303)


@router.post("/admin/memos/{mid}/delete")
def memo_delete(mid: int, db: Session = Depends(get_db)):
    m = db.get(OfficeMemo, mid)
    if m:
        db.delete(m); db.commit()
    return RedirectResponse("/admin/memos", status_code=303)


@router.get("/admin/memos/{mid}/generate")
def memo_generate(mid: int, db: Session = Depends(get_db)):
    m = db.get(OfficeMemo, mid)
    if not m:
        return RedirectResponse("/admin/memos", status_code=303)
    path = render_memo(m, get_school(db))
    return serve_generated(path, _DOCX)


# ---------------- คำสั่งโรงเรียน ----------------
@router.get("/admin/orders", response_class=HTMLResponse)
def orders_page(request: Request, db: Session = Depends(get_db)):
    fy = current_fiscal_year()
    rows = db.query(SchoolOrder).order_by(SchoolOrder.id.desc()).all()
    return templates.TemplateResponse("orders.html", {
        "request": request, "rows": rows, "fiscal_year": fy, "school": get_school(db),
        "sug_order": suggest_doc_no(db, "command", fy),
    })


@router.post("/admin/orders")
def order_create(db: Session = Depends(get_db), fiscal_year: str = Form(""), order_no: str = Form(""),
                 date: str = Form(""), subject: str = Form(""), body: str = Form(""),
                 pending_file: str = Form("")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    no = order_no.strip() or suggest_doc_no(db, "command", fy)
    from app.services.doc_number import parse_seq
    fp = pending_file.strip() if _SAFE_FILE_NAME.match(pending_file.strip()) else ""
    o = SchoolOrder(fiscal_year=fy, order_no=no, seq=parse_seq(no), date=parse_be_date(date),
                    subject=subject.strip(), body=body, file_path=fp)
    db.add(o); db.flush()
    commit_doc_no(db, "command", fy, no, source="admin", ref_id=o.id, subject=o.subject)
    db.commit(); db.refresh(o)
    return RedirectResponse(f"/admin/orders/{o.id}", status_code=303)


@router.post("/admin/orders/upload")
async def order_upload(request: Request, db: Session = Depends(get_db), file: UploadFile = File(...)):
    data = await file.read()
    ext = _detect_ext(data, file.filename or "")
    if not ext:
        return RedirectResponse("/admin/orders?err=notpdf", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "order", name, fields)


@router.post("/admin/orders/fetch")
async def order_fetch(request: Request, db: Session = Depends(get_db), url: str = Form("")):
    data, ext, err = _fetch_file(url)
    if err:
        return RedirectResponse(f"/admin/orders?err={err}", status_code=303)
    name = _save_upload(data, ext)
    fields = extract_letter_fields(str(_uploads_dir() / name))
    return _render_letter_confirm(request, db, "order", name, fields)


@router.get("/admin/orders/{oid}", response_class=HTMLResponse)
def order_detail(oid: int, request: Request, db: Session = Depends(get_db)):
    o = db.get(SchoolOrder, oid)
    if not o:
        return RedirectResponse("/admin/orders", status_code=303)
    return templates.TemplateResponse("order_detail.html", {
        "request": request, "o": o, "school": get_school(db),
    })


@router.post("/admin/orders/{oid}/update")
def order_update(oid: int, db: Session = Depends(get_db), order_no: str = Form(""),
                 date: str = Form(""), subject: str = Form(""), body: str = Form("")):
    o = db.get(SchoolOrder, oid)
    if o:
        from app.services.doc_number import parse_seq
        o.order_no = order_no.strip(); o.seq = parse_seq(order_no)
        o.date = parse_be_date(date); o.subject = subject.strip(); o.body = body
        commit_doc_no(db, "command", o.fiscal_year, o.order_no, source="admin", ref_id=o.id, subject=o.subject)
        db.commit()
    return RedirectResponse(f"/admin/orders/{oid}?saved=1", status_code=303)


@router.post("/admin/orders/{oid}/delete")
def order_delete(oid: int, db: Session = Depends(get_db)):
    o = db.get(SchoolOrder, oid)
    if o:
        db.delete(o); db.commit()
    return RedirectResponse("/admin/orders", status_code=303)


@router.get("/admin/orders/{oid}/generate")
def order_generate(oid: int, db: Session = Depends(get_db)):
    o = db.get(SchoolOrder, oid)
    if not o:
        return RedirectResponse("/admin/orders", status_code=303)
    path = render_order(o, get_school(db))
    return serve_generated(path, _DOCX)


# ---------------- นำเข้า Excel (ทะเบียนหนังสือรับ/ส่ง) ----------------
@router.get("/admin/import", response_class=HTMLResponse)
def admin_import_page(request: Request, db: Session = Depends(get_db),
                      imported: str | None = None, import_err: str | None = None):
    lines = []
    if imported and imported != "none":
        for part in imported.split(","):
            if ":" in part:
                sheet, n = part.split(":", 1)
                lines.append(f"{sheet}: เพิ่ม {n} รายการ")
    err = {"type": "ไฟล์ต้องเป็น .xlsx เท่านั้น",
           "read": "อ่านไฟล์ไม่สำเร็จ ตรวจสอบว่าใช้เทมเพลตที่ถูกต้อง"}.get(import_err)
    return templates.TemplateResponse("admin_import.html", {
        "request": request, "import_lines": lines, "import_err": err,
    })


@router.get("/admin/letters/export.xlsx")
def admin_letters_export(db: Session = Depends(get_db), year: int | None = None):
    fy = year or current_fiscal_year()
    incoming = (db.query(IncomingLetter).filter_by(fiscal_year=fy)
                .order_by(IncomingLetter.recv_no).all())
    outgoing = (db.query(OutgoingLetter).filter_by(fiscal_year=fy)
                .order_by(OutgoingLetter.send_seq).all())
    path = export_admin_register(incoming, outgoing, fy)
    return serve_generated(path, _XLSX)


@router.get("/admin/template.xlsx")
def admin_template():
    path = build_admin_template()
    return serve_generated(path, _XLSX)


@router.post("/admin/import")
async def admin_import(db: Session = Depends(get_db), file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/admin/import?import_err=type", status_code=303)
    content = await file.read()
    try:
        summary = import_admin_workbook(content, db)
    except Exception:
        return RedirectResponse("/admin/import?import_err=read", status_code=303)
    q = ",".join(f"{k}:{v}" for k, v in summary.items()) or "none"
    return RedirectResponse(f"/admin/import?imported={q}", status_code=303)


# ============================================================
# หนังสือราชการภายนอก (แม่แบบสำเร็จ + ออก Word)
# ============================================================
# แม่แบบสำเร็จ: key -> (label, เรื่อง, โครงร่างเนื้อความ)  ("…" = เว้นให้เติม)
LETTER_PRESETS = {
    "blank": ("ฟอร์มเปล่า", "", ""),
    "invite_meeting": ("เชิญประชุม", "ขอเชิญประชุม………………………",
        "ด้วย…………………………… กำหนดจัดประชุม…………………… ในวันที่…………… เวลา………… น. "
        "ณ…………………………………\n\nในการนี้ จึงขอเรียนเชิญท่านเข้าร่วมประชุมตามวัน เวลา และสถานที่ดังกล่าว "
        "จะเป็นพระคุณยิ่ง"),
    "invite_speaker": ("เชิญเป็นวิทยากร", "ขอเชิญเป็นวิทยากร",
        "ด้วย…………………………… กำหนดจัด…………………… ในวันที่…………… ณ……………………………\n\n"
        "ในการนี้ เห็นว่าท่านเป็นผู้มีความรู้ความสามารถ จึงขอเรียนเชิญท่านเป็นวิทยากรในหัวข้อ"
        "…………………………… จะเป็นพระคุณยิ่ง"),
    "ask_permission": ("ขออนุญาต", "ขออนุญาต……………………",
        "ด้วย…………………………… มีความประสงค์ขออนุญาต…………………… เนื่องด้วย……………………………\n\n"
        "จึงเรียนมาเพื่อโปรดพิจารณาอนุญาต"),
    "ask_support": ("ขอความอนุเคราะห์", "ขอความอนุเคราะห์……………………",
        "ด้วย…………………………… จึงใคร่ขอความอนุเคราะห์…………………… จากท่าน เพื่อ……………………………\n\n"
        "จึงเรียนมาเพื่อโปรดพิจารณาให้ความอนุเคราะห์"),
    "thanks": ("ขอบคุณ", "ขอขอบคุณ……………………",
        "ตามที่ท่านได้……………………………… นั้น\n\n…………………………… ขอขอบคุณในความอนุเคราะห์ของท่านมา ณ "
        "โอกาสนี้ และหวังเป็นอย่างยิ่งว่าจะได้รับความอนุเคราะห์จากท่านในโอกาสต่อไป"),
    "send_person": ("ส่งตัว", "ส่งตัว……………………",
        "ตามที่…………………………… นั้น\n\n…………………………… ขอส่งตัว…………………… เพื่อ……………………………"),
}


def _letter_ctx(db, extra=None):
    school = get_school(db)
    fy = current_fiscal_year()
    ctx = {
        "school": school, "fiscal_year": fy,
        "sug_no": suggest_doc_no(db, "outgoing", fy),
        "presets": LETTER_PRESETS,
        "persons": db.query(Person).order_by(Person.name).all(),
        "today_input": be_date_input(datetime.now()),
    }
    if extra:
        ctx.update(extra)
    return ctx


@router.get("/admin/letters", response_class=HTMLResponse)
def letters_page(request: Request, db: Session = Depends(get_db)):
    rows = db.query(OfficialLetter).order_by(OfficialLetter.id.desc()).all()
    return templates.TemplateResponse("official_letters.html",
                                      _letter_ctx(db, {"request": request, "rows": rows}))


@router.post("/admin/letters")
def letter_create(db: Session = Depends(get_db), fiscal_year: str = Form(""), doc_no: str = Form(""),
                  date: str = Form(""), subject: str = Form(""), to: str = Form(""),
                  ref: str = Form(""), enclosure: str = Form(""), body: str = Form(""),
                  closing: str = Form("ขอแสดงความนับถือ"), signer_name: str = Form(""),
                  signer_position: str = Form(""), preset: str = Form("")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    no = doc_no.strip() or suggest_doc_no(db, "outgoing", fy)
    lt = OfficialLetter(fiscal_year=fy, doc_no=no, seq=parse_seq(no), date=parse_be_date(date),
                        subject=subject.strip(), to=to.strip(), ref=ref.strip(),
                        enclosure=enclosure.strip(), body=body,
                        closing=closing.strip() or "ขอแสดงความนับถือ",
                        signer_name=signer_name.strip(), signer_position=signer_position.strip(),
                        preset=preset.strip())
    db.add(lt); db.flush()
    commit_doc_no(db, "outgoing", fy, no, source="admin", ref_id=lt.id, subject=lt.subject)
    db.commit(); db.refresh(lt)
    return RedirectResponse(f"/admin/letters/{lt.id}", status_code=303)


@router.post("/admin/letters/ai-write")
def letter_ai_write(db: Session = Depends(get_db), subject: str = Form(""),
                    to: str = Form(""), points: str = Form(""), detail: str = Form("")):
    """ให้ AI ร่างเนื้อความหนังสือราชการจากข้อมูลสำคัญ -> คืน JSON {subject, body}"""
    from fastapi.responses import JSONResponse
    from app.services.ai_extract import write_official_letter
    school = get_school(db)
    key = _ai_key()
    if not key:
        return JSONResponse({"error": "ฟีเจอร์ AI สำหรับสมาชิกเท่านั้น — ต่ออายุ/เป็นสมาชิกเพื่อใช้งาน"}, status_code=400)
    res = write_official_letter({
        "school": school.name or "", "subject": subject, "to": to,
        "points": points, "detail": detail,
    }, key)
    if res.get("error"):
        return JSONResponse({"error": "AI เขียนไม่สำเร็จ ลองใหม่อีกครั้ง"}, status_code=502)
    return JSONResponse({"subject": res.get("subject", ""), "body": res.get("body", "")})


@router.post("/admin/memos/ai-write")
def memo_ai_write(db: Session = Depends(get_db), subject: str = Form(""),
                  from_dept: str = Form(""), to: str = Form(""), points: str = Form("")):
    """ให้ AI ร่างบันทึกข้อความ -> คืน JSON {subject, body}"""
    from fastapi.responses import JSONResponse
    from app.services.ai_extract import write_memo
    school = get_school(db)
    key = _ai_key()
    if not key:
        return JSONResponse({"error": "ฟีเจอร์ AI สำหรับสมาชิกเท่านั้น — ต่ออายุ/เป็นสมาชิกเพื่อใช้งาน"}, status_code=400)
    res = write_memo({"school": school.name or "", "subject": subject,
                      "from_dept": from_dept, "to": to, "points": points}, key)
    if res.get("error"):
        return JSONResponse({"error": "AI เขียนไม่สำเร็จ ลองใหม่อีกครั้ง"}, status_code=502)
    return JSONResponse({"subject": res.get("subject", ""), "body": res.get("body", "")})


@router.post("/admin/orders/ai-write")
def order_ai_write(db: Session = Depends(get_db), subject: str = Form(""), points: str = Form("")):
    """ให้ AI ร่างคำสั่งโรงเรียน -> คืน JSON {subject, body}"""
    from fastapi.responses import JSONResponse
    from app.services.ai_extract import write_order
    school = get_school(db)
    key = _ai_key()
    if not key:
        return JSONResponse({"error": "ฟีเจอร์ AI สำหรับสมาชิกเท่านั้น — ต่ออายุ/เป็นสมาชิกเพื่อใช้งาน"}, status_code=400)
    res = write_order({"school": school.name or "", "subject": subject, "points": points}, key)
    if res.get("error"):
        return JSONResponse({"error": "AI เขียนไม่สำเร็จ ลองใหม่อีกครั้ง"}, status_code=502)
    return JSONResponse({"subject": res.get("subject", ""), "body": res.get("body", "")})


@router.get("/admin/letters/{lid}", response_class=HTMLResponse)
def letter_detail(lid: int, request: Request, db: Session = Depends(get_db)):
    lt = db.get(OfficialLetter, lid)
    if not lt:
        return RedirectResponse("/admin/letters", status_code=303)
    return templates.TemplateResponse("official_letter_detail.html",
                                      _letter_ctx(db, {"request": request, "lt": lt}))


@router.post("/admin/letters/{lid}/update")
def letter_update(lid: int, db: Session = Depends(get_db), doc_no: str = Form(""),
                  date: str = Form(""), subject: str = Form(""), to: str = Form(""),
                  ref: str = Form(""), enclosure: str = Form(""), body: str = Form(""),
                  closing: str = Form("ขอแสดงความนับถือ"), signer_name: str = Form(""),
                  signer_position: str = Form("")):
    lt = db.get(OfficialLetter, lid)
    if lt:
        lt.doc_no = doc_no.strip(); lt.seq = parse_seq(doc_no)
        lt.date = parse_be_date(date); lt.subject = subject.strip(); lt.to = to.strip()
        lt.ref = ref.strip(); lt.enclosure = enclosure.strip(); lt.body = body
        lt.closing = closing.strip() or "ขอแสดงความนับถือ"
        lt.signer_name = signer_name.strip(); lt.signer_position = signer_position.strip()
        commit_doc_no(db, "outgoing", lt.fiscal_year, lt.doc_no, source="admin",
                      ref_id=lt.id, subject=lt.subject)
        db.commit()
    return RedirectResponse(f"/admin/letters/{lid}?saved=1", status_code=303)


@router.post("/admin/letters/{lid}/delete")
def letter_delete(lid: int, db: Session = Depends(get_db)):
    lt = db.get(OfficialLetter, lid)
    if lt:
        db.delete(lt); db.commit()
    return RedirectResponse("/admin/letters", status_code=303)


@router.get("/admin/letters/{lid}/generate")
def letter_generate(lid: int, db: Session = Depends(get_db)):
    lt = db.get(OfficialLetter, lid)
    if not lt:
        return RedirectResponse("/admin/letters", status_code=303)
    path = render_official_letter(lt, get_school(db))
    return serve_generated(path, _DOCX)


# ============================================================
# เกียรติบัตร (อัปโหลดพื้นหลัง + พิมพ์ชื่อทับ -> PDF หลายหน้า)
# ============================================================
from fastapi.responses import JSONResponse


@router.get("/admin/certificates", response_class=HTMLResponse)
def certificates_page(request: Request, db: Session = Depends(get_db)):
    batches = db.query(CertificateBatch).order_by(CertificateBatch.id.desc()).all()
    saved = [{
        "id": b.id, "title": b.title or f"ชุดที่ {b.id}", "bg_image": b.bg_image or "",
        "url": f"/admin/uploaded/{b.bg_image}" if b.bg_image else "",
        "name_x": b.name_x, "name_y": b.name_y, "name_size": b.name_size,
        "name_color": b.name_color or "#1a1a1a", "sub_text": b.sub_text or "",
    } for b in batches if b.bg_image]
    return templates.TemplateResponse("certificates.html", {
        "request": request, "school": get_school(db), "saved": saved,
    })


@router.post("/admin/certificates/{bid}/delete")
def certificate_delete(bid: int, db: Session = Depends(get_db)):
    """ลบชุดเกียรติบัตรที่บันทึกไว้"""
    b = db.get(CertificateBatch, bid)
    if b:
        db.delete(b); db.commit()
    return RedirectResponse("/admin/certificates", status_code=303)


@router.post("/admin/certificates/bg")
async def certificate_bg(file: UploadFile = File(...)):
    """อัปโหลดรูปพื้นหลังเกียรติบัตร -> คืนชื่อไฟล์ + ขนาด (สำหรับพรีวิว/ปักตำแหน่ง)"""
    data = await file.read()
    ext = file_upload.detect_ext(data, file.filename or "")
    if ext not in ("png", "jpg", "webp"):
        return JSONResponse({"error": "ไฟล์ต้องเป็นรูปภาพ (png/jpg/webp)"})
    name = file_upload.save_upload(data, ext)
    try:
        from PIL import Image
        import io
        w, h = Image.open(io.BytesIO(data)).size
    except Exception:
        w = h = 0
    return JSONResponse({"file": name, "w": w, "h": h, "url": f"/admin/uploaded/{name}"})


@router.post("/admin/certificates/names-excel")
async def certificate_names_excel(file: UploadFile = File(...)):
    """อ่านรายชื่อจาก Excel (คอลัมน์แรก) -> คืนเป็นรายการ"""
    if not (file.filename or "").lower().endswith(".xlsx"):
        return JSONResponse({"error": "ไฟล์ต้องเป็น .xlsx"})
    data = await file.read()
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        names = []
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None:
                v = str(row[0]).strip()
                if v and v not in ("ชื่อ", "ชื่อ-นามสกุล", "รายชื่อ", "ลำดับ"):
                    names.append(v)
    except Exception:
        return JSONResponse({"error": "อ่านไฟล์ Excel ไม่สำเร็จ"})
    return JSONResponse({"names": names})


@router.post("/admin/certificates")
def certificate_generate(db: Session = Depends(get_db), title: str = Form(""),
                         sub_text: str = Form(""), bg_image: str = Form(""),
                         name_x: str = Form("50"), name_y: str = Form("45"),
                         name_size: str = Form("48"), name_color: str = Form("#1a1a1a"),
                         names: str = Form("")):
    from app.services.cert_doc import render_certificates
    name_list = [n.strip() for n in (names or "").splitlines() if n.strip()]
    if not bg_image.strip() or not name_list:
        return RedirectResponse("/admin/certificates?err=1", status_code=303)
    batch = CertificateBatch(
        title=title.strip(), sub_text=sub_text.strip(), bg_image=bg_image.strip(),
        name_x=_to_float(name_x, 50.0), name_y=_to_float(name_y, 45.0),
        name_size=_to_int(name_size, 48), name_color=name_color.strip() or "#1a1a1a",
        names="\n".join(name_list))
    db.add(batch); db.commit(); db.refresh(batch)
    path = render_certificates(batch, name_list, get_school(db))
    return serve_generated(path, _PDF)
