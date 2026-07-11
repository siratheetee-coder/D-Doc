"""
pages.py
--------
เส้นทาง (URL) ของทุกหน้าในระบบ (เฟส 2)

หน้าหลัก:
  /                      Dashboard + ทะเบียนจัดซื้อ
  /settings              ตั้งค่าโรงเรียน + เกณฑ์วงเงิน + ผู้ลงนาม
  /masters               จัดการรายชื่อบุคลากร / ฝ่าย / โครงการ
  /vendors               ผู้ขาย/ผู้รับจ้าง
  /procurement/new       สร้างเรื่องจัดซื้อใหม่ (เลือกผู้ตรวจรับคนเดียว/คณะกรรมการ)
  /procurement/{id}      รายละเอียด + ออกเอกสารจากแม่แบบ
  /register.xlsx         ดาวน์โหลดทะเบียน Excel
"""
from pathlib import Path
from datetime import datetime
from types import SimpleNamespace

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File, Response, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from app.database import get_db
from app.models import (
    School, Vendor, Procurement, ProcurementItem, Document,
    Person, Department, Project, ProjectBudgetRevision, Committee, CommitteeMember,
    Asset, MaterialItem, MaterialTxn, Requisition, RequisitionItem, IssuedDocNo,
    DocNumberCounter, OfficeMemo, SchoolOrder, IncomingLetter, OutgoingLetter,
    DisburseMemo, ItemCatalog, Student, ProcurementPlan, Contract,
)
from app.services.asset_utils import (
    CATEGORIES, CATEGORY_LIFE, annual_depreciation, accumulated_depreciation,
    net_book_value, depreciation_schedule, material_balance,
)
from app.services.doc_number import suggest_doc_no, commit_doc_no, check_doc_no, COUNTER_TYPES, parse_seq
from app.services.budget import current_plan_year, plan_year_label, project_budget, project_spent
from app.services.render import render_document, render_bundle, AVAILABLE_KINDS
from app.services.register_export import export_register
from app.services.thai_holidays import holiday_map, year_range_for
from app.services.bulk_io import build_import_template, import_workbook
from app.services import file_upload
from app.services.pdf_extract import extract_letter_fields, extract_procurement_fields, extract_text_any
from app.services.ai_extract import extract_with_ai
from app.thai_utils import current_fiscal_year, thai_date, bahttext, be_date_input, parse_be_date
from app.templating import templates

router = APIRouter()


def _to_int(v, default=0):
    """แปลงเป็นจำนวนเต็มแบบปลอดภัย (คืน default ถ้าไม่ใช่ตัวเลข)"""
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return default


def _to_float(v, default=0.0):
    """แปลงเป็นทศนิยมแบบปลอดภัย (คืน default ถ้าไม่ใช่ตัวเลข)"""
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return default


def _order_sort_key(p):
    """เรียงตามเลขใบสั่งซื้อ/จ้าง (ลำดับตัวเลข น้อย->มาก) เรื่องที่ยังไม่ออกใบสั่งไปท้าย"""
    return (0, parse_seq(p.order_no)) if (p.order_no or "").strip() else (1, p.id)


def get_school(db: Session) -> School:
    school = db.query(School).first()
    if school is None:
        school = School()
        db.add(school)
        db.commit()
        db.refresh(school)
    return school


# ลำดับเดือนตามปีงบประมาณไทย (ต.ค. -> ก.ย.) + ป้ายเดือน
_FY_MONTHS = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
_FY_MONTH_LABELS = ["ต.ค.", "พ.ย.", "ธ.ค.", "ม.ค.", "ก.พ.", "มี.ค.",
                    "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย."]
# สถานะเอกสาร + สีโดนัท (เรียงตามขั้นตอนงาน)
_STATUS_COLORS = [("ร่าง", "#9aa6b8"), ("อนุมัติ", "#16b364"),
                  ("ตรวจรับแล้ว", "#4880FF"), ("เบิกจ่ายแล้ว", "#3568e0")]


def _dashboard_charts(procurements):
    """คำนวณข้อมูลกราฟแดชบอร์ด: ยอดจัดซื้อรายเดือน (แท่ง) + สถานะเอกสาร (โดนัท)
    คืน dict สำหรับวาด SVG ในเทมเพลต (ออฟไลน์ ไม่ใช้ไลบรารีภายนอก)"""
    # --- รายเดือน ---
    sums = {m: 0.0 for m in _FY_MONTHS}
    for p in procurements:
        if p.order_date:                     # อิงวันที่ใบสั่งซื้อ/จ้าง (เรื่องที่ยังไม่ออกใบสั่ง = ยังไม่นับเป็นยอดซื้อ)
            sums[p.order_date.month] = sums.get(p.order_date.month, 0) + (p.total_amount or 0)
    maxv = max(sums.values()) or 1
    bars = [{"label": lab, "value": sums[m], "frac": sums[m] / maxv}
            for m, lab in zip(_FY_MONTHS, _FY_MONTH_LABELS)]
    monthly_max = max(sums.values())

    # --- สถานะ (โดนัท) ---
    counts = {s: 0 for s, _ in _STATUS_COLORS}
    for p in procurements:
        st = p.status if p.status in counts else "ร่าง"
        counts[st] += 1
    total = sum(counts.values())
    circ = 2 * 3.141592653589793 * 70.0       # รัศมีวงโดนัท r=70
    acc = 0.0
    segs = []
    for s, color in _STATUS_COLORS:
        frac = (counts[s] / total) if total else 0
        segs.append({
            "label": s, "color": color, "count": counts[s],
            "pct": round(frac * 100),
            "dash": frac * circ, "gap": circ - frac * circ,
            "offset": -acc * circ,
        })
        acc += frac
    return {
        "bars": bars, "monthly_max": monthly_max,
        "segs": segs, "status_total": total, "circ": circ,
    }


# ---------------- หน้าเลือกงาน (Hub) ----------------
@router.get("/", response_class=HTMLResponse)
def hub(request: Request, db: Session = Depends(get_db)):
    """หน้าหลัก: เลือกเข้าใช้งาน ธุรการ / พัสดุ / การเงิน + ภาพรวมปีงบ"""
    fy = current_fiscal_year()
    stats = {
        "proc": db.query(Procurement).filter(Procurement.fiscal_year == fy).count(),
        "proc_amount": db.query(func.coalesce(func.sum(Procurement.total_amount), 0.0))
                         .filter(Procurement.fiscal_year == fy).scalar() or 0,
        "memo": db.query(OfficeMemo).filter(OfficeMemo.fiscal_year == fy).count(),
        "order": db.query(SchoolOrder).filter(SchoolOrder.fiscal_year == fy).count(),
        "incoming": db.query(IncomingLetter).filter(IncomingLetter.fiscal_year == fy).count(),
        "outgoing": db.query(OutgoingLetter).filter(OutgoingLetter.fiscal_year == fy).count(),
        "asset": db.query(Asset).count(),
        "material": db.query(MaterialItem).count(),
    }
    admin_total = stats["memo"] + stats["order"] + stats["incoming"] + stats["outgoing"]
    return templates.TemplateResponse("hub.html", {
        "request": request, "school": get_school(db), "fiscal_year": fy,
        "stats": stats, "admin_total": admin_total,
    })


# ---------------- ค้นหากลาง (ข้ามทุกงาน) ----------------
@router.get("/search", response_class=HTMLResponse)
def global_search(request: Request, db: Session = Depends(get_db), q: str | None = None):
    """ค้นหารวมทุกงาน: พัสดุ / บันทึก / คำสั่ง / หนังสือรับ-ส่ง / ครุภัณฑ์ / วัสดุ"""
    q = (q or "").strip()
    groups = []
    if q:
        like = f"%{q}%"
        # งานพัสดุ
        procs = (db.query(Procurement).outerjoin(Procurement.vendor)
                 .filter(or_(Procurement.subject.ilike(like), Procurement.memo_no.ilike(like),
                             Procurement.order_no.ilike(like), Vendor.name.ilike(like)))
                 .order_by(Procurement.id.desc()).limit(20).all())
        groups.append(("งานพัสดุ: เรื่องจัดซื้อจัดจ้าง", "procurement", [{
            "title": f"{p.order_no or p.memo_no or '-'} · {p.proc_type or ''}{p.subject or ''}",
            "sub": (p.vendor.name if p.vendor else "") + " · " + (p.status or ""),
            "url": f"/procurement/{p.id}"} for p in procs]))
        # บันทึกข้อความ
        memos = (db.query(OfficeMemo).filter(or_(OfficeMemo.subject.ilike(like),
                 OfficeMemo.memo_no.ilike(like))).order_by(OfficeMemo.id.desc()).limit(20).all())
        groups.append(("งานธุรการ: บันทึกข้อความ", "admin", [{
            "title": f"{m.memo_no or '-'} · {m.subject or ''}", "sub": m.to_person or "",
            "url": f"/admin/memos/{m.id}"} for m in memos]))
        # คำสั่ง
        orders = (db.query(SchoolOrder).filter(or_(SchoolOrder.subject.ilike(like),
                  SchoolOrder.order_no.ilike(like))).order_by(SchoolOrder.id.desc()).limit(20).all())
        groups.append(("งานธุรการ: คำสั่งโรงเรียน", "admin", [{
            "title": f"{o.order_no or '-'} · {o.subject or ''}", "sub": "",
            "url": f"/admin/orders/{o.id}"} for o in orders]))
        # หนังสือรับ
        inc = (db.query(IncomingLetter).filter(or_(IncomingLetter.subject.ilike(like),
               IncomingLetter.letter_no.ilike(like), IncomingLetter.from_org.ilike(like)))
               .order_by(IncomingLetter.id.desc()).limit(20).all())
        groups.append(("งานธุรการ: หนังสือรับ", "admin", [{
            "title": f"รับ {r.recv_no} · {r.subject or ''}", "sub": (r.letter_no or "") + " · " + (r.from_org or ""),
            "url": "/admin/incoming"} for r in inc]))
        # หนังสือส่ง
        out = (db.query(OutgoingLetter).filter(or_(OutgoingLetter.subject.ilike(like),
               OutgoingLetter.send_no.ilike(like), OutgoingLetter.to_org.ilike(like)))
               .order_by(OutgoingLetter.id.desc()).limit(20).all())
        groups.append(("งานธุรการ: หนังสือส่ง", "admin", [{
            "title": f"{r.send_no or '-'} · {r.subject or ''}", "sub": r.to_org or "",
            "url": "/admin/outgoing"} for r in out]))
        # ครุภัณฑ์
        assets = (db.query(Asset).filter(or_(Asset.name.ilike(like), Asset.asset_code.ilike(like)))
                  .order_by(Asset.id.desc()).limit(20).all())
        groups.append(("งานพัสดุ: ครุภัณฑ์", "procurement", [{
            "title": f"{a.asset_code or '-'} · {a.name}", "sub": a.category or "",
            "url": "/assets"} for a in assets]))
        # วัสดุ
        mats = (db.query(MaterialItem).filter(MaterialItem.name.ilike(like))
                .order_by(MaterialItem.name).limit(20).all())
        groups.append(("งานพัสดุ: วัสดุ", "procurement", [{
            "title": it.name, "sub": "คงเหลือ " + format(material_balance(it), "g") + " " + (it.unit or ""),
            "url": f"/materials/{it.id}"} for it in mats]))
        # การเงิน: บันทึกขอเบิกจ่าย
        disb = (db.query(DisburseMemo).filter(or_(DisburseMemo.subject.ilike(like),
                DisburseMemo.memo_no.ilike(like), DisburseMemo.payee.ilike(like)))
                .order_by(DisburseMemo.id.desc()).limit(20).all())
        groups.append(("งานการเงิน: บันทึกขอเบิกจ่าย", "finance", [{
            "title": f"{d.memo_no or '-'} · {d.subject or ''}",
            "sub": (d.payee or "") + " · " + "{:,.2f}".format(d.amount or 0) + " บาท",
            "url": f"/finance/disburse/{d.id}"} for d in disb]))
    total = sum(len(g[2]) for g in groups)
    return templates.TemplateResponse("search_results.html", {
        "request": request, "q": q, "groups": groups, "total": total,
    })


# ---------------- ทะเบียนเลขหนังสือกลาง (ใช้ร่วมทุกงาน) ----------------
@router.get("/docnos", response_class=HTMLResponse)
def docnos_page(request: Request, db: Session = Depends(get_db),
                year: int | None = None):
    fy = year or current_fiscal_year()
    rows = (db.query(IssuedDocNo)
            .filter(IssuedDocNo.fiscal_year == fy)
            .order_by(IssuedDocNo.doc_type, IssuedDocNo.seq).all())
    years = [r[0] for r in db.query(IssuedDocNo.fiscal_year).distinct().all()]
    if fy not in years:
        years.append(fy)
    years = sorted(set(years), reverse=True)
    # จัดกลุ่มตามชนิดเอกสาร
    groups = {}
    for r in rows:
        groups.setdefault(r.doc_type, []).append(r)
    return templates.TemplateResponse("docnos.html", {
        "request": request, "groups": groups, "type_labels": COUNTER_TYPES,
        "fiscal_year": fy, "years": years,
    })


@router.post("/docnos/{did}/update")
def docno_update(did: int, db: Session = Depends(get_db), subject: str = Form(""),
                 date: str = Form("")):
    """แก้ไขเรื่อง/วันที่ของเลขที่ในทะเบียนกลาง"""
    r = db.get(IssuedDocNo, did)
    fy = r.fiscal_year if r else current_fiscal_year()
    if r:
        r.subject = (subject or "").strip()
        d = parse_be_date(date)
        if d:
            r.date = d
        db.commit()
    return RedirectResponse(f"/docnos?year={fy}&saved=1", status_code=303)


@router.post("/docnos/{did}/delete")
def docno_delete(did: int, db: Session = Depends(get_db)):
    """ลบเลขที่ออกจากทะเบียนกลาง (ลบเฉพาะรายการทะเบียน ไม่ลบเอกสารต้นทาง)"""
    r = db.get(IssuedDocNo, did)
    fy = r.fiscal_year if r else current_fiscal_year()
    if r:
        db.delete(r); db.commit()
    return RedirectResponse(f"/docnos?year={fy}", status_code=303)


# ---------------- งานพัสดุ: Dashboard / ทะเบียน ----------------
@router.get("/procurement", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db),
              year: int | None = None, q: str | None = None):
    fy = year or current_fiscal_year()
    q = (q or "").strip()
    if q:
        # ค้นหาข้ามทุกปีงบ: เรื่อง / เลขที่บันทึก / ชื่อผู้ขาย
        like = f"%{q}%"
        procurements = (
            db.query(Procurement)
            .outerjoin(Procurement.vendor)
            .filter(or_(Procurement.subject.ilike(like),
                        Procurement.memo_no.ilike(like),
                        Vendor.name.ilike(like)))
            .order_by(Procurement.id.desc()).all()
        )
    else:
        procurements = (
            db.query(Procurement).filter(Procurement.fiscal_year == fy)
            .order_by(Procurement.id.desc()).all()
        )
    years = [r[0] for r in db.query(Procurement.fiscal_year).distinct().all()]
    if fy not in years:
        years.append(fy)
    years = sorted(set(years), reverse=True)
    total_amount = sum(p.total_amount or 0 for p in procurements)
    # กราฟแสดงเฉพาะมุมมองรายปีงบ (ไม่แสดงตอนค้นหาข้ามปี)
    charts = None if q else _dashboard_charts(procurements)
    # แยกทะเบียนคุม: จัดซื้อ (ซื้อ) / จัดจ้าง (จ้าง) เป็นคนละเล่ม
    buys = sorted([p for p in procurements if (p.proc_type or "ซื้อ") == "ซื้อ"], key=_order_sort_key)
    hires = sorted([p for p in procurements if (p.proc_type or "ซื้อ") == "จ้าง"], key=_order_sort_key)
    return templates.TemplateResponse("dashboard.html", {
        "request": request, "school": get_school(db), "procurements": procurements,
        "buys": buys, "hires": hires,
        "buy_total": sum(p.total_amount or 0 for p in buys),
        "hire_total": sum(p.total_amount or 0 for p in hires),
        "fiscal_year": fy, "years": years, "total_amount": total_amount,
        "count": len(procurements), "q": q, "charts": charts,
    })


# ---------------- ตั้งค่าโรงเรียน ----------------
@router.get("/settings", response_class=HTMLResponse)
def settings_form(request: Request, db: Session = Depends(get_db),
                  imported: str | None = None, import_err: str | None = None):
    err_msg = {"type": "ไฟล์ต้องเป็น .xlsx เท่านั้น",
               "read": "อ่านไฟล์ไม่สำเร็จ ตรวจสอบว่าใช้เทมเพลตที่ถูกต้อง"}.get(import_err)
    return templates.TemplateResponse("settings.html", {
        "request": request, "school": get_school(db),
        "persons": db.query(Person).order_by(Person.name).all(),
        "import_lines": _decode_import_summary(imported),
        "import_err": err_msg,
    })


@router.post("/settings")
def settings_save(
    db: Session = Depends(get_db),
    name: str = Form(""), address: str = Form(""),
    district: str = Form(""), province: str = Form(""),
    director_name: str = Form(""), director_position: str = Form("ผู้อำนวยการโรงเรียน"),
    officer_name: str = Form(""), head_officer_name: str = Form(""),
    finance_officer_name: str = Form(""), finance_head_name: str = Form(""),
    admin_officer_name: str = Form(""),
    doc_prefix: str = Form("ศธ"), doc_set_threshold: float = Form(5000.0),
    project_year_mode: str = Form("budget"),
):
    s = get_school(db)
    s.name, s.address, s.district, s.province = name, address, district, province
    s.director_name, s.director_position = director_name, director_position
    s.officer_name, s.head_officer_name = officer_name, head_officer_name
    s.finance_officer_name = finance_officer_name.strip()
    s.finance_head_name = finance_head_name.strip()
    s.admin_officer_name = admin_officer_name.strip()
    s.project_year_mode = "academic" if project_year_mode == "academic" else "budget"
    s.doc_prefix, s.doc_set_threshold = doc_prefix, doc_set_threshold
    db.commit()
    return RedirectResponse("/settings?saved=1", status_code=303)


# ---------------- ติดต่อเจ้าหน้าที่ / รีวิว ----------------
@router.get("/support", response_class=HTMLResponse)
def support_page(request: Request):
    from app.seller_config import SELLER
    return templates.TemplateResponse("support.html", {
        "request": request, "facebook": SELLER.get("facebook", ""),
        "sent": request.query_params.get("sent"), "rated": request.query_params.get("rated"),
    })


@router.post("/support")
def support_submit(request: Request, message: str = Form("")):
    from app.accounts import add_lead
    if message.strip():
        add_lead(kind="support", school_name=request.session.get("name", ""),
                 email=request.session.get("username", ""), tenant_id=request.session.get("tid"),
                 login_user=request.session.get("username", ""), note=message.strip(), status="ใหม่")
    return RedirectResponse("/support?sent=1", status_code=303)


@router.post("/support/review")
def support_review(request: Request, stars: str = Form("5"), comment: str = Form("")):
    from app.accounts import add_lead
    add_lead(kind="review", school_name=request.session.get("name", ""),
             email=request.session.get("username", ""), tenant_id=request.session.get("tid"),
             login_user=request.session.get("username", ""), amount=_to_float(stars, 5),
             note=comment.strip(), status="ใหม่")
    return RedirectResponse("/support?rated=1", status_code=303)


# ---------------- สำรอง / กู้คืนข้อมูล ----------------
@router.get("/backup")
def backup_download():
    """ดาวน์โหลดไฟล์สำรองฐานข้อมูลทั้งหมด (.db)"""
    from app.database import current_db_path, _checkpoint
    _checkpoint()
    fname = f"school-backup-{datetime.now():%Y%m%d-%H%M}.db"
    return serve_generated(str(current_db_path()), "application/octet-stream", count=False)


@router.post("/restore")
async def restore_upload(file: UploadFile = File(...)):
    """กู้คืนฐานข้อมูลจากไฟล์สำรอง (.db) ที่อัปโหลด"""
    from app.database import restore_db
    data = await file.read()
    if data[:16] != b"SQLite format 3\x00":
        return RedirectResponse("/settings?restore_err=type", status_code=303)
    try:
        restore_db(data)
    except Exception:
        return RedirectResponse("/settings?restore_err=fail", status_code=303)
    return RedirectResponse("/settings?restored=1", status_code=303)


# ---------------- มาสเตอร์ลิสต์ (บุคลากร/ฝ่าย/โครงการ) ----------------
def _decode_import_summary(imported: str | None) -> list:
    """แปลง query 'บุคลากร:12+2,โครงการ:5+0,โรงเรียน:u6' เป็นข้อความอ่านง่าย"""
    if not imported or imported == "none":
        return []
    lines = []
    for part in imported.split(","):
        if ":" not in part:
            continue
        sheet, val = part.split(":", 1)
        if val.startswith("u"):
            lines.append(f"{sheet}: อัปเดต {val[1:]} ช่อง")
        elif "+" in val:
            added, skipped = val.split("+", 1)
            msg = f"{sheet}: เพิ่ม {added} รายการ"
            if skipped and skipped != "0":
                msg += f" (ข้ามชื่อซ้ำ {skipped})"
            lines.append(msg)
    return lines


@router.get("/masters", response_class=HTMLResponse)
def masters_page(request: Request, db: Session = Depends(get_db),
                 imported: str | None = None, import_err: str | None = None):
    err_msg = {"type": "ไฟล์ต้องเป็น .xlsx เท่านั้น",
               "read": "อ่านไฟล์ไม่สำเร็จ ตรวจสอบว่าใช้เทมเพลตที่ถูกต้อง"}.get(import_err)
    return templates.TemplateResponse("masters.html", {
        "request": request,
        "persons": db.query(Person).order_by(Person.name).all(),
        "departments": db.query(Department).order_by(Department.name).all(),
        "projects": db.query(Project).order_by(Project.name).all(),
        "positions": POSITION_CHOICES,
        "import_lines": _decode_import_summary(imported),
        "import_err": err_msg,
    })


@router.get("/masters/template.xlsx")
def masters_template():
    """ดาวน์โหลดเทมเพลต Excel สำหรับกรอกข้อมูลตั้งต้นทีละมาก ๆ"""
    path = build_import_template()
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/masters/import")
async def masters_import(db: Session = Depends(get_db), file: UploadFile = File(...)):
    """อัปโหลดไฟล์ Excel ที่กรอกแล้ว นำเข้าข้อมูลทุกชีต"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/settings?import_err=type", status_code=303)
    content = await file.read()
    try:
        summary = import_workbook(content, db)
    except Exception:
        return RedirectResponse("/settings?import_err=read", status_code=303)
    # เข้ารหัสสรุปผลเป็น query string สั้น ๆ เช่น บุคลากร:12+2,โครงการ:5+0
    parts = []
    for sheet, r in summary.items():
        if "updated" in r:
            parts.append(f"{sheet}:u{r['updated']}")
        else:
            parts.append(f"{sheet}:{r.get('added',0)}+{r.get('skipped',0)}")
    q = ",".join(parts) if parts else "none"
    return RedirectResponse(f"/settings?imported={q}", status_code=303)


@router.post("/masters/person")
def add_person(db: Session = Depends(get_db), name: str = Form(...), position: str = Form("ครู")):
    db.add(Person(name=name, position=position)); db.commit()
    return RedirectResponse("/masters", status_code=303)


@router.post("/masters/department")
def add_department(db: Session = Depends(get_db), name: str = Form(...)):
    db.add(Department(name=name)); db.commit()
    return RedirectResponse("/masters", status_code=303)


@router.post("/masters/project")
def add_project(db: Session = Depends(get_db), name: str = Form(...),
                budget: str = Form("0"), budget_note: str = Form("")):
    db.add(Project(name=name, budget=_to_float(budget, 0.0),
                   budget_note=(budget_note or "").strip()))
    db.commit()
    return RedirectResponse("/masters", status_code=303)


@router.post("/masters/{kind}/{item_id}/update")
def update_master(kind: str, item_id: int, db: Session = Depends(get_db),
                  name: str = Form(...), position: str = Form("ครู"),
                  budget: str = Form("0"), budget_note: str = Form("")):
    model = {"person": Person, "department": Department, "project": Project}.get(kind)
    if model:
        obj = db.get(model, item_id)
        if obj:
            obj.name = name
            if kind == "person":
                obj.position = position
            elif kind == "project":
                obj.budget = _to_float(budget, 0.0)
                obj.budget_note = (budget_note or "").strip()
            db.commit()
    return RedirectResponse("/masters?saved=1", status_code=303)


@router.post("/masters/{kind}/{item_id}/delete")
def delete_master(kind: str, item_id: int, db: Session = Depends(get_db)):
    model = {"person": Person, "department": Department, "project": Project}.get(kind)
    if model:
        obj = db.get(model, item_id)
        if obj:
            if kind == "person" and getattr(obj, "signature", ""):
                from app.services.signature import delete_signature
                delete_signature(obj.signature)
            db.delete(obj); db.commit()
    return RedirectResponse("/masters", status_code=303)


# ---------------- ลายเซ็นบุคลากร ----------------
@router.post("/masters/person/{pid}/signature")
async def upload_signature(pid: int, db: Session = Depends(get_db), file: UploadFile = File(...)):
    from app.services.signature import process_signature, delete_signature
    p = db.get(Person, pid)
    if not p:
        return RedirectResponse("/masters", status_code=303)
    data = await file.read()
    if len(data) > 8 * 1024 * 1024:      # จำกัด ~8MB
        return RedirectResponse("/masters?sig=big", status_code=303)
    new_name = process_signature(data, remove_white=True)
    if not new_name:
        return RedirectResponse("/masters?sig=bad", status_code=303)
    if p.signature:                       # ลบไฟล์เก่า
        delete_signature(p.signature)
    p.signature = new_name
    db.commit()
    return RedirectResponse("/masters?sig=ok", status_code=303)


@router.post("/masters/person/{pid}/signature/delete")
def remove_signature(pid: int, db: Session = Depends(get_db)):
    from app.services.signature import delete_signature
    p = db.get(Person, pid)
    if p and p.signature:
        delete_signature(p.signature)
        p.signature = ""
        db.commit()
    return RedirectResponse("/masters", status_code=303)


@router.get("/signatures/{name}")
def serve_signature(name: str):
    from app.services.signature import signature_full_path
    path = signature_full_path(name)
    if not path:
        raise HTTPException(status_code=404)
    return FileResponse(str(path), media_type="image/png",
                        headers={"Cache-Control": "no-cache"})


# ---------------- คลังรายการพัสดุมาตรฐาน (ItemCatalog) ----------------
def _catalog_upsert(db, name: str, unit: str, price: float) -> None:
    """เก็บรายการพัสดุเข้าคลังอัตโนมัติ (dedupe ตามชื่อ, อัปเดตหน่วย/ราคาล่าสุด)"""
    name = (name or "").strip()
    if not name:
        return
    row = db.query(ItemCatalog).filter(ItemCatalog.name == name).first()
    if row:
        if unit:
            row.unit = unit
        if price:
            row.unit_price = price
    else:
        db.add(ItemCatalog(name=name, unit=unit or "ชิ้น", unit_price=price or 0))


@router.get("/catalog", response_class=HTMLResponse)
def catalog_page(request: Request, db: Session = Depends(get_db), saved: str | None = None):
    rows = db.query(ItemCatalog).order_by(ItemCatalog.category, ItemCatalog.name).all()
    return templates.TemplateResponse("catalog.html", {
        "request": request, "school": get_school(db), "rows": rows, "saved": saved,
    })


@router.post("/catalog/add")
def catalog_add(db: Session = Depends(get_db), name: str = Form(""), unit: str = Form("ชิ้น"),
                unit_price: str = Form("0"), category: str = Form("")):
    nm = (name or "").strip()
    if nm and not db.query(ItemCatalog).filter(ItemCatalog.name == nm).first():
        db.add(ItemCatalog(name=nm, unit=(unit or "ชิ้น").strip(),
                           unit_price=_to_float(unit_price, 0.0), category=(category or "").strip()))
        db.commit()
    return RedirectResponse("/catalog?saved=1", status_code=303)


@router.post("/catalog/{item_id}/update")
def catalog_update(item_id: int, db: Session = Depends(get_db), name: str = Form(""),
                   unit: str = Form("ชิ้น"), unit_price: str = Form("0"), category: str = Form("")):
    row = db.get(ItemCatalog, item_id)
    if row:
        row.name = (name or "").strip() or row.name
        row.unit = (unit or "ชิ้น").strip()
        row.unit_price = _to_float(unit_price, 0.0)
        row.category = (category or "").strip()
        db.commit()
    return RedirectResponse("/catalog?saved=1", status_code=303)


@router.post("/catalog/{item_id}/delete")
def catalog_delete(item_id: int, db: Session = Depends(get_db)):
    row = db.get(ItemCatalog, item_id)
    if row:
        db.delete(row); db.commit()
    return RedirectResponse("/catalog", status_code=303)


# ---------------- ทะเบียนนักเรียนกลาง (Student master) ----------------
_SEX_MAP = {"ช": "M", "ชาย": "M", "m": "M", "M": "M",
            "ญ": "F", "หญิง": "F", "f": "F", "F": "F"}


def _xlsx_download(wb, filename: str) -> Response:
    """ส่งไฟล์ Excel จากหน่วยความจำ (ไม่เขียนดิสก์) — กันปัญหา path/permission บนเซิร์ฟเวอร์"""
    import io as _io
    from urllib.parse import quote
    buf = _io.BytesIO()
    wb.save(buf)
    dispo = f"attachment; filename*=utf-8''{quote(filename)}"
    return Response(content=buf.getvalue(), media_type=_XLSX_MT,
                    headers={"Content-Disposition": dispo})


def serve_generated(path, media_type: str | None = None, *,
                    inline: bool = False, download_name: str | None = None,
                    count: bool = True) -> Response:
    """อ่านไฟล์ที่สร้างไว้เข้าหน่วยความจำแล้วส่งกลับ พร้อม Content-Disposition แบบ RFC 5987
    (filename*=utf-8'') — เลี่ยงปัญหา FileResponse กับชื่อไฟล์ภาษาไทยบนเซิร์ฟเวอร์
    count=True: นับโควตาเอกสาร (ทดลองใช้) — ครบแล้วพาไปหน้าต่ออายุ"""
    if count:
        from app.tenancy import current_school_id
        from app.accounts import consume_doc_quota
        ok, _info = consume_doc_quota(current_school_id.get())
        if not ok:
            return RedirectResponse("/trial-limit", status_code=303)
    from urllib.parse import quote
    p = Path(path)
    data = p.read_bytes()
    name = download_name or p.name
    disp = "inline" if inline else "attachment"
    dispo = f"{disp}; filename*=utf-8''{quote(name)}"
    return Response(content=data, media_type=media_type or "application/octet-stream",
                    headers={"Content-Disposition": dispo})


def _ai_key() -> str:
    """AI key กลาง (หลังบ้าน) ของโรงเรียนที่ล็อกอิน — คืน '' ถ้าไม่ใช่สมาชิก/ไม่ได้ตั้ง key"""
    from app.tenancy import current_school_id
    from app.accounts import ai_key_for
    return ai_key_for(current_school_id.get())


def _norm_sex(s: str) -> str:
    return _SEX_MAP.get((s or "").strip(), "")


@router.get("/students", response_class=HTMLResponse)
def students_page(request: Request, db: Session = Depends(get_db)):
    rows = db.query(Student).order_by(Student.level, Student.name).all()
    grad = sum(1 for s in rows if (s.level or "").strip() == "จบการศึกษา")
    return templates.TemplateResponse("students.html", {
        "request": request, "school": get_school(db), "rows": rows,
        "graduated": grad, "today_input": be_date_input(datetime.now()),
    })


@router.post("/students")
def student_add(db: Session = Depends(get_db), name: str = Form(""), sex: str = Form(""),
                birthdate: str = Form(""), level: str = Form(""), student_no: str = Form("")):
    nm = (name or "").strip()
    if nm:
        db.add(Student(name=nm, sex=_norm_sex(sex), birthdate=parse_be_date(birthdate),
                       level=(level or "").strip(), student_no=(student_no or "").strip()))
        db.commit()
    return RedirectResponse("/students", status_code=303)


@router.post("/students/{sid}/update")
def student_master_update(sid: int, request: Request, db: Session = Depends(get_db), name: str = Form(""),
                          sex: str = Form(""), birthdate: str = Form(""),
                          level: str = Form(""), student_no: str = Form("")):
    s = db.get(Student, sid)
    if s:
        s.name = (name or "").strip() or s.name
        s.sex = _norm_sex(sex)
        s.birthdate = parse_be_date(birthdate)
        s.level = (level or "").strip()
        s.student_no = (student_no or "").strip()
        db.commit()
    if request.headers.get("x-requested-with") == "fetch":
        return JSONResponse({"ok": bool(s)})
    return RedirectResponse("/students?saved=1", status_code=303)


@router.post("/students/{sid}/delete")
def student_master_delete(sid: int, db: Session = Depends(get_db)):
    s = db.get(Student, sid)
    if s:
        db.delete(s); db.commit()
    return RedirectResponse("/students", status_code=303)


# ลำดับชั้นเรียนสำหรับเลื่อนชั้นขึ้นปีใหม่
_LEVEL_LADDER = ["อ.1", "อ.2", "อ.3", "ป.1", "ป.2", "ป.3", "ป.4", "ป.5", "ป.6",
                 "ม.1", "ม.2", "ม.3"]
_GRADUATED = "จบการศึกษา"


@router.post("/students/promote")
def students_promote(db: Session = Depends(get_db)):
    """เลื่อนชั้นทุกคนขึ้น 1 ระดับ (ขึ้นปีการศึกษาใหม่) — ชั้นสูงสุดตั้งเป็น 'จบการศึกษา'"""
    idx = {lv: i for i, lv in enumerate(_LEVEL_LADDER)}
    n = 0
    for s in db.query(Student).all():
        i = idx.get((s.level or "").strip())
        if i is None:
            continue
        s.level = _GRADUATED if i >= len(_LEVEL_LADDER) - 1 else _LEVEL_LADDER[i + 1]
        n += 1
    db.commit()
    return RedirectResponse(f"/students?promoted={n}", status_code=303)


@router.post("/students/remove-graduated")
def students_remove_graduated(db: Session = Depends(get_db)):
    """ลบนักเรียนที่จบการศึกษาแล้ว"""
    rows = db.query(Student).filter(Student.level == _GRADUATED).all()
    n = len(rows)
    for s in rows:
        db.delete(s)
    db.commit()
    return RedirectResponse(f"/students?removed={n}", status_code=303)


@router.post("/students/bulk")
def students_master_bulk(db: Session = Depends(get_db), bulk: str = Form("")):
    """เพิ่มนักเรียนทีละหลายคน: 1 บรรทัด = ชื่อ, เพศ(ช/ญ), วันเกิด, ชั้น, เลขประจำตัว"""
    import re as _re
    n = 0
    for line in (bulk or "").splitlines():
        parts = [p.strip() for p in _re.split(r"[,\t]", line.strip())]
        if not parts or not parts[0]:
            continue
        db.add(Student(
            name=parts[0], sex=_norm_sex(parts[1] if len(parts) > 1 else ""),
            birthdate=parse_be_date(parts[2]) if len(parts) > 2 else None,
            level=parts[3] if len(parts) > 3 else "",
            student_no=parts[4] if len(parts) > 4 else ""))
        n += 1
    db.commit()
    return RedirectResponse(f"/students?added={n}", status_code=303)


@router.get("/students/template.xlsx")
def students_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "นักเรียน"
    headers = ["ชื่อ-นามสกุล", "เพศ (ช/ญ)", "วันเกิด (วว/ดด/ปปปป)", "ระดับชั้น", "เลขประจำตัว"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name="TH Sarabun New", bold=True, size=14)
        ws.column_dimensions[chr(64 + c)].width = 24
    ws.append(["เด็กชายสมชาย ใจดี", "ช", "15/05/2562", "ป.1", "10001"])
    return _xlsx_download(wb, "แบบฟอร์มนำเข้านักเรียน.xlsx")


@router.post("/students/import")
async def students_import(db: Session = Depends(get_db), file: UploadFile = File(...)):
    import io as _io
    from openpyxl import load_workbook
    data = await file.read()
    n = 0
    try:
        ws = load_workbook(_io.BytesIO(data), data_only=True).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            db.add(Student(
                name=str(row[0]).strip(),
                sex=_norm_sex(str(row[1]).strip() if len(row) > 1 and row[1] else ""),
                birthdate=parse_be_date(str(row[2])) if len(row) > 2 and row[2] else None,
                level=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                student_no=str(row[4]).strip() if len(row) > 4 and row[4] else ""))
            n += 1
        db.commit()
    except Exception:
        return RedirectResponse("/students?err=read", status_code=303)
    return RedirectResponse(f"/students?added={n}", status_code=303)


# ---------------- แผนการจัดซื้อจัดจ้างประจำปี ----------------
@router.get("/procurement/plan", response_class=HTMLResponse)
def proc_plan_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    fy = year or current_fiscal_year()
    rows = (db.query(ProcurementPlan).filter_by(fiscal_year=fy)
            .order_by(ProcurementPlan.seq, ProcurementPlan.id).all())
    years = sorted({r[0] for r in db.query(ProcurementPlan.fiscal_year).distinct()} |
                   set(range(fy - 2, fy + 2)), reverse=True)
    return templates.TemplateResponse("proc_plan.html", {
        "request": request, "school": get_school(db), "rows": rows, "fiscal_year": fy,
        "years": years, "total": sum(float(r.budget or 0) for r in rows),
        "today_input": be_date_input(datetime.now()),
    })


@router.post("/procurement/plan/add")
def proc_plan_add(db: Session = Depends(get_db), fiscal_year: str = Form(""), name: str = Form(""),
                  budget: str = Form("0"), method: str = Form("เฉพาะเจาะจง"),
                  expected_period: str = Form(""), source: str = Form("เงินอุดหนุน")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    nm = (name or "").strip()
    if nm:
        seq = (db.query(ProcurementPlan).filter_by(fiscal_year=fy).count()) + 1
        db.add(ProcurementPlan(fiscal_year=fy, seq=seq, name=nm, budget=_to_float(budget, 0.0),
                               method=(method or "").strip(), expected_period=(expected_period or "").strip(),
                               source=(source or "").strip()))
        db.commit()
    return RedirectResponse(f"/procurement/plan?year={fy}", status_code=303)


@router.post("/procurement/plan/{pid}/update")
def proc_plan_update(pid: int, db: Session = Depends(get_db), name: str = Form(""),
                     budget: str = Form("0"), method: str = Form(""),
                     expected_period: str = Form(""), source: str = Form("")):
    p = db.get(ProcurementPlan, pid)
    fy = p.fiscal_year if p else current_fiscal_year()
    if p:
        p.name = (name or "").strip() or p.name
        p.budget = _to_float(budget, 0.0)
        p.method = (method or "").strip()
        p.expected_period = (expected_period or "").strip()
        p.source = (source or "").strip()
        db.commit()
    return RedirectResponse(f"/procurement/plan?year={fy}&saved=1", status_code=303)


@router.post("/procurement/plan/{pid}/delete")
def proc_plan_delete(pid: int, db: Session = Depends(get_db)):
    p = db.get(ProcurementPlan, pid)
    fy = p.fiscal_year if p else current_fiscal_year()
    if p:
        db.delete(p); db.commit()
    return RedirectResponse(f"/procurement/plan?year={fy}", status_code=303)


@router.post("/procurement/plan/pull-projects")
def proc_plan_pull(db: Session = Depends(get_db), fiscal_year: str = Form("")):
    """ดึงรายการจากโครงการ/แผนงบ (ปีเดียวกัน) มาเป็นแผนจัดซื้อ (ข้ามชื่อที่มีแล้ว)"""
    fy = _to_int(fiscal_year, current_fiscal_year())
    have = {(p.name or "").strip() for p in db.query(ProcurementPlan).filter_by(fiscal_year=fy).all()}
    projs = db.query(Project).filter(Project.plan_year == fy).order_by(Project.name).all()
    seq = len(have)
    for pj in projs:
        nm = (pj.name or "").strip()
        if not nm or nm in have:
            continue
        seq += 1
        db.add(ProcurementPlan(fiscal_year=fy, seq=seq, name=nm, budget=pj.budget or 0.0,
                               project_id=pj.id, source="เงินอุดหนุน"))
        have.add(nm)
    db.commit()
    return RedirectResponse(f"/procurement/plan?year={fy}", status_code=303)


@router.get("/procurement/plan/announce.docx")
def proc_plan_announce(db: Session = Depends(get_db), year: int | None = None, date: str = ""):
    from app.services.proc_plan_doc import render_plan_announcement
    fy = year or current_fiscal_year()
    rows = (db.query(ProcurementPlan).filter_by(fiscal_year=fy)
            .order_by(ProcurementPlan.seq, ProcurementPlan.id).all())
    path = render_plan_announcement(get_school(db), fy, rows, parse_be_date(date))
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


# ---------------- ทะเบียนคุมสัญญา + แจ้งเตือนครบกำหนด ----------------
CONTRACT_TYPES = ["สัญญาจ้าง", "สัญญาซื้อ", "ใบสั่งจ้าง", "ใบสั่งซื้อ", "ข้อตกลง"]
CONTRACT_STATUS = ["ระหว่างดำเนินการ", "ส่งมอบแล้ว", "ตรวจรับแล้ว", "สิ้นสุด"]
_ALERT_DAYS = 15   # เตือนล่วงหน้ากี่วันก่อนครบกำหนด


def _contract_alert(c, today):
    """คืน (days_left, level) : level = overdue/soon/None (เฉพาะสัญญาที่ยังไม่สิ้นสุด)"""
    if c.status == "สิ้นสุด" or not c.end_date:
        return None, None
    dl = (c.end_date.date() - today).days
    if dl < 0:
        return dl, "overdue"
    if dl <= _ALERT_DAYS:
        return dl, "soon"
    return dl, None


@router.get("/procurement/contracts", response_class=HTMLResponse)
def contracts_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    fy = year or current_fiscal_year()
    today = datetime.now().date()
    rows = (db.query(Contract).filter_by(fiscal_year=fy)
            .order_by(Contract.end_date.is_(None), Contract.end_date).all())
    items = []
    for c in rows:
        dl, lvl = _contract_alert(c, today)
        items.append({"c": c, "days_left": dl, "level": lvl})
    alerts = [it for it in items if it["level"]]
    years = sorted({r[0] for r in db.query(Contract.fiscal_year).distinct()} |
                   set(range(fy - 2, fy + 2)), reverse=True)
    return templates.TemplateResponse("contracts.html", {
        "request": request, "school": get_school(db), "items": items, "alerts": alerts,
        "fiscal_year": fy, "years": years, "types": CONTRACT_TYPES, "statuses": CONTRACT_STATUS,
        "vendors": db.query(Vendor).order_by(Vendor.name).all(),
        "today_input": be_date_input(datetime.now()),
    })


@router.post("/procurement/contracts/add")
def contract_add(db: Session = Depends(get_db), fiscal_year: str = Form(""), contract_no: str = Form(""),
                 ctype: str = Form("ใบสั่งจ้าง"), party: str = Form(""), subject: str = Form(""),
                 amount: str = Form("0"), sign_date: str = Form(""), end_date: str = Form(""),
                 status: str = Form("ระหว่างดำเนินการ")):
    fy = _to_int(fiscal_year, current_fiscal_year())
    if (subject or party or contract_no).strip():
        db.add(Contract(fiscal_year=fy, contract_no=contract_no.strip(), ctype=ctype.strip() or "ใบสั่งจ้าง",
                        party=party.strip(), subject=subject.strip(), amount=_to_float(amount, 0.0),
                        sign_date=parse_be_date(sign_date), end_date=parse_be_date(end_date),
                        status=status.strip() or "ระหว่างดำเนินการ"))
        db.commit()
    return RedirectResponse(f"/procurement/contracts?year={fy}", status_code=303)


@router.post("/procurement/contracts/{cid}/update")
def contract_update(cid: int, db: Session = Depends(get_db), contract_no: str = Form(""),
                    ctype: str = Form(""), party: str = Form(""), subject: str = Form(""),
                    amount: str = Form("0"), sign_date: str = Form(""), end_date: str = Form(""),
                    status: str = Form("")):
    c = db.get(Contract, cid)
    fy = c.fiscal_year if c else current_fiscal_year()
    if c:
        c.contract_no = contract_no.strip(); c.ctype = ctype.strip() or c.ctype
        c.party = party.strip(); c.subject = subject.strip(); c.amount = _to_float(amount, 0.0)
        c.sign_date = parse_be_date(sign_date); c.end_date = parse_be_date(end_date)
        c.status = status.strip() or c.status
        db.commit()
    return RedirectResponse(f"/procurement/contracts?year={fy}&saved=1", status_code=303)


@router.post("/procurement/contracts/{cid}/delete")
def contract_delete(cid: int, db: Session = Depends(get_db)):
    c = db.get(Contract, cid)
    fy = c.fiscal_year if c else current_fiscal_year()
    if c:
        db.delete(c); db.commit()
    return RedirectResponse(f"/procurement/contracts?year={fy}", status_code=303)


@router.post("/procurement/contracts/pull-procurement")
def contracts_pull_procurement(db: Session = Depends(get_db), fiscal_year: str = Form("")):
    """ดึงเรื่องจัดซื้อจัดจ้างที่มีเลขใบสั่ง มาเป็นสัญญาในทะเบียน (ข้ามที่ดึงแล้ว)"""
    fy = _to_int(fiscal_year, current_fiscal_year())
    have = {(c.source, c.ref_id) for c in db.query(Contract).filter_by(fiscal_year=fy).all()}
    procs = (db.query(Procurement).filter(Procurement.fiscal_year == fy,
             Procurement.order_no != "").all())
    for p in procs:
        if ("procurement", p.id) in have or not (p.order_no or "").strip():
            continue
        ctype = "ใบสั่งจ้าง" if p.proc_type == "จ้าง" else "ใบสั่งซื้อ"
        db.add(Contract(fiscal_year=fy, contract_no=(p.order_no or "").strip(), ctype=ctype,
                        party=(p.vendor.name if p.vendor else ""), subject=p.subject,
                        amount=p.total_amount or 0.0, sign_date=p.order_date,
                        end_date=p.delivery_due_date or p.delivery_date,
                        status="ตรวจรับแล้ว" if p.status in ("ตรวจรับแล้ว", "เบิกจ่ายแล้ว") else "ระหว่างดำเนินการ",
                        source="procurement", ref_id=p.id))
    db.commit()
    return RedirectResponse(f"/procurement/contracts?year={fy}", status_code=303)


@router.post("/procurement/contracts/pull-lunch")
def contracts_pull_lunch(db: Session = Depends(get_db), fiscal_year: str = Form("")):
    """ดึงรอบจ้างเหมาอาหารกลางวันที่มีเลขใบสั่งจ้าง มาเป็นสัญญาในทะเบียน"""
    from app.models import LunchHireRound
    fy = _to_int(fiscal_year, current_fiscal_year())
    have = {(c.source, c.ref_id) for c in db.query(Contract).filter_by(fiscal_year=fy).all()}
    for r in db.query(LunchHireRound).filter(LunchHireRound.order_no != "").all():
        if ("lunch", r.id) in have or not (r.order_no or "").strip():
            continue
        db.add(Contract(fiscal_year=fy, contract_no=(r.order_no or "").strip(), ctype="ใบสั่งจ้าง",
                        party=(r.vendor.name if r.vendor else ""),
                        subject=f"จ้างเหมาประกอบอาหารกลางวัน รอบที่ {r.seq}",
                        amount=r.amount or 0.0, sign_date=r.order_date,
                        start_date=r.start_date, end_date=r.end_date,
                        status="จ่ายแล้ว" if r.status == "จ่ายแล้ว" else "ระหว่างดำเนินการ",
                        source="lunch", ref_id=r.id))
    db.commit()
    return RedirectResponse(f"/procurement/contracts?year={fy}", status_code=303)


@router.get("/procurement/contracts/export.xlsx")
def contracts_export(db: Session = Depends(get_db), year: int | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    fy = year or current_fiscal_year()
    today = datetime.now().date()
    rows = (db.query(Contract).filter_by(fiscal_year=fy)
            .order_by(Contract.end_date.is_(None), Contract.end_date).all())
    wb = Workbook(); ws = wb.active; ws.title = f"ทะเบียนคุมสัญญา {fy}"[:31]
    ws.append([f"ทะเบียนคุมสัญญา/ใบสั่งซื้อสั่งจ้าง ประจำปีงบประมาณ {fy}"])
    ws.append([])
    headers = ["ลำดับ", "เลขที่สัญญา/ใบสั่ง", "ประเภท", "คู่สัญญา", "เรื่อง/งาน", "วงเงิน (บาท)",
               "วันทำสัญญา", "ครบกำหนด", "สถานะ", "คงเหลือ (วัน)"]
    ws.append(headers)
    thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(3, c); cell.font = Font(name="TH Sarabun New", bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border; cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1).font = Font(name="TH Sarabun New", bold=True, size=16)
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    for i, c in enumerate(rows, start=1):
        dl, lvl = _contract_alert(c, today)
        left = ("เกินกำหนด %d วัน" % abs(dl)) if lvl == "overdue" else (str(dl) if dl is not None else "-")
        ws.append([i, c.contract_no or "-", c.ctype, c.party or "-", c.subject or "-", c.amount or 0,
                   thai_date(c.sign_date) if c.sign_date else "-",
                   thai_date(c.end_date) if c.end_date else "-", c.status, left])
    widths = [7, 16, 12, 22, 30, 14, 16, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.font = Font(name="TH Sarabun New", size=13); cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[5].number_format = "#,##0.00"
    return _xlsx_download(wb, f"ทะเบียนคุมสัญญา_ปีงบ{fy}.xlsx")


# ---------------- ผู้ขาย ----------------
@router.get("/vendors", response_class=HTMLResponse)
def vendors_page(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("vendors.html", {
        "request": request, "vendors": db.query(Vendor).order_by(Vendor.name).all(),
    })


@router.get("/vendors/template.xlsx")
def vendors_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = "ผู้ขาย"
    headers = ["ชื่อร้าน/บริษัท", "ชื่อเจ้าของ/ผู้ลงนาม", "เลขประจำตัวผู้เสียภาษี",
               "ที่อยู่", "เบอร์โทร", "เลขบัญชีธนาคาร"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name="TH Sarabun New", bold=True, size=14)
        ws.column_dimensions[chr(64 + c)].width = 24
    ws.append(["หจก. ตัวอย่างการค้า", "นายสมชาย ใจดี", "0123456789012", "123 ต.ในเมือง", "0812345678", ""])
    return _xlsx_download(wb, "แบบฟอร์มนำเข้าผู้ขาย.xlsx")


@router.post("/vendors/import")
async def vendors_import(db: Session = Depends(get_db), file: UploadFile = File(...)):
    import io as _io
    from openpyxl import load_workbook
    data = await file.read()
    existing = {v.name for v in db.query(Vendor).all()}
    n = 0
    try:
        ws = load_workbook(_io.BytesIO(data), data_only=True).active
        for row in ws.iter_rows(min_row=2, values_only=True):
            nm = (str(row[0]).strip() if row and row[0] else "")
            if not nm or nm in existing:
                continue
            db.add(Vendor(name=nm,
                          owner_name=str(row[1]).strip() if len(row) > 1 and row[1] else "",
                          tax_id=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                          address=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                          phone=str(row[4]).strip() if len(row) > 4 and row[4] else "",
                          bank_account=str(row[5]).strip() if len(row) > 5 and row[5] else ""))
            existing.add(nm); n += 1
        db.commit()
    except Exception:
        return RedirectResponse("/vendors?err=read", status_code=303)
    return RedirectResponse(f"/vendors?added={n}", status_code=303)


@router.post("/vendors")
def vendor_add(db: Session = Depends(get_db), name: str = Form(...), tax_id: str = Form(""),
               address: str = Form(""), phone: str = Form(""), bank_account: str = Form(""),
               owner_name: str = Form("")):
    db.add(Vendor(name=name, tax_id=tax_id, address=address, phone=phone,
                  bank_account=bank_account, owner_name=owner_name))
    db.commit()
    return RedirectResponse("/vendors", status_code=303)


@router.post("/vendors/quick-add")
def vendor_quick_add(db: Session = Depends(get_db), name: str = Form(...), tax_id: str = Form(""),
                     address: str = Form(""), phone: str = Form(""), bank_account: str = Form(""),
                     owner_name: str = Form("")):
    """เพิ่มผู้ขายแบบ AJAX (จาก modal หน้าสร้างเรื่อง) คืน JSON ไม่เปลี่ยนหน้า
    ถ้าชื่อซ้ำ ใช้รายเดิม (ไม่สร้างซ้ำ)"""
    name = (name or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "no_name"}, status_code=400)
    v = db.query(Vendor).filter(Vendor.name == name).first()
    if not v:
        v = Vendor(name=name, tax_id=tax_id.strip(), address=address.strip(),
                   phone=phone.strip(), bank_account=bank_account.strip(), owner_name=owner_name.strip())
        db.add(v); db.commit(); db.refresh(v)
    return JSONResponse({"ok": True, "id": v.id, "name": v.name})


@router.post("/vendors/{vendor_id}/update")
def vendor_update(vendor_id: int, db: Session = Depends(get_db), name: str = Form(...),
                  tax_id: str = Form(""), address: str = Form(""),
                  phone: str = Form(""), bank_account: str = Form(""),
                  owner_name: str = Form("")):
    v = db.get(Vendor, vendor_id)
    if v:
        v.name, v.tax_id, v.address = name, tax_id, address
        v.phone, v.bank_account = phone, bank_account
        v.owner_name = owner_name
        db.commit()
    return RedirectResponse("/vendors?saved=1", status_code=303)


@router.post("/vendors/{vendor_id}/delete")
def vendor_delete(vendor_id: int, db: Session = Depends(get_db)):
    v = db.get(Vendor, vendor_id)
    if v:
        db.delete(v); db.commit()
    return RedirectResponse("/vendors", status_code=303)


# ---------------- สร้าง/แก้ไขเรื่องจัดซื้อ ----------------
def _form_lists(db: Session) -> dict:
    """รายการมาสเตอร์ลิสต์ที่ฟอร์มสร้าง/แก้ไขใช้ร่วมกัน"""
    persons = db.query(Person).order_by(Person.name).all()
    return {
        "vendors": db.query(Vendor).order_by(Vendor.name).all(),
        "persons": persons,
        "departments": db.query(Department).order_by(Department.name).all(),
        "projects": db.query(Project).order_by(Project.name).all(),
        "catalog": db.query(ItemCatalog).order_by(ItemCatalog.name).all(),
        # แผนที่ ชื่อ -> ตำแหน่ง สำหรับเติมตำแหน่งอัตโนมัติเมื่อเลือกชื่อ
        "persons_pos": {p.name: p.position for p in persons},
    }


# ตำแหน่งมาตรฐานในโรงเรียน (ไว้ทำ dropdown ให้เลือก) เรียงผู้บริหารบนสุด
POSITION_CHOICES = [
    # ผู้บริหาร
    "ผู้อำนวยการโรงเรียน", "รองผู้อำนวยการโรงเรียน", "รักษาการในตำแหน่งผู้อำนวยการโรงเรียน",
    # ครู
    "ครู", "ครูผู้ช่วย", "ครูชำนาญการ", "ครูชำนาญการพิเศษ", "ครูเชี่ยวชาญ",
    "ครูอัตราจ้าง", "พนักงานราชการ",
    # เจ้าหน้าที่/สนับสนุน
    "หัวหน้าเจ้าหน้าที่", "เจ้าหน้าที่", "เจ้าหน้าที่พัสดุ", "เจ้าหน้าที่การเงิน",
    "ธุรการ", "นักการภารโรง",
    # คณะกรรมการสถานศึกษา / ผู้มีส่วนได้ส่วนเสีย
    "ประธานคณะกรรมการสถานศึกษาขั้นพื้นฐาน", "กรรมการสถานศึกษาขั้นพื้นฐาน",
    "ผู้แทนผู้ปกครอง", "ผู้แทนครู", "ผู้แทนองค์กรชุมชน", "ผู้แทนศิษย์เก่า",
    "ผู้ทรงคุณวุฒิ", "ผู้นำชุมชน", "ผู้ใหญ่บ้าน", "กำนัน",
]


def _assign_roles(members, kind: str) -> None:
    """เติมบทบาทอัตโนมัติ 'เฉพาะช่องที่ผู้ใช้ไม่ได้เลือก' (ปล่อยว่าง)
    ค่าที่ผู้ใช้เลือกเองจะไม่ถูกเขียนทับ
    - 1 คน: ผู้ตรวจรับ (ชุดตรวจรับ) / ประธานกรรมการ (ชุดคุณลักษณะ)
    - หลายคน: คนแรก=ประธานกรรมการ, คนสุดท้าย=กรรมการและเลขานุการ, ที่เหลือ=กรรมการ
    """
    n = len(members)
    for i, m in enumerate(members):
        if m.role:               # ผู้ใช้เลือกเองแล้ว -> ไม่ทับ
            continue
        if n == 1:
            m.role = "ผู้ตรวจรับ" if kind == "inspect" else "ประธานกรรมการ"
        elif i == 0:
            m.role = "ประธานกรรมการ"
        elif i == n - 1:
            m.role = "กรรมการและเลขานุการ"
        else:
            m.role = "กรรมการ"


def _build_committee(form, kind: str, mode: str, prefix: str) -> Committee:
    """สร้างคณะกรรมการจากฟิลด์ในฟอร์ม (prefix แยกชุด เช่น '' หรือ 'spec_')
    ใช้บทบาทที่ผู้ใช้เลือก ถ้าเว้นว่างจะเติมค่าอัตโนมัติให้"""
    committee = Committee(kind=kind, mode=mode)
    m_names = form.getlist(prefix + "member_name")
    m_pos = form.getlist(prefix + "member_position")
    m_role = form.getlist(prefix + "member_role")
    for i, mn in enumerate(m_names):
        mn = (mn or "").strip()
        if not mn:
            continue
        committee.members.append(CommitteeMember(
            name=mn,
            position=(m_pos[i] if i < len(m_pos) else "") or "ครู",
            role=(m_role[i] if i < len(m_role) else "").strip(),   # ค่าที่ผู้ใช้เลือก
            seq=i,
        ))
    _assign_roles(committee.members, kind)   # เติมเฉพาะช่องที่ปล่อยว่าง
    return committee


def _resolve_vendor(db: Session, form) -> int | None:
    """หา vendor_id จากชื่อที่กรอก — ถ้ายังไม่มีในระบบให้สร้างผู้ขายใหม่ให้อัตโนมัติ
    (รองรับ vendor_id เดิมเพื่อความเข้ากันได้ย้อนหลัง)"""
    name = (form.get("vendor_name") or "").strip()
    if name:
        v = db.query(Vendor).filter(Vendor.name == name).first()
        if not v:
            v = Vendor(name=name)
            db.add(v)
            db.flush()
        return v.id
    vid = form.get("vendor_id")
    return _to_int(vid, None) if vid else None


# รูปแบบการจัดซื้อ -> ป้าย/คำอธิบาย/เอกสารวิธีพิเศษ (altkind=None แปลว่าชุดเอกสารปกติ)
PROC_CASES = {
    "normal": {"label": "จัดซื้อ/จัดจ้าง (ปกติ)", "altkind": None,
               "short": "เฉพาะเจาะจงเต็มรูปแบบ",
               "desc": "ชุดเอกสารครบ (รายงานขอซื้อ/ตั้งกรรมการ/ใบสั่ง/ตรวจรับ/เบิกจ่าย ฯลฯ) ใช้กับงานจัดซื้อทั่วไป"},
    "w804": {"label": "ว.804 (ซื้อไม่เกิน 50,000 บาท)", "altkind": "w804",
             "short": "ซื้อของไม่เกิน 50,000 บาท",
             "desc": "เอกสารแบบย่อ เจ้าหน้าที่คนเดียวเป็นทั้งผู้ซื้อและผู้ตรวจรับ ไม่ต้องตั้งคณะกรรมการ"},
    "w119t1": {"label": "ว.119 ตาราง 1 (ซื้อพัสดุไม่เกิน 10,000 บาท)", "altkind": "w119t1",
               "short": "ซื้อพัสดุไม่เกิน 10,000 บาท",
               "desc": "ซื้อของ/วัสดุ/ครุภัณฑ์ จ่ายไปก่อนได้ แล้วรายงานผลภายใน 5 วันทำการ เอกสารใบเดียวจบ"},
    "w119t2": {"label": "ว.119 ตาราง 2 (ค่าบริหาร/ฝึกอบรม)", "altkind": "w119t2",
               "short": "ค่าบริหารงาน/ฝึกอบรม",
               "desc": "ค่าวิทยากร/ค่าอาหาร/จัดประชุม ไม่ถือเป็นการซื้อพัสดุ ไม่จำกัดวงเงิน เอกสารใบเดียวจบ"},
    "clause79": {"label": "ข้อ 79 วรรค 2 (จำเป็นเร่งด่วน)", "altkind": "clause79",
                 "short": "จำเป็นเร่งด่วน ทำไปก่อน",
                 "desc": "กรณีจำเป็นเร่งด่วนที่ไม่ได้คาดหมายไว้ก่อนและทำตามปกติไม่ทัน ดำเนินการจัดซื้อไปก่อน แล้วรายงานขอความเห็นชอบ (หัวหน้าเจ้าหน้าที่อนุมัติ) ถือรายงานเป็นการตรวจรับโดยอนุโลม"},
}


# ฟิลด์เสริมเฉพาะรูปแบบ (key, ป้ายกำกับ, ชนิด) — ใช้สร้างช่องกรอกในฟอร์ม + เก็บลง case_extra (JSON)
CASE_EXTRA_FIELDS = {
    "w804": [
        ("report2_no", "เลขที่บันทึกรายงานสรุปผล/ขอเบิกจ่าย", "text"),
        ("report2_date", "วันที่รายงานสรุปผล/ขอเบิกจ่าย", "date"),
        ("receipt_book", "ใบเสร็จรับเงิน เล่มที่", "text"),
        ("receipt_no", "ใบเสร็จรับเงิน เลขที่", "text"),
        ("receipt_date", "วันที่ใบเสร็จรับเงิน", "date"),
        ("deliver_date", "วันที่รับมอบพัสดุ", "date"),
    ],
    "w119t1": [
        ("budget_kind", "ประเภทงบ (เช่น รายหัว / 15 ปี / อื่น ๆ)", "text"),
        ("advance_payer", "ผู้ทดรองจ่าย/ผู้ยืมเงิน", "text"),
    ],
    "w119t2": [
        ("receipt_book", "ใบเสร็จ/ใบส่งของ เล่มที่", "text"),
        ("receipt_no", "ใบเสร็จ/ใบส่งของ เลขที่", "text"),
        ("receipt_date", "วันที่ใบเสร็จ/ใบส่งของ", "date"),
    ],
    "clause79": [
        ("receipt_book", "ใบเสร็จ/ใบส่งของ เล่มที่", "text"),
        ("receipt_no", "ใบเสร็จ/ใบส่งของ เลขที่", "text"),
        ("receipt_date", "วันที่ใบเสร็จ/ใบส่งของ", "date"),
    ],
}


def _collect_case_extra(proc: Procurement, form) -> None:
    """เก็บค่าฟิลด์เสริมเฉพาะรูปแบบ -> JSON ลง proc.case_extra"""
    import json
    spec = CASE_EXTRA_FIELDS.get(proc.proc_case or "normal")
    if not spec:
        proc.case_extra = ""
        return
    data = {key: (form.get(key) or "").strip() for key, _label, _t in spec}
    proc.case_extra = json.dumps(data, ensure_ascii=False)


def _populate_proc_from_form(proc: Procurement, form, db: Session, threshold: float = 0) -> None:
    """อ่านค่าจากฟอร์ม เติมลง proc + สร้างรายการพัสดุ + คณะกรรมการ (ตรวจรับ + คุณลักษณะ)
    (ใช้ทั้งตอนสร้างและแก้ไข ไม่ยุ่งกับเลขที่/วันที่เอกสาร)"""
    proc.fiscal_year = _to_int(form.get("fiscal_year"), current_fiscal_year())
    proc.subject = (form.get("subject") or "").strip()
    proc.project_name = (form.get("project_name") or "").strip()
    # ผูกกับโครงการในแผน (ถ้าชื่อตรงกับโครงการที่มีอยู่) -> ใช้ติดตามงบโครงการ
    proc.project_id = None
    if proc.project_name:
        q = db.query(Project).filter(Project.name == proc.project_name)
        proj = q.filter(Project.plan_year == current_plan_year(get_school(db))).first() or q.first()
        if proj:
            proc.project_id = proj.id
    proc.department = (form.get("department") or "").strip()
    proc.purpose = (form.get("purpose") or "").strip()
    proc.proc_type = form.get("proc_type") or "ซื้อ"
    proc.method = form.get("method") or "เฉพาะเจาะจง"
    case = (form.get("proc_case") or "").strip()
    if case in PROC_CASES:
        proc.proc_case = case
    elif not proc.proc_case:
        proc.proc_case = "normal"
    _collect_case_extra(proc, form)
    proc.budget_source = (form.get("budget_source") or "อุดหนุน").strip()
    proc.price_ref_source = (form.get("price_ref_source") or "การสืบราคาจากท้องตลาด").strip()
    proc.delivery_days = _to_int(form.get("delivery_days"), 7)
    proc.vat_mode = form.get("vat_mode") or "none"
    proc.wht_rate = _to_float(form.get("wht_rate"), 0.0)
    proc.order_signer = form.get("order_signer") or "director"
    proc.inspection_mode = form.get("inspection_mode") or "single"
    proc.vendor_id = _resolve_vendor(db, form)
    # วันที่รายงานขอซื้อ/จ้าง (ถ้ากรอกมา) — กันไม่ให้ขึ้นเป็นวันปัจจุบันเสมอ
    rd = (form.get("request_date") or "").strip()
    if rd:
        d = parse_be_date(rd)
        if d:
            proc.request_date = d

    # รายการพัสดุ (ล้างของเดิมแล้วสร้างใหม่)
    proc.items.clear()
    names, qtys = form.getlist("item_name"), form.getlist("item_qty")
    units, prices = form.getlist("item_unit"), form.getlist("item_price")
    total = 0.0
    for i, nm in enumerate(names):
        nm = (nm or "").strip()
        if not nm:
            continue
        qty = _to_float(qtys[i]) if i < len(qtys) else 0
        price = _to_float(prices[i]) if i < len(prices) else 0
        unit = (units[i] if i < len(units) else "") or "หน่วย"
        proc.items.append(ProcurementItem(name=nm, quantity=qty, unit=unit, unit_price=price))
        total += qty * price
        _catalog_upsert(db, nm, unit, price)   # เก็บเข้าคลังรายการพัสดุอัตโนมัติ
    proc.total_amount = total

    # บังคับโหมดตามวงเงิน: เกินเกณฑ์ต้องใช้คณะกรรมการตรวจรับ
    if threshold and total > threshold:
        proc.inspection_mode = "committee"

    # ลบคณะกรรมการเดิมทั้งหมด แล้วสร้างใหม่ (ตรวจรับ + คุณลักษณะ)
    for c in list(proc.committees):
        proc.committees.remove(c)
    inspect = _build_committee(form, "inspect", proc.inspection_mode, "")
    if inspect.members:
        proc.committees.append(inspect)
    spec = _build_committee(form, "spec", "committee", "spec_")
    if spec.members:
        proc.committees.append(spec)


@router.post("/procurement/ocr-items")
async def procurement_ocr_items(file: UploadFile = File(...)):
    """OCR รูปรายการพัสดุ -> คืนรายการ (JSON) ให้ฟอร์มเติมในตารางรายการ (ไม่ออกจากหน้า)"""
    from app.services.pdf_extract import ocr_image, _parse_items_from_text
    import tempfile, os
    data = await file.read()
    ext = file_upload.detect_ext(data, file.filename or "")
    if ext not in ("png", "jpg", "webp"):
        return JSONResponse({"items": [], "error": "ไฟล์ต้องเป็นรูปภาพ (png/jpg/webp)"})
    tmp = None
    try:
        with tempfile.NamedTemporaryFile(suffix="." + ext, delete=False) as tf:
            tf.write(data)
            tmp = tf.name
        text = ocr_image(tmp)
        items = _parse_items_from_text(text)
    finally:
        if tmp:
            try:
                os.unlink(tmp)
            except OSError:
                pass
    return JSONResponse({"items": items, "ocr": bool((text or '').strip())})


@router.post("/procurement/ai-items")
async def procurement_ai_items(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """ให้ AI (vision) อ่านรูปรายการพัสดุ -> คืนรายการ (JSON) ใช้ API key ของโรงเรียน"""
    from app.services.ai_extract import extract_items_from_image
    key = _ai_key()
    if not key:
        return JSONResponse({"items": [], "error": "ฟีเจอร์ AI สำหรับสมาชิก — ต่ออายุ/เป็นสมาชิกเพื่อใช้งาน"})
    data = await file.read()
    ext = file_upload.detect_ext(data, file.filename or "")
    if ext not in ("png", "jpg", "webp"):
        return JSONResponse({"items": [], "error": "ไฟล์ต้องเป็นรูปภาพ (png/jpg/webp)"})
    # ย่อรูปก่อนส่ง (ลดค่าใช้จ่าย + เข้าขนาดที่โมเดลรองรับ) แล้วส่งเป็น JPEG
    try:
        from PIL import Image, ImageOps
        import io as _io
        img = ImageOps.exif_transpose(Image.open(_io.BytesIO(data))).convert("RGB")
        img.thumbnail((1600, 1600))
        buf = _io.BytesIO(); img.save(buf, "JPEG", quality=85)
        img_bytes, media = buf.getvalue(), "image/jpeg"
    except Exception:
        media = {"jpg": "image/jpeg", "png": "image/png", "webp": "image/webp"}[ext]
        img_bytes = data
    res = extract_items_from_image(img_bytes, media, key)
    if res.get("error"):
        msg = {"request": "เรียก AI ไม่สำเร็จ (ตรวจ API key/อินเทอร์เน็ต)",
               "no_json": "AI ตอบไม่เป็นรูปแบบที่อ่านได้", "bad_json": "AI ตอบไม่เป็นรูปแบบที่อ่านได้"}
        return JSONResponse({"items": [], "error": msg.get(res["error"], res["error"])})
    return JSONResponse({"items": res.get("items", []), "ok": True})


@router.get("/procurement/new", response_class=HTMLResponse)
def procurement_new_form(request: Request, db: Session = Depends(get_db),
                         case: str | None = None):
    school = get_school(db)
    fy = current_fiscal_year()
    # ยังไม่เลือกรูปแบบ -> แสดงหน้าเลือกรูปแบบก่อน
    if case not in PROC_CASES:
        return templates.TemplateResponse("procurement_choose.html", {
            "request": request, "cases": PROC_CASES,
        })
    return templates.TemplateResponse("procurement_form.html", {
        "request": request, "p": None, "action": "/procurement/new",
        "prefill_items": [], "prefill_members": [], "prefill_spec_members": [],
        "prefill_subject": "", "prefill_memo": "", "prefill_date": "", "pending_file": "",
        "today_input": be_date_input(datetime.now()),
        "fiscal_year": fy, "today_thai": thai_date(),
        "sug_memo": suggest_doc_no(db, "memo", fy),
        "threshold": school.doc_set_threshold or 5000, "positions": POSITION_CHOICES,
        "proc_case": case, "case_info": PROC_CASES[case],
        "case_extra_fields": CASE_EXTRA_FIELDS.get(case, []), "case_extra": {},
        **_form_lists(db),
    })


@router.post("/procurement/new")
async def procurement_create(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    fy = _to_int(form.get("fiscal_year"), current_fiscal_year())
    # เลขบันทึก: ถ้าไม่ได้กรอกมา ระบบออกเลขถัดไปให้
    memo_no = (form.get("memo_no") or "").strip() or suggest_doc_no(db, "memo", fy)
    proc = Procurement(memo_no=memo_no)
    _populate_proc_from_form(proc, form, db, get_school(db).doc_set_threshold or 5000)
    # ไฟล์แนบ (กรณีสร้างเรื่องจากไฟล์)
    pf = (form.get("pending_file") or "").strip()
    if file_upload.SAFE_FILE_NAME.match(pf):
        proc.file_path = pf
    db.add(proc)
    db.flush()
    commit_doc_no(db, "memo", proc.fiscal_year, proc.memo_no, source="procurement",
                  ref_id=proc.id, subject=f"{proc.proc_type or ''}{proc.subject or ''}".strip())
    db.commit()
    db.refresh(proc)
    return RedirectResponse(f"/procurement/{proc.id}", status_code=303)


# ---------------- สร้างเรื่องจากไฟล์ (อ่าน PDF/Word อัตโนมัติ) ----------------
def _norm_items(raw):
    """แปลงรายการพัสดุจากการสกัด -> objects ที่เทมเพลตใช้ (name/quantity/unit/unit_price)"""
    out = []
    for it in (raw or []):
        if not isinstance(it, dict):
            continue
        name = (it.get("name") or "").strip()
        if not name:
            continue
        out.append(SimpleNamespace(
            name=name, quantity=_to_float(it.get("qty", it.get("quantity", 1)), 1) or 1,
            unit=(it.get("unit") or "ชิ้น"), unit_price=_to_float(it.get("unit_price", 0), 0)))
    return out


def _norm_members(raw):
    out = []
    for m in (raw or []):
        if not isinstance(m, dict):
            continue
        name = (m.get("name") or "").strip()
        if not name:
            continue
        out.append(SimpleNamespace(name=name, position=(m.get("position") or "ครู"),
                                   role=(m.get("role") or "กรรมการ")))
    return out


def _render_proc_from_file(request, db, pending_file: str, fields: dict, ai_note=None):
    """แสดงฟอร์มสร้างเรื่องที่เติมค่าจากการสกัดไฟล์ (heuristic หรือ AI)"""
    school = get_school(db)
    fy = current_fiscal_year()
    items = _norm_items(fields.get("items"))
    members = _norm_members(fields.get("inspectors"))
    ptype = "จ้าง" if "จ้าง" in (fields.get("proc_type") or "") else "ซื้อ"
    return templates.TemplateResponse("procurement_form.html", {
        "request": request, "p": None, "action": "/procurement/new",
        "prefill_items": items, "prefill_members": members, "prefill_spec_members": [],
        "prefill_subject": fields.get("subject", ""),
        "prefill_memo": fields.get("memo_no", "") or fields.get("letter_no", ""),
        "prefill_type": ptype,
        "prefill_project": fields.get("project_name", ""),
        "prefill_dept": fields.get("department", ""),
        "prefill_purpose": fields.get("purpose", ""),
        "prefill_budget": fields.get("budget_source", ""),
        "prefill_delivery": _to_int(fields.get("delivery_days"), 7) or 7,
        "prefill_vendor": fields.get("vendor_name", ""),
        "prefill_mode": "committee" if len(members) > 1 else "single",
        "prefill_date": be_date_input(fields.get("request_date")),
        "today_input": be_date_input(datetime.now()),
        "pending_file": pending_file, "ai_note": ai_note,
        "fiscal_year": fy, "today_thai": thai_date(),
        "sug_memo": suggest_doc_no(db, "memo", fy),
        "threshold": school.doc_set_threshold or 5000, "positions": POSITION_CHOICES,
        **_form_lists(db),
    })


def _extract_proc(path: str, use_ai: bool, db):
    """สกัดข้อมูล: ถ้าเลือก AI และตั้งค่า key ไว้ ใช้ AI ไม่งั้น heuristic; คืน (fields, ai_note)"""
    if use_ai:
        key = _ai_key()
        if not key:
            return extract_procurement_fields(path), "ฟีเจอร์ AI สำหรับสมาชิก — ใช้การอ่านแบบปกติแทน (ต่ออายุเพื่อเปิด AI)"
        res = extract_with_ai(extract_text_any(path), key)
        if res.get("ok"):
            return res, "อ่านด้วย AI สำเร็จ — โปรดตรวจสอบความถูกต้องก่อนบันทึก"
        return extract_procurement_fields(path), "อ่านด้วย AI ไม่สำเร็จ ใช้การอ่านแบบปกติแทน (ตรวจ API key/อินเทอร์เน็ต)"
    return extract_procurement_fields(path), None


@router.get("/procurement/from-file", response_class=HTMLResponse)
def proc_from_file_page(request: Request, db: Session = Depends(get_db), err: str | None = None):
    err_msg = {"notpdf": "ไฟล์/ลิงก์ไม่ใช่ PDF หรือ Word (.docx) หรือต้องล็อกอินก่อน",
               "url": "ลิงก์ไม่ถูกต้อง",
               "fetch": "ดึงไฟล์จากลิงก์ไม่สำเร็จ (เครื่องต้องต่อเน็ตและเข้าถึงลิงก์ได้)"}.get(err)
    return templates.TemplateResponse("procurement_from_file.html", {
        "request": request, "err": err_msg, "has_ai_key": bool(_ai_key()),
    })


@router.post("/procurement/from-file/upload")
async def proc_from_file_upload(request: Request, db: Session = Depends(get_db),
                                file: UploadFile = File(...), use_ai: str = Form("")):
    data = await file.read()
    ext = file_upload.detect_ext(data, file.filename or "")
    if not ext:
        return RedirectResponse("/procurement/from-file?err=notpdf", status_code=303)
    name = file_upload.save_upload(data, ext)
    fields, ai_note = _extract_proc(str(file_upload.uploads_dir() / name), bool(use_ai), db)
    return _render_proc_from_file(request, db, name, fields, ai_note)


@router.post("/procurement/from-file/fetch")
async def proc_from_file_fetch(request: Request, db: Session = Depends(get_db),
                               url: str = Form(""), use_ai: str = Form("")):
    data, ext, err = file_upload.fetch_file(url)
    if err:
        return RedirectResponse(f"/procurement/from-file?err={err}", status_code=303)
    name = file_upload.save_upload(data, ext)
    fields, ai_note = _extract_proc(str(file_upload.uploads_dir() / name), bool(use_ai), db)
    return _render_proc_from_file(request, db, name, fields, ai_note)


@router.get("/procurement/file/{name}")
def proc_serve_file(name: str):
    if not file_upload.SAFE_FILE_NAME.match(name):
        return RedirectResponse("/procurement", status_code=303)
    path = file_upload.uploads_dir() / name
    if not path.exists():
        return RedirectResponse("/procurement", status_code=303)
    if name.lower().endswith(".pdf"):
        return serve_generated(str(path), "application/pdf", inline=True)
    return serve_generated(str(path), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.get("/procurement/{proc_id}/edit", response_class=HTMLResponse)
def procurement_edit_form(proc_id: int, request: Request, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    school = get_school(db)
    inspect = next((c for c in proc.committees if c.kind == "inspect"), None)
    spec = next((c for c in proc.committees if c.kind == "spec"), None)
    import json
    try:
        extra_vals = json.loads(proc.case_extra) if proc.case_extra else {}
    except Exception:
        extra_vals = {}
    return templates.TemplateResponse("procurement_form.html", {
        "request": request, "p": proc, "action": f"/procurement/{proc.id}/edit",
        "prefill_items": list(proc.items),
        "prefill_members": list(inspect.members) if inspect else [],
        "prefill_spec_members": list(spec.members) if spec else [],
        "fiscal_year": proc.fiscal_year, "today_thai": thai_date(),
        "threshold": school.doc_set_threshold or 5000, "positions": POSITION_CHOICES,
        "case_extra_fields": CASE_EXTRA_FIELDS.get(proc.proc_case or "normal", []),
        "case_extra": extra_vals,
        **_form_lists(db),
    })


@router.post("/procurement/{proc_id}/edit")
async def procurement_edit_save(proc_id: int, request: Request, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    form = await request.form()
    _populate_proc_from_form(proc, form, db, get_school(db).doc_set_threshold or 5000)
    db.commit()
    return RedirectResponse(f"/procurement/{proc_id}?saved=1", status_code=303)


# ---------------- รายละเอียด ----------------
@router.get("/procurement/{proc_id}", response_class=HTMLResponse)
def procurement_detail(proc_id: int, request: Request, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    inspect = next((c for c in proc.committees if c.kind == "inspect"), None)
    fy = proc.fiscal_year
    order_type = "purchase_order" if proc.proc_type == "ซื้อ" else "hire_order"
    docs = sorted(proc.documents, key=lambda d: d.id, reverse=True)
    # ปฏิทินวันหยุด (สำหรับเตือนเมื่อวันที่ลงนามตรงวันหยุด) — ครอบคลุมปีที่เกี่ยวข้อง
    holiday_years = year_range_for(proc.request_date, proc.order_date,
                                   proc.delivery_due_date, proc.inspect_date)
    # ข้อมูลเลขที่ใช้ไปแล้ว (สำหรับเตือนเลขซ้ำ/เลยเลข) แยกตามชนิด — ไม่รวมเลขของเรื่องนี้เอง
    docno_used = {}
    for dt in ("memo", "command", order_type):
        rows = (db.query(IssuedDocNo)
                .filter(IssuedDocNo.fiscal_year == fy, IssuedDocNo.doc_type == dt)
                .all())
        counter = next((c for c in db.query(DocNumberCounter)
                        .filter_by(doc_type=dt, fiscal_year=fy)), None)
        docno_used[dt] = {
            "last": counter.last_number if counter else 0,
            "used": {r.seq: (r.subject or "") for r in rows
                     if not (r.source == "procurement" and r.ref_id == proc.id)},
        }
    return templates.TemplateResponse("procurement_detail.html", {
        "request": request, "p": proc, "school": get_school(db),
        "doc_kinds": AVAILABLE_KINDS, "inspect": inspect, "documents": docs,
        "statuses": ["ร่าง", "อนุมัติ", "ตรวจรับแล้ว", "เบิกจ่ายแล้ว"],
        # เลขที่ที่ระบบเสนอ (ไว้เติมช่องว่างในฟอร์มแก้ไข)
        "sug_order": suggest_doc_no(db, order_type, fy),
        "sug_command": suggest_doc_no(db, "command", fy),
        "sug_memo": suggest_doc_no(db, "memo", fy),
        "holidays_json": holiday_map(holiday_years),
        "docno_used": docno_used, "order_type": order_type,
        "case_info": PROC_CASES.get(proc.proc_case or "normal", PROC_CASES["normal"]),
    })


@router.get("/document/{doc_id}/download")
def document_download(doc_id: int, db: Session = Depends(get_db)):
    """ดาวน์โหลดไฟล์เอกสารที่เคยสร้างไว้ (ถ้าไฟล์ยังอยู่)"""
    doc = db.get(Document, doc_id)
    if not doc or not doc.file_path or not Path(doc.file_path).exists():
        # ไฟล์ถูกลบ/ย้าย ให้กลับไปหน้ารายละเอียดเพื่อสร้างใหม่
        ref = doc.procurement_id if doc else ""
        return RedirectResponse(f"/procurement/{ref}", status_code=303)
    return serve_generated(doc.file_path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.post("/procurement/{proc_id}/status")
def procurement_set_status(proc_id: int, status: str = Form(...),
                           ajax: str = Form(""), db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if proc:
        proc.status = status; db.commit()
    if ajax:                              # เรียกจาก dropdown ในตาราง (ไม่ต้องเปลี่ยนหน้า)
        return Response(status_code=204)
    return RedirectResponse(f"/procurement/{proc_id}", status_code=303)


def _parse_date(s: str):
    """แปลงวันที่จากช่องกรอก (พ.ศ. วว/ดด/ปปปป) เป็น datetime"""
    return parse_be_date(s)


@router.post("/procurement/{proc_id}/update-refs")
async def procurement_update_refs(proc_id: int, request: Request, db: Session = Depends(get_db)):
    """แก้ไขเลขที่เอกสารและวันที่ภายหลัง (ใบสั่งซื้อ/จ้าง, คำสั่ง, วันที่ตรวจรับ ฯลฯ)"""
    from datetime import timedelta
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    form = await request.form()

    # เลขที่เอกสาร (ข้อความ)
    proc.memo_no = (form.get("memo_no") or "").strip()
    proc.result_memo_no = (form.get("result_memo_no") or "").strip()
    proc.spec_memo_no = (form.get("spec_memo_no") or "").strip()
    proc.inspect_memo_no = (form.get("inspect_memo_no") or "").strip()
    proc.command_no = (form.get("command_no") or "").strip()
    proc.order_no = (form.get("order_no") or "").strip()

    # วันที่หลักของแต่ละเอกสาร
    rd = _parse_date(form.get("request_date"))
    if rd:
        proc.request_date = rd
    proc.spec_memo_date   = _parse_date(form.get("spec_memo_date"))
    proc.command_date     = _parse_date(form.get("command_date"))
    proc.result_memo_date = _parse_date(form.get("result_memo_date"))
    proc.quotation_date   = _parse_date(form.get("quotation_date"))
    proc.order_date       = _parse_date(form.get("order_date"))
    proc.inspect_date     = _parse_date(form.get("inspect_date"))

    # วันครบกำหนดส่งมอบ — ถ้าไม่ได้กรอกให้คำนวณอัตโนมัติจากวันที่ใบสั่ง + จำนวนวัน
    due = _parse_date(form.get("delivery_due_date"))
    if due:
        proc.delivery_due_date = due
    elif proc.order_date:
        proc.delivery_due_date = proc.order_date + timedelta(days=int(proc.delivery_days or 7))
    # วันที่ส่งมอบจริง (ใบส่งมอบงาน) — ถ้าเว้นว่างใช้วันครบกำหนดในเอกสารแทน
    proc.delivery_date = _parse_date(form.get("delivery_date"))

    # ค่าปรับ + ใบส่งของ (ตรวจรับ)
    proc.overdue_days = _to_int(form.get("overdue_days"), 0)
    proc.delivery_note_no = (form.get("delivery_note_no") or "").strip()
    proc.delivery_note_book = (form.get("delivery_note_book") or "").strip()

    # bump counters + บันทึกลงทะเบียนเลขกลาง (เลขรันรวมทั้งโรงเรียน)
    fy = proc.fiscal_year
    subj = f"{proc.proc_type or ''}{proc.subject or ''}".strip()
    for no in (proc.memo_no, proc.result_memo_no, proc.spec_memo_no, proc.inspect_memo_no):
        commit_doc_no(db, "memo", fy, no, source="procurement", ref_id=proc.id, subject=subj)
    commit_doc_no(db, "command", fy, proc.command_no, source="procurement", ref_id=proc.id, subject=subj)
    commit_doc_no(db, "purchase_order" if proc.proc_type == "ซื้อ" else "hire_order", fy,
                  proc.order_no, source="procurement", ref_id=proc.id, subject=subj)

    db.commit()
    return RedirectResponse(f"/procurement/{proc_id}?saved=1", status_code=303)


@router.post("/procurement/{proc_id}/duplicate")
def procurement_duplicate(proc_id: int, db: Session = Depends(get_db)):
    """คัดลอกเรื่องเดิมเป็นฉบับร่างใหม่ (เลขบันทึกใหม่ ไม่ก๊อปเลข/วันที่เอกสารอื่น)"""
    src = db.get(Procurement, proc_id)
    if not src:
        return RedirectResponse("/procurement", status_code=303)
    fy = src.fiscal_year
    new = Procurement(
        fiscal_year=fy, memo_no=suggest_doc_no(db, "memo", fy),
        subject=(src.subject or "") + " (สำเนา)", project_name=src.project_name,
        department=src.department, purpose=src.purpose, proc_type=src.proc_type,
        method=src.method, budget_source=src.budget_source, price_ref_source=src.price_ref_source,
        total_amount=src.total_amount, delivery_days=src.delivery_days, penalty_rate=src.penalty_rate,
        vat_mode=src.vat_mode, order_signer=src.order_signer,
        inspection_mode=src.inspection_mode, vendor_id=src.vendor_id, status="ร่าง",
    )
    for it in src.items:
        new.items.append(ProcurementItem(name=it.name, quantity=it.quantity,
                                         unit=it.unit, unit_price=it.unit_price))
    for c in src.committees:
        nc = Committee(kind=c.kind, mode=c.mode)
        for m in c.members:
            nc.members.append(CommitteeMember(name=m.name, position=m.position, role=m.role, seq=m.seq))
        new.committees.append(nc)
    db.add(new)
    commit_doc_no(db, "memo", fy, new.memo_no)
    db.commit()
    db.refresh(new)
    return RedirectResponse(f"/procurement/{new.id}/edit", status_code=303)


@router.post("/procurement/{proc_id}/delete")
def procurement_delete(proc_id: int, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if proc:
        db.delete(proc); db.commit()
    return RedirectResponse("/", status_code=303)


@router.post("/procurement/{proc_id}/generate")
def procurement_generate(proc_id: int, doc_kind: str = Form(...), db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc or doc_kind not in AVAILABLE_KINDS:
        return RedirectResponse(f"/procurement/{proc_id}", status_code=303)
    file_path = render_document(doc_kind, proc, get_school(db))
    db.add(Document(procurement_id=proc.id, doc_kind=doc_kind, file_path=file_path))
    db.commit()
    return serve_generated(file_path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.get("/procurement/{proc_id}/altdoc/{kind}")
def procurement_altdoc(proc_id: int, kind: str, db: Session = Depends(get_db)):
    """ออกเอกสารจัดซื้อวิธีพิเศษ (ว.804 / ว.119 ตาราง 1 / ว.119 ตาราง 2)"""
    from app.services.proc_alt_doc import RENDERERS
    proc = db.get(Procurement, proc_id)
    renderer = RENDERERS.get(kind)
    if not proc or renderer is None:
        return RedirectResponse(f"/procurement/{proc_id}", status_code=303)
    file_path = renderer(proc, get_school(db))
    return serve_generated(file_path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


# ---------------- ทะเบียน Excel ----------------
# เอกสารที่ "ไม่ติ๊กอัตโนมัติ" (ยังเลือกเองได้) แยกตามวงเงิน
EXCLUDE_SMALL = {  # วงเงิน ≤ เกณฑ์: ไม่ต้องประกาศ/ตั้ง กก.คุณลักษณะ/TOR/คำสั่งแต่งตั้ง
    "ประกาศผู้ชนะ", "แต่งตั้งกรรมการคุณลักษณะ",
    "รายละเอียดคุณลักษณะ(TOR)", "คำสั่งแต่งตั้งผู้ตรวจรับ",
}
EXCLUDE_LARGE = {  # วงเงิน > เกณฑ์: ไม่ต้องตั้ง กก.คุณลักษณะ/TOR
    "แต่งตั้งกรรมการคุณลักษณะ", "รายละเอียดคุณลักษณะ(TOR)",
}


# แผนที่ "การเบิกจ่าย" -> งบใน/นอก พรบ.รายจ่าย (e-GP)
_EGP_INPB = {"เงินงบประมาณ", "งบประมาณ", "เงินรายได้แผ่นดิน"}  # เงินในงบ พรบ.รายจ่าย

# คำนำหน้าชื่อที่พบบ่อย (เรียงยาว->สั้น เพื่อจับแบบ longest-match) สำหรับแยกคำนำหน้า/ชื่อ/สกุล
_NAME_PREFIXES = [
    "ว่าที่ร้อยตรีหญิง", "ว่าที่ร้อยตรี", "ว่าที่ร้อยเอก", "ว่าที่ร้อยโท", "ว่าที่พันตรี",
    "นางสาว", "นาง", "นาย", "ดร.", "ดร",
]


def _split_name(full: str) -> dict:
    """แยกชื่อเต็ม -> {prefix, first, last} สำหรับช่อง P1/P2/P3 ใน e-GP (เดาแบบง่าย แก้เองได้)"""
    full = (full or "").strip()
    prefix = ""
    for p in _NAME_PREFIXES:
        if full.startswith(p):
            prefix = p
            full = full[len(p):].strip()
            break
    parts = full.split()
    first = parts[0] if parts else ""
    last = " ".join(parts[1:]) if len(parts) > 1 else ""
    return {"prefix": prefix, "first": first, "last": last}


@router.get("/procurement/{proc_id}/egp", response_class=HTMLResponse)
def procurement_egp(proc_id: int, request: Request, db: Session = Depends(get_db)):
    """หน้า 'ช่วยกรอก e-GP' (copy-assist) — จัดข้อมูลเรื่องจัดซื้อให้คัดลอกไปวางลง e-GP ทีละช่อง/ทีละขั้นตอน
    อิงโครงสร้างจริงของ e-GP (จัดทำโครงการ + รายงานขอซื้อขอจ้าง + ผู้ค้า/ราคา)
    ไม่ยุ่งกับบัญชี/รหัสผ่าน e-GP และไม่ทำ autofill อัตโนมัติ"""
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    school = get_school(db)
    src = (proc.budget_source or "").strip()
    in_pb = any(k in src for k in _EGP_INPB)   # เงินในงบ พรบ.รายจ่าย ?
    buy_word = "ซื้อ" if (proc.proc_type or "ซื้อ") == "ซื้อ" else "จ้าง"

    # เรียน = ผู้อำนวยการ + ชื่อโรงเรียน (ถ้าชื่อขึ้นต้นด้วย "โรงเรียน")
    sname = (school.name or "").strip()
    if sname.startswith("โรงเรียน"):
        attn = "ผู้อำนวยการ" + sname
    else:
        attn = (school.director_position or "ผู้อำนวยการโรงเรียน")

    # ผู้ลงนามในรายงานขอซื้อขอจ้าง = เจ้าหน้าที่พัสดุ (ผู้รายงาน) ถ้าไม่มีใช้หัวหน้าเจ้าหน้าที่
    signer = (school.officer_name or school.head_officer_name or "").strip()
    signer_pos = "เจ้าหน้าที่พัสดุ" if school.officer_name else "หัวหน้าเจ้าหน้าที่พัสดุ"

    egp = {
        "in_pb": in_pb,
        "buy_word": buy_word,                       # ซื้อ / จ้าง
        "school_name": sname,
        "attn": attn,                               # เรียน
        "report_subject": f"รายงานขอ{buy_word}",    # B2 เรื่อง
        "committee": "ไม่จัดทำ" if (proc.inspection_mode or "single") == "single" else "จัดทำ",
        "duration_text": f"{proc.delivery_days or 7} วัน",
        "signer_name": signer,
        "signer_pos": signer_pos,
        "signer_split": _split_name(signer),
        "owner_split": _split_name(proc.vendor.owner_name if proc.vendor else ""),
    }
    # คณะกรรมการ/ผู้ตรวจรับ (สำหรับช่อง "ชื่อผู้ตรวจรับ")
    inspect = next((c for c in proc.committees if c.kind == "inspect"), None)
    return templates.TemplateResponse("egp_helper.html", {
        "request": request, "p": proc, "school": school,
        "vendor": proc.vendor, "items": proc.items,
        "budget_in_pb": in_pb, "egp": egp, "inspect": inspect,
    })


@router.get("/procurement/{proc_id}/bundle", response_class=HTMLResponse)
def bundle_page(proc_id: int, request: Request, db: Session = Depends(get_db)):
    """หน้าเลือกเอกสารที่จะออกเป็นชุด (ติ๊กเลือกได้)"""
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    school = get_school(db)
    threshold = school.doc_set_threshold or 5000
    is_large = (proc.total_amount or 0) > threshold
    # ติ๊กให้อัตโนมัติ = ทุกใบ ยกเว้นรายการที่ไม่จำเป็นตามวงเงิน (ยังเลือกเองได้)
    exclude = EXCLUDE_LARGE if is_large else EXCLUDE_SMALL
    kinds = [{"name": k, "checked": k not in exclude} for k in AVAILABLE_KINDS]
    return templates.TemplateResponse("bundle.html", {
        "request": request, "p": proc, "kinds": kinds,
        "is_large": is_large, "threshold": threshold,
    })


@router.post("/procurement/{proc_id}/bundle")
async def bundle_generate(proc_id: int, request: Request, db: Session = Depends(get_db)):
    """สร้างเอกสารทุกใบที่เลือก แล้วรวมเป็นไฟล์ .docx เดียว (แต่ละใบขึ้นหน้าใหม่)"""
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    form = await request.form()
    # คงลำดับตาม AVAILABLE_KINDS (ขั้นตอนเอกสาร) ไม่ใช่ลำดับที่ติ๊ก
    chosen = set(form.getlist("kinds"))
    selected = [k for k in AVAILABLE_KINDS if k in chosen]
    if not selected:
        return RedirectResponse(f"/procurement/{proc_id}/bundle", status_code=303)

    school = get_school(db)
    out_path = render_bundle(selected, proc, school)
    db.add(Document(procurement_id=proc.id,
                    doc_kind=f"ชุดเอกสาร ({len(selected)} ใบ)", file_path=out_path))
    db.commit()

    # เสิร์ฟจากหน่วยความจำ + Content-Disposition แบบ RFC 5987 (รองรับชื่อไฟล์ไทยบนเซิร์ฟเวอร์)
    return serve_generated(out_path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.get("/register.xlsx")
def download_register(db: Session = Depends(get_db), year: int | None = None,
                      kind: str = "all"):
    """ดาวน์โหลดทะเบียนคุม Excel — kind: all / buy (จัดซื้อ) / hire (จัดจ้าง)
    แยกเล่มจัดซื้อ-จัดจ้างตามระเบียบพัสดุ"""
    fy = year or current_fiscal_year()
    query = db.query(Procurement).filter(Procurement.fiscal_year == fy)
    if kind == "buy":
        query = query.filter(Procurement.proc_type == "ซื้อ")
    elif kind == "hire":
        query = query.filter(Procurement.proc_type == "จ้าง")
    procurements = sorted(query.all(), key=_order_sort_key)   # เรียงตามเลขใบสั่ง น้อย->มาก
    path = export_register(procurements, fy, kind=kind)
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/summary.xlsx")
def download_monthly_summary(db: Session = Depends(get_db), year: int | None = None):
    """ดาวน์โหลดแบบ สขร.1 (สรุปผลการจัดซื้อจัดจ้างรายเดือน) แยกชีตตามเดือน"""
    from app.services.register_export import export_monthly_summary
    fy = year or current_fiscal_year()
    procurements = db.query(Procurement).filter(Procurement.fiscal_year == fy).all()
    school = get_school(db)
    path = export_monthly_summary(procurements, fy, school_name=(school.name if school else ""))
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ============================================================
# เฟส 3.1 — ทะเบียนครุภัณฑ์ + ค่าเสื่อมราคา
# ============================================================
_XLSX_MT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ============================================================
# โครงการ / แผนปฏิบัติการ (งบรายปี + ใช้จริง + ประวัติการปรับงบ)
# ============================================================
def _plan_years(db, cur):
    # ปีที่มีโครงการอยู่จริง + ช่วงปีรอบ ๆ ปีปัจจุบัน (ย้อนหลัง 3 ปี ถึงปีหน้า) ให้เลือกล่วงหน้าได้
    ys = {r[0] for r in db.query(Project.plan_year).distinct() if r[0]}
    ys.update(range(cur - 3, cur + 2))
    return sorted(ys, reverse=True)


@router.get("/projects", response_class=HTMLResponse)
def projects_page(request: Request, db: Session = Depends(get_db), year: int | None = None):
    school = get_school(db)
    cur = year or current_plan_year(school)
    rows = db.query(Project).filter(Project.plan_year == cur).order_by(Project.name).all()
    # โครงการเดิมที่ยังไม่ระบุปี (ก่อนมีโมดูลนี้) แสดงให้กำหนดปีได้ เมื่อดูปีปัจจุบัน
    legacy = []
    if cur == current_plan_year(school):
        legacy = db.query(Project).filter(Project.plan_year.is_(None)).order_by(Project.name).all()
    return templates.TemplateResponse("projects.html", {
        "request": request, "school": school, "rows": rows, "legacy": legacy,
        "fiscal_year": cur, "year_label": plan_year_label(school),
        "years": _plan_years(db, cur),
        "departments": db.query(Department).order_by(Department.name).all(),
        "total_budget": sum(project_budget(p) for p in rows),
        "total_spent": sum(project_spent(p) for p in rows),
    })


@router.post("/projects")
def project_add(db: Session = Depends(get_db), name: str = Form(...), budget: str = Form("0"),
                responsible: str = Form(""), plan_year: str = Form("")):
    py = _to_int(plan_year, current_plan_year(get_school(db)))
    if name.strip():
        db.add(Project(name=name.strip(), budget=_to_float(budget, 0.0),
                       responsible=responsible.strip(), plan_year=py, active=True))
        db.commit()
    return RedirectResponse(f"/projects?year={py}", status_code=303)


def _parse_projects_xlsx(content: bytes) -> list[dict]:
    """อ่านไฟล์ Excel โครงการ คืนค่า [{name, budget, responsible}] (รองรับมีหัวตาราง/ไม่มี)"""
    from openpyxl import load_workbook
    import io as _io
    wb = load_workbook(_io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]

    def find(*keys):
        for i, h in enumerate(header):
            if any(k in h for k in keys):
                return i
        return None

    ci_name = find("ชื่อโครงการ", "โครงการ", "ชื่อ")
    ci_budget = find("งบประมาณ", "งบ", "จำนวนเงิน", "วงเงิน")
    ci_resp = find("ผู้รับผิดชอบ", "รับผิดชอบ", "ฝ่าย")
    has_header = ci_name is not None
    if not has_header:
        ci_name, ci_budget, ci_resp = 0, 1, 2
    data = rows[1:] if has_header else rows
    out = []
    for r in data:
        if not r:
            continue
        def cell(i):
            return r[i] if (i is not None and i < len(r) and r[i] is not None) else None
        name = str(cell(ci_name)).strip() if cell(ci_name) is not None else ""
        if not name:
            continue
        b = cell(ci_budget)
        budget = _to_float(str(b).replace(",", ""), 0.0) if b is not None else 0.0
        resp = str(cell(ci_resp)).strip() if cell(ci_resp) is not None else ""
        out.append({"name": name, "budget": budget, "responsible": resp})
    return out


@router.get("/projects/import", response_class=HTMLResponse)
def projects_import_form(request: Request, db: Session = Depends(get_db)):
    school = get_school(db)
    cur = current_plan_year(school)
    return templates.TemplateResponse("projects_import.html", {
        "request": request, "school": school, "rows": None,
        "year_label": plan_year_label(school), "years": _plan_years(db, cur),
        "fiscal_year": cur,
    })


@router.get("/projects/import/template.xlsx")
def projects_import_template():
    from openpyxl import Workbook
    import tempfile
    wb = Workbook(); ws = wb.active; ws.title = "โครงการ"
    ws.append(["ชื่อโครงการ", "งบประมาณ", "ฝ่าย/ผู้รับผิดชอบ"])
    ws.append(["โครงการพัฒนาการเรียนการสอน", 50000, "ฝ่ายบริหารงานวิชาการ"])
    ws.append(["โครงการกีฬาสี", 20000, "ฝ่ายบริหารงานทั่วไป"])
    for col, w in (("A", 40), ("B", 16), ("C", 28)):
        ws.column_dimensions[col].width = w
    path = Path(tempfile.gettempdir()) / "ตัวอย่างนำเข้าโครงการ.xlsx"
    wb.save(path)
    return serve_generated(path, _XLSX_MT)


@router.post("/projects/import")
async def projects_import_preview(request: Request, db: Session = Depends(get_db),
                                  file: UploadFile = File(...), plan_year: str = Form("")):
    school = get_school(db)
    py = _to_int(plan_year, current_plan_year(school))
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/projects/import?err=type", status_code=303)
    content = await file.read()
    try:
        rows = _parse_projects_xlsx(content)
    except Exception:
        return RedirectResponse("/projects/import?err=read", status_code=303)
    import json
    # ทำเครื่องหมายแถวที่ชื่อซ้ำกับโครงการเดิมในปีเดียวกัน (จะอัปเดตงบ ไม่สร้างซ้ำ)
    existing = {p.name for p in db.query(Project).filter(Project.plan_year == py).all()}
    for r in rows:
        r["dup"] = r["name"] in existing
    n_update = sum(1 for r in rows if r["dup"])
    return templates.TemplateResponse("projects_import.html", {
        "request": request, "school": school, "rows": rows,
        "year_label": plan_year_label(school), "fiscal_year": py,
        "years": _plan_years(db, py), "payload": json.dumps(rows, ensure_ascii=False),
        "total_budget": sum(r["budget"] for r in rows),
        "n_update": n_update, "n_new": len(rows) - n_update,
    })


@router.post("/projects/import/confirm")
def projects_import_confirm(db: Session = Depends(get_db),
                            payload: str = Form(""), plan_year: str = Form("")):
    import json
    py = _to_int(plan_year, current_plan_year(get_school(db)))
    try:
        rows = json.loads(payload or "[]")
    except Exception:
        rows = []
    # โครงการเดิมในปีนี้ (ไว้เช็คชื่อซ้ำ -> อัปเดตงบแทนสร้างใหม่)
    existing = {p.name: p for p in db.query(Project).filter(Project.plan_year == py).all()}
    added = updated = 0
    for r in rows:
        name = (r.get("name") or "").strip()
        if not name:
            continue
        budget = _to_float(str(r.get("budget", 0)), 0.0)
        resp = (r.get("responsible") or "").strip()
        if name in existing:
            p = existing[name]
            p.budget = budget
            if resp:
                p.responsible = resp
            updated += 1
        else:
            p = Project(name=name, budget=budget, responsible=resp,
                        plan_year=py, active=True)
            db.add(p)
            existing[name] = p
            added += 1
    db.commit()
    return RedirectResponse(f"/projects?year={py}&added={added}&updated={updated}", status_code=303)


@router.get("/projects/{pid}", response_class=HTMLResponse)
def project_detail(pid: int, request: Request, db: Session = Depends(get_db)):
    p = db.get(Project, pid)
    if not p:
        return RedirectResponse("/projects", status_code=303)
    from app.services.budget import project_procurements
    procs = project_procurements(p)
    disb = (db.query(DisburseMemo).filter(DisburseMemo.project_id == pid)
            .order_by(DisburseMemo.id.desc()).all())
    next_seq = max([r.seq or 0 for r in p.revisions], default=0) + 1
    return templates.TemplateResponse("project_detail.html", {
        "request": request, "school": get_school(db), "p": p,
        "procs": procs, "disb": disb, "next_seq": next_seq,
        "year_label": plan_year_label(get_school(db)),
        "departments": db.query(Department).order_by(Department.name).all(),
        "years": _plan_years(db, p.plan_year or current_plan_year(get_school(db))),
    })


@router.post("/projects/{pid}/revise")
def project_revise(pid: int, db: Session = Depends(get_db), amount: str = Form("0"),
                   date: str = Form(""), reason: str = Form("")):
    p = db.get(Project, pid)
    if p:
        seq = max([r.seq or 0 for r in p.revisions], default=0) + 1
        db.add(ProjectBudgetRevision(project_id=pid, seq=seq, amount=_to_float(amount, 0.0),
                                     date=parse_be_date(date) or datetime.now(), reason=reason.strip()))
        db.commit()
    return RedirectResponse(f"/projects/{pid}?saved=1", status_code=303)


@router.post("/projects/{pid}/update")
def project_update(pid: int, db: Session = Depends(get_db), name: str = Form(...),
                   responsible: str = Form(""), plan_year: str = Form(""), active: str = Form("")):
    p = db.get(Project, pid)
    if p:
        p.name = name.strip()
        p.responsible = responsible.strip()
        p.plan_year = _to_int(plan_year, p.plan_year)
        p.active = (active == "1")
        db.commit()
    return RedirectResponse(f"/projects/{pid}?saved=1", status_code=303)


@router.post("/projects/{pid}/set-year")
def project_set_year(pid: int, db: Session = Depends(get_db),
                     plan_year: str = Form(""), back: str = Form("")):
    """เปลี่ยนปีของโครงการเดียว (จากตารางหรือหน้ารายละเอียด)"""
    p = db.get(Project, pid)
    if p:
        p.plan_year = _to_int(plan_year, p.plan_year)
        db.commit()
        py = p.plan_year
    else:
        py = current_plan_year(get_school(db))
    return RedirectResponse(back or f"/projects?year={py}", status_code=303)


@router.post("/projects/{pid}/set-responsible")
def project_set_responsible(pid: int, db: Session = Depends(get_db),
                            responsible: str = Form(""), back: str = Form("")):
    """เปลี่ยนฝ่าย/ผู้รับผิดชอบของโครงการเดียว (เรียกแบบ AJAX จากตาราง ไม่ reload หน้า)"""
    p = db.get(Project, pid)
    if p:
        p.responsible = responsible.strip()
        db.commit()
    # back มีค่า = ฟอร์มแบบเดิม (ไม่มี JS) ให้ redirect; ปกติเรียกผ่าน fetch ตอบ 204
    if back:
        return RedirectResponse(back, status_code=303)
    return Response(status_code=204)


@router.post("/projects/assign-year")
def projects_assign_year(db: Session = Depends(get_db),
                         plan_year: str = Form(""), only_legacy: str = Form("1")):
    """กำหนดปีให้หลายโครงการพร้อมกัน (ค่าเริ่มต้น = เฉพาะที่ยังไม่ระบุปี)"""
    py = _to_int(plan_year, current_plan_year(get_school(db)))
    q = db.query(Project)
    if only_legacy == "1":
        q = q.filter(Project.plan_year.is_(None))
    for p in q.all():
        p.plan_year = py
    db.commit()
    return RedirectResponse(f"/projects?year={py}", status_code=303)


@router.post("/projects/{pid}/delete")
def project_delete(pid: int, db: Session = Depends(get_db)):
    p = db.get(Project, pid)
    if p:
        db.delete(p); db.commit()
    return RedirectResponse("/projects", status_code=303)


@router.get("/assets/export.xlsx")
def assets_export(db: Session = Depends(get_db)):
    from app.services.asset_export import export_asset_register
    assets = db.query(Asset).order_by(Asset.id).all()
    path = export_asset_register(assets)
    return serve_generated(path, _XLSX_MT)


@router.get("/assets/form.xlsx")
def assets_form_export(db: Session = Depends(get_db)):
    """ออกแบบฟอร์มทะเบียนคุมทรัพย์สิน (แบบ 2) — การ์ดต่อ 1 ครุภัณฑ์"""
    from app.services.asset_export import export_asset_cards
    assets = db.query(Asset).order_by(Asset.id).all()
    path = export_asset_cards(assets, get_school(db))
    return serve_generated(path, _XLSX_MT)


@router.get("/assets", response_class=HTMLResponse)
def assets_page(request: Request, db: Session = Depends(get_db)):
    assets = db.query(Asset).order_by(Asset.id.desc()).all()
    total_cost = sum(a.cost or 0 for a in assets)
    total_nbv = sum(net_book_value(a.cost, a.salvage_value, a.useful_life,
                                   a.acquired_date) for a in assets)
    return templates.TemplateResponse("assets.html", {
        "request": request, "assets": assets, "categories": CATEGORIES,
        "category_life": CATEGORY_LIFE, "total_cost": total_cost, "total_nbv": total_nbv,
    })


@router.get("/assets/audit", response_class=HTMLResponse)
def asset_audit_page(request: Request, db: Session = Depends(get_db)):
    school = get_school(db)
    live = db.query(Asset).filter(Asset.status != "จำหน่ายแล้ว").count()
    return templates.TemplateResponse("asset_audit.html", {
        "request": request, "school": school,
        "year": current_fiscal_year(), "live_count": live,
        "today_input": be_date_input(datetime.now()),
    })


@router.post("/assets/audit")
async def asset_audit_generate(request: Request, db: Session = Depends(get_db)):
    from app.services.asset_audit_doc import render_audit_bundle
    form = await request.form()
    members = []
    for i in range(1, 6):
        nm = (form.get(f"m{i}_name") or "").strip()
        if nm:
            members.append({"name": nm,
                            "position": (form.get(f"m{i}_pos") or "ครู").strip(),
                            "role": (form.get(f"m{i}_role") or "กรรมการ").strip()})
    ctx = {
        "year": _to_int(form.get("year"), current_fiscal_year()),
        "date": parse_be_date(form.get("date") or ""),
        "memo_no": (form.get("memo_no") or "").strip(),
        "order_no": (form.get("order_no") or "").strip(),
        "result_memo_no": (form.get("result_memo_no") or "").strip(),
        "damaged_count": (form.get("damaged_count") or "").strip(),
        "members": members,
    }
    assets = db.query(Asset).filter(Asset.status != "จำหน่ายแล้ว").order_by(Asset.asset_code, Asset.id).all()
    path = render_audit_bundle(get_school(db), ctx, assets)
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


def _asset_from_form(asset: Asset, form) -> None:
    asset.asset_code = (form.get("asset_code") or "").strip()
    asset.name = (form.get("name") or "").strip()
    asset.category = form.get("category") or "ครุภัณฑ์สำนักงาน"
    asset.acquired_date = parse_be_date(form.get("acquired_date"))
    asset.cost = _to_float(form.get("cost"), 0.0)
    asset.useful_life = _to_int(form.get("useful_life"), CATEGORY_LIFE.get(asset.category, 10))
    asset.salvage_value = _to_float(form.get("salvage_value"), 1.0)
    asset.location = (form.get("location") or "").strip()
    asset.funding_source = (form.get("funding_source") or "").strip()
    asset.vendor_name = (form.get("vendor_name") or "").strip()
    asset.note = (form.get("note") or "").strip()
    asset.status = form.get("status") or "ใช้งาน"
    # ฟิลด์ตามแบบฟอร์มทะเบียนคุมทรัพย์สิน
    asset.brand_model = (form.get("brand_model") or "").strip()
    asset.vendor_address = (form.get("vendor_address") or "").strip()
    asset.fund_type = form.get("fund_type") or "เงินงบประมาณ"
    asset.acquire_method = form.get("acquire_method") or "วิธีเฉพาะเจาะจง"
    asset.doc_ref = (form.get("doc_ref") or "").strip()
    asset.quantity = _to_float(form.get("quantity"), 1) or 1
    asset.unit = (form.get("unit") or "หน่วย").strip() or "หน่วย"


@router.post("/assets")
async def asset_add(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    a = Asset()
    _asset_from_form(a, form)
    if a.name:
        db.add(a); db.commit()
    return RedirectResponse("/assets", status_code=303)


@router.post("/assets/{asset_id}/update")
async def asset_update(asset_id: int, request: Request, db: Session = Depends(get_db)):
    a = db.get(Asset, asset_id)
    if a:
        form = await request.form()
        _asset_from_form(a, form)
        db.commit()
    return RedirectResponse("/assets?saved=1", status_code=303)


@router.post("/assets/{asset_id}/delete")
def asset_delete(asset_id: int, db: Session = Depends(get_db)):
    a = db.get(Asset, asset_id)
    if a:
        db.delete(a); db.commit()
    return RedirectResponse("/assets", status_code=303)


@router.get("/assets/dispose", response_class=HTMLResponse)
def assets_dispose_page(request: Request, db: Session = Depends(get_db)):
    """หน้าเลือกครุภัณฑ์เพื่อขออนุมัติจำหน่าย (เลือกหลายรายการพร้อมกัน)"""
    active = (db.query(Asset).filter(Asset.status != "จำหน่ายแล้ว")
              .order_by(Asset.id.desc()).all())
    disposed = (db.query(Asset).filter(Asset.status == "จำหน่ายแล้ว")
                .order_by(Asset.id.desc()).all())
    sug = suggest_doc_no(db, "memo", current_fiscal_year())
    return templates.TemplateResponse("assets_dispose.html", {
        "request": request, "active": active, "disposed": disposed, "sug_no": sug,
    })


@router.post("/assets/dispose")
async def assets_dispose_submit(request: Request, db: Session = Depends(get_db)):
    """อัปเดตสถานะครุภัณฑ์ที่เลือกเป็น 'จำหน่ายแล้ว' + ออกเอกสารบันทึกข้อความขออนุมัติจำหน่าย"""
    from app.services.asset_dispose_doc import render_asset_disposal
    form = await request.form()
    ids = form.getlist("asset_ids")
    if not ids:
        return RedirectResponse("/assets/dispose?err=none", status_code=303)
    doc_no = (form.get("doc_no") or "").strip()
    doc_date = parse_be_date(form.get("doc_date")) or datetime.now()
    method = (form.get("method") or "").strip()
    reason = (form.get("reason") or "").strip()
    note = (form.get("note") or "").strip()
    value = _to_float(form.get("dispose_value"), 0.0)
    school = get_school(db)
    assets = (db.query(Asset).filter(Asset.id.in_([int(i) for i in ids]))
              .order_by(Asset.id).all())
    if not assets:
        return RedirectResponse("/assets/dispose?err=none", status_code=303)
    # แบ่งมูลค่าที่ขายได้รวม ไปยังแต่ละรายการตามสัดส่วนราคาทุน (ถ้าระบุมูลค่ารวม)
    total_cost = sum(float(a.cost or 0) for a in assets) or 1.0
    for a in assets:
        a.status = "จำหน่ายแล้ว"
        a.disposed_date = doc_date
        a.dispose_method = method
        a.dispose_reason = reason
        a.dispose_doc_ref = doc_no
        a.dispose_value = round(value * (float(a.cost or 0) / total_cost), 2) if value else 0.0
    path = render_asset_disposal(assets, school, doc_no=doc_no, doc_date=doc_date,
                                 method=method, reason=reason, note=note)
    if doc_no:
        commit_doc_no(db, "memo", current_fiscal_year(), doc_no,
                      source="asset", subject="ขออนุมัติจำหน่ายครุภัณฑ์")
    db.commit()
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


@router.get("/assets/{asset_id}/schedule", response_class=HTMLResponse)
def asset_schedule(asset_id: int, request: Request, db: Session = Depends(get_db)):
    a = db.get(Asset, asset_id)
    if not a:
        return RedirectResponse("/assets", status_code=303)
    sched = depreciation_schedule(a.cost, a.salvage_value, a.useful_life, a.acquired_date)
    return templates.TemplateResponse("asset_schedule.html", {
        "request": request, "a": a, "schedule": sched,
        "annual": annual_depreciation(a.cost, a.salvage_value, a.useful_life),
        "school": get_school(db),
    })


# ============================================================
# เฟส 3.2 — บัญชีวัสดุ (รับ-จ่าย-คงเหลือ)
# ============================================================
@router.get("/materials", response_class=HTMLResponse)
def materials_page(request: Request, db: Session = Depends(get_db)):
    items = db.query(MaterialItem).order_by(MaterialItem.name).all()
    return templates.TemplateResponse("materials.html", {
        "request": request, "items": items,
    })


@router.post("/materials")
def material_add(db: Session = Depends(get_db), name: str = Form(...),
                 unit: str = Form("หน่วย"), min_stock: str = Form("0"),
                 opening: str = Form("0")):
    if name.strip():
        it = MaterialItem(name=name.strip(), unit=unit.strip() or "หน่วย",
                          min_stock=_to_float(min_stock, 0.0))
        db.add(it); db.flush()
        op = _to_float(opening, 0.0)
        if op > 0:                       # มียอดเริ่มต้น -> ลงรับเป็น "ยอดยกมา"
            db.add(MaterialTxn(material_id=it.id, kind="in", qty=op,
                               date=datetime.now(), note="ยอดยกมา"))
        db.commit()
    return RedirectResponse("/materials", status_code=303)


@router.post("/materials/bulk")
async def materials_bulk_add(request: Request, db: Session = Depends(get_db)):
    """เพิ่มวัสดุหลายรายการพร้อมกัน (จากการสแกนรูป/กรอกในตาราง)"""
    form = await request.form()
    names = form.getlist("m_name")
    units, qtys, mins = form.getlist("m_unit"), form.getlist("m_qty"), form.getlist("m_min")
    n = 0
    for i, nm in enumerate(names):
        nm = (nm or "").strip()
        if not nm:
            continue
        it = MaterialItem(name=nm,
                          unit=(units[i].strip() if i < len(units) and units[i] else "") or "หน่วย",
                          min_stock=_to_float(mins[i] if i < len(mins) else "0", 0.0))
        db.add(it); db.flush()
        op = _to_float(qtys[i] if i < len(qtys) else "0", 0.0)
        if op > 0:
            db.add(MaterialTxn(material_id=it.id, kind="in", qty=op,
                               date=datetime.now(), note="ยอดยกมา"))
        n += 1
    db.commit()
    return RedirectResponse(f"/materials?added={n}", status_code=303)


@router.post("/materials/{item_id}/delete")
def material_delete(item_id: int, db: Session = Depends(get_db)):
    it = db.get(MaterialItem, item_id)
    if it:
        db.delete(it); db.commit()
    return RedirectResponse("/materials", status_code=303)


@router.get("/materials/{item_id}", response_class=HTMLResponse)
def material_ledger(item_id: int, request: Request, db: Session = Depends(get_db)):
    it = db.get(MaterialItem, item_id)
    if not it:
        return RedirectResponse("/materials", status_code=303)
    # คำนวณยอดคงเหลือสะสมแต่ละแถว (เรียงตามวันที่/ไอดี)
    rows = []
    bal = 0.0
    for t in sorted(it.txns, key=lambda x: (x.date or datetime.min, x.id)):
        if t.kind == "in":
            bal += (t.qty or 0)
        else:
            bal -= (t.qty or 0)
        rows.append({"t": t, "balance": bal})
    return templates.TemplateResponse("material_ledger.html", {
        "request": request, "item": it, "rows": rows, "balance": bal,
    })


@router.post("/materials/{item_id}/txn")
def material_txn_add(item_id: int, db: Session = Depends(get_db),
                     kind: str = Form("in"), qty: str = Form("0"),
                     unit_price: str = Form("0"), date: str = Form(""),
                     ref: str = Form(""), note: str = Form("")):
    it = db.get(MaterialItem, item_id)
    if it:
        db.add(MaterialTxn(
            material_id=it.id, kind=("out" if kind == "out" else "in"),
            qty=_to_float(qty, 0.0), unit_price=_to_float(unit_price, 0.0),
            date=parse_be_date(date) or datetime.now(),
            ref=ref.strip(), note=note.strip(),
        ))
        db.commit()
    return RedirectResponse(f"/materials/{item_id}", status_code=303)


@router.post("/materials/txn/{txn_id}/delete")
def material_txn_delete(txn_id: int, db: Session = Depends(get_db)):
    t = db.get(MaterialTxn, txn_id)
    mid = t.material_id if t else None
    if t:
        db.delete(t); db.commit()
    return RedirectResponse(f"/materials/{mid}" if mid else "/materials", status_code=303)


# ============================================================
# เฟส 3.3 — ใบเบิกวัสดุ
# ============================================================
@router.get("/requisitions", response_class=HTMLResponse)
def requisitions_page(request: Request, db: Session = Depends(get_db)):
    reqs = db.query(Requisition).order_by(Requisition.id.desc()).all()
    return templates.TemplateResponse("requisitions.html", {
        "request": request, "reqs": reqs,
    })


@router.get("/requisitions/new", response_class=HTMLResponse)
def requisition_new(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse("requisition_form.html", {
        "request": request, "items": db.query(MaterialItem).order_by(MaterialItem.name).all(),
        "persons": db.query(Person).order_by(Person.name).all(),
        "departments": db.query(Department).order_by(Department.name).all(),
        "today_thai": be_date_input(datetime.now()),
    })


@router.post("/requisitions/new")
async def requisition_create(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    req = Requisition(
        req_no=(form.get("req_no") or "").strip(),
        date=parse_be_date(form.get("date")) or datetime.now(),
        requester=(form.get("requester") or "").strip(),
        department=(form.get("department") or "").strip(),
        purpose=(form.get("purpose") or "").strip(),
    )
    names = form.getlist("item_name")
    qtys = form.getlist("item_qty")
    units = form.getlist("item_unit")
    mids = form.getlist("item_mid")
    for i, nm in enumerate(names):
        nm = (nm or "").strip()
        if not nm:
            continue
        req.items.append(RequisitionItem(
            name=nm, qty=_to_float(qtys[i] if i < len(qtys) else 0, 0.0),
            unit=(units[i] if i < len(units) else "") or "หน่วย",
            material_id=_to_int(mids[i], None) if i < len(mids) and mids[i] else None,
        ))
    if req.items:
        db.add(req); db.commit(); db.refresh(req)
        return RedirectResponse(f"/requisitions/{req.id}", status_code=303)
    return RedirectResponse("/requisitions/new", status_code=303)


@router.get("/requisitions/{req_id}", response_class=HTMLResponse)
def requisition_detail(req_id: int, request: Request, db: Session = Depends(get_db)):
    req = db.get(Requisition, req_id)
    if not req:
        return RedirectResponse("/requisitions", status_code=303)
    # ตรวจยอดคงเหลือพอจ่ายไหม (เฉพาะรายการที่ผูกกับวัสดุในระบบ)
    stock = {}
    for it in req.items:
        if it.material_id:
            m = db.get(MaterialItem, it.material_id)
            stock[it.id] = material_balance(m) if m else None
    return templates.TemplateResponse("requisition_detail.html", {
        "request": request, "req": req, "school": get_school(db), "stock": stock,
    })


@router.post("/requisitions/{req_id}/issue")
def requisition_issue(req_id: int, db: Session = Depends(get_db)):
    """ยืนยันจ่ายวัสดุ: สร้างรายการจ่ายออกในบัญชีวัสดุ + เปลี่ยนสถานะ"""
    req = db.get(Requisition, req_id)
    if req and req.status != "จ่ายแล้ว":
        for it in req.items:
            if it.material_id:
                db.add(MaterialTxn(
                    material_id=it.material_id, kind="out", qty=it.qty or 0,
                    date=req.date or datetime.now(),
                    ref=f"ใบเบิก {req.req_no or req.id}", requisition_id=req.id,
                    note=f"เบิกโดย {req.requester}",
                ))
        req.status = "จ่ายแล้ว"
        db.commit()
    return RedirectResponse(f"/requisitions/{req_id}?issued=1", status_code=303)


@router.post("/requisitions/{req_id}/delete")
def requisition_delete(req_id: int, db: Session = Depends(get_db)):
    req = db.get(Requisition, req_id)
    if req:
        # ลบรายการจ่ายออกที่ผูกกับใบเบิกนี้ด้วย (คืนสต๊อก)
        db.query(MaterialTxn).filter(MaterialTxn.requisition_id == req_id).delete()
        db.delete(req); db.commit()
    return RedirectResponse("/requisitions", status_code=303)


@router.get("/requisitions/{req_id}/print")
def requisition_print(req_id: int, db: Session = Depends(get_db)):
    from app.services.req_doc import render_requisition
    req = db.get(Requisition, req_id)
    if not req:
        return RedirectResponse("/requisitions", status_code=303)
    path = render_requisition(req, get_school(db))
    return serve_generated(path, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")


# ============================================================
# เฟส 3.4 — นำเข้าครุภัณฑ์/วัสดุ จากเรื่องจัดซื้อ (อัตโนมัติ)
# ============================================================
@router.get("/procurement/{proc_id}/to-register", response_class=HTMLResponse)
def to_register_page(proc_id: int, request: Request, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    # เดาประเภทให้: ราคา/หน่วย ≥ เกณฑ์ครุภัณฑ์ (10,000) -> ครุภัณฑ์, ไม่งั้น -> วัสดุ
    ASSET_THRESHOLD = 10000
    rows = []
    for it in proc.items:
        guess = "asset" if (it.unit_price or 0) >= ASSET_THRESHOLD else "material"
        rows.append({"it": it, "guess": guess})
    return templates.TemplateResponse("to_register.html", {
        "request": request, "p": proc, "rows": rows,
        "categories": CATEGORIES, "asset_threshold": ASSET_THRESHOLD,
    })


@router.post("/procurement/{proc_id}/to-register")
async def to_register_save(proc_id: int, request: Request, db: Session = Depends(get_db)):
    proc = db.get(Procurement, proc_id)
    if not proc:
        return RedirectResponse("/procurement", status_code=303)
    form = await request.form()
    vendor_name = proc.vendor.name if proc.vendor else ""
    n_asset = n_material = 0
    for idx, it in enumerate(proc.items):
        dest = form.get(f"dest_{idx}")        # asset / material / skip
        if dest == "asset":
            qty = int(it.quantity or 1)
            category = form.get(f"cat_{idx}") or "ครุภัณฑ์สำนักงาน"
            life = CATEGORY_LIFE.get(category, 10)
            # ครุภัณฑ์ 1 ชิ้น = 1 ระเบียน (แยกตามจำนวน)
            for _ in range(max(qty, 1)):
                db.add(Asset(
                    name=it.name, category=category, cost=it.unit_price or 0,
                    useful_life=life, salvage_value=1.0,
                    acquired_date=proc.order_date or proc.request_date,
                    funding_source=proc.budget_source or "", vendor_name=vendor_name,
                    procurement_id=proc.id,
                ))
                n_asset += 1
        elif dest == "material":
            # หาวัสดุชื่อเดียวกัน ถ้าไม่มีสร้างใหม่ แล้วบันทึกรับเข้า
            m = db.query(MaterialItem).filter(MaterialItem.name == it.name).first()
            if not m:
                m = MaterialItem(name=it.name, unit=it.unit or "หน่วย")
                db.add(m); db.flush()
            db.add(MaterialTxn(
                material_id=m.id, kind="in", qty=it.quantity or 0,
                unit_price=it.unit_price or 0,
                date=proc.order_date or proc.request_date or datetime.now(),
                ref=f"เรื่องจัดซื้อ {proc.memo_no or proc.id}",
            ))
            n_material += 1
    db.commit()
    return RedirectResponse(
        f"/procurement/{proc_id}?registered={n_asset}a{n_material}m", status_code=303)
