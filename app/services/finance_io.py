# -*- coding: utf-8 -*-
"""
finance_io.py - นำเข้าข้อมูลการเงินจาก Excel (สำหรับโรงเรียนที่มีทะเบียนเดิม)
ไฟล์เดียว 3 ชีต: บัญชีเงิน / รายการรับ-จ่าย / ใบเสร็จ-ใบสำคัญ
(แถว 1 = คำอธิบาย, แถว 2 = หัวตาราง, แถว 3+ = ข้อมูล)
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.models import FinanceAccount, FinanceTxn, Receipt
from app.thai_utils import current_fiscal_year, parse_be_date

THAI_FONT = "TH Sarabun New"
_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD = PatternFill("solid", fgColor="DCFCE7")     # เขียวอ่อน (ธีมการเงิน)
_NOTE = PatternFill("solid", fgColor="FFF2CC")

SHEETS = {
    "บัญชีเงิน": {
        "headers": ["ชื่อบัญชี", "ยอดยกมา", "หมายเหตุ"],
        "note": "กรอกรายการบัญชี/ประเภทเงินตั้งแต่แถวที่ 3 (เช่น เงินอุดหนุน, รายได้สถานศึกษา)",
        "widths": [30, 16, 36],
    },
    "รายการรับ-จ่าย": {
        "headers": ["บัญชี (ชื่อ)", "วันที่", "ประเภท (รับ/จ่าย)", "จำนวนเงิน", "หมวด", "อ้างอิง", "หมายเหตุ"],
        "note": "กรอกการเคลื่อนไหวเงินตั้งแต่แถวที่ 3 (วันที่ พ.ศ. เช่น 09/06/2569 · ประเภทใส่ รับ หรือ จ่าย)",
        "widths": [26, 16, 18, 16, 20, 20, 26],
    },
    "ใบเสร็จ-ใบสำคัญ": {
        "headers": ["เลขที่", "วันที่", "ประเภท (รับ/จ่าย)", "คู่สัญญา/ผู้รับเงิน", "จำนวนเงิน", "บัญชี (ชื่อ)", "หมายเหตุ"],
        "note": "กรอกทะเบียนใบเสร็จ/ใบสำคัญตั้งแต่แถวที่ 3",
        "widths": [16, 16, 18, 28, 16, 24, 24],
    },
}


def _style_head(c):
    c.font = Font(name=THAI_FONT, bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _BORDER; c.fill = _HEAD


def build_finance_template() -> str:
    wb = Workbook(); wb.remove(wb.active)
    for title, spec in SHEETS.items():
        ws = wb.create_sheet(title)
        last = chr(64 + len(spec["headers"]))
        ws.merge_cells(f"A1:{last}1")
        ws["A1"] = spec["note"]
        ws["A1"].font = Font(name=THAI_FONT, bold=True, size=12, color="9C6500")
        ws["A1"].fill = _NOTE
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col, h in enumerate(spec["headers"], start=1):
            _style_head(ws.cell(row=2, column=col, value=h))
        for i, w in enumerate(spec["widths"], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A3"
    out_dir = get_data_dir() / "documents"; out_dir.mkdir(exist_ok=True)
    path = out_dir / "เทมเพลตนำเข้าข้อมูลการเงิน.xlsx"
    wb.save(str(path))
    return str(path)


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _f(v) -> float:
    s = _s(v).replace(",", "")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def _kind(v) -> str:
    return "out" if ("จ่าย" in _s(v) or _s(v).lower() == "out") else "in"


def import_finance_workbook(file_bytes: bytes, db) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    fy = current_fiscal_year()
    summary = {}

    # แคชบัญชีตามชื่อ (สร้างใหม่ถ้ายังไม่มี เพื่อให้ผูกรายการได้)
    accounts = {a.name.strip(): a for a in db.query(FinanceAccount).all()}

    def _get_account(name: str):
        name = name.strip()
        if not name:
            return None
        a = accounts.get(name)
        if a is None:
            a = FinanceAccount(name=name)
            db.add(a); db.flush()
            accounts[name] = a
        return a

    if "บัญชีเงิน" in wb.sheetnames:
        added = 0
        for row in wb["บัญชีเงิน"].iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            name = _s(row[0])
            if not name:
                continue
            a = accounts.get(name)
            if a is None:
                a = FinanceAccount(name=name)
                db.add(a); db.flush()
                accounts[name] = a
            a.opening_balance = _f(row[1]) if len(row) > 1 else 0.0
            a.note = _s(row[2]) if len(row) > 2 else ""
            added += 1
        if added:
            summary["บัญชีเงิน"] = added

    if "รายการรับ-จ่าย" in wb.sheetnames:
        added = 0
        for row in wb["รายการรับ-จ่าย"].iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            acc = _get_account(_s(row[0]) if len(row) > 0 else "")
            if acc is None:
                continue
            db.add(FinanceTxn(
                account_id=acc.id, fiscal_year=fy,
                date=parse_be_date(_s(row[1])) if len(row) > 1 else None,
                kind=_kind(row[2]) if len(row) > 2 else "in",
                amount=_f(row[3]) if len(row) > 3 else 0.0,
                category=_s(row[4]) if len(row) > 4 else "",
                ref=_s(row[5]) if len(row) > 5 else "",
                note=_s(row[6]) if len(row) > 6 else "",
            ))
            added += 1
        if added:
            summary["รายการรับ-จ่าย"] = added

    if "ใบเสร็จ-ใบสำคัญ" in wb.sheetnames:
        added = 0
        for row in wb["ใบเสร็จ-ใบสำคัญ"].iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            receipt_no = _s(row[0]) if len(row) > 0 else ""
            party = _s(row[3]) if len(row) > 3 else ""
            if not (receipt_no or party):
                continue
            acc = _get_account(_s(row[5]) if len(row) > 5 else "")
            db.add(Receipt(
                fiscal_year=fy, receipt_no=receipt_no,
                date=parse_be_date(_s(row[1])) if len(row) > 1 else None,
                kind="จ่าย" if (len(row) > 2 and "จ่าย" in _s(row[2])) else "รับ",
                party=party,
                amount=_f(row[4]) if len(row) > 4 else 0.0,
                account_id=acc.id if acc else None,
                note=_s(row[6]) if len(row) > 6 else "",
            ))
            added += 1
        if added:
            summary["ใบเสร็จ-ใบสำคัญ"] = added

    db.commit()
    return summary
