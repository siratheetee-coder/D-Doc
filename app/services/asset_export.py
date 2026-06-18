# -*- coding: utf-8 -*-
"""
asset_export.py — ส่งออกทะเบียนครุภัณฑ์ (+ค่าเสื่อมราคา) เป็น Excel (.xlsx)
"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.thai_utils import thai_date
from app.services.asset_utils import (
    annual_depreciation, accumulated_depreciation, net_book_value, depreciation_schedule,
)

THAI_FONT = "TH Sarabun New"
_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD = PatternFill("solid", fgColor="E8F0FF")


def export_asset_register(assets) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "ทะเบียนครุภัณฑ์"
    today = date.today()

    ws.append([f"ทะเบียนคุมครุภัณฑ์ (คำนวณค่าเสื่อม ณ {thai_date(today)})"])
    ws.append([])
    headers = ["เลขครุภัณฑ์", "ชื่อครุภัณฑ์", "ประเภท", "วันที่ได้มา", "ราคาทุน",
               "อายุ (ปี)", "มูลค่าซาก", "ค่าเสื่อม/ปี", "ค่าเสื่อมสะสม", "มูลค่าสุทธิ",
               "สถานที่/ผู้รับผิดชอบ", "แหล่งเงิน", "ผู้ขาย", "สถานะ"]
    ws.append(headers)
    n = len(headers)
    HEAD_ROW = 3

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1)
    t.font = Font(name=THAI_FONT, bold=True, size=18)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, n + 1):
        c = ws.cell(row=HEAD_ROW, column=col)
        c.font = Font(name=THAI_FONT, bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER; c.fill = _HEAD

    tot_cost = tot_acc = tot_nbv = 0.0
    for a in assets:
        ann = annual_depreciation(a.cost, a.salvage_value, a.useful_life)
        acc = accumulated_depreciation(a.cost, a.salvage_value, a.useful_life, a.acquired_date, today)
        nbv = net_book_value(a.cost, a.salvage_value, a.useful_life, a.acquired_date, today)
        tot_cost += (a.cost or 0); tot_acc += acc; tot_nbv += nbv
        ws.append([
            a.asset_code or "", a.name, a.category or "", thai_date(a.acquired_date) if a.acquired_date else "",
            a.cost or 0, a.useful_life or 0, a.salvage_value or 0, ann, acc, nbv,
            a.location or "", a.funding_source or "", a.vendor_name or "", a.status or "",
        ])
    ws.append(["", "", "", "รวม", tot_cost, "", "", "", round(tot_acc, 2), round(tot_nbv, 2), "", "", "", ""])

    widths = [16, 26, 18, 14, 13, 9, 11, 11, 13, 13, 22, 16, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    money_cols = [5, 7, 8, 9, 10]
    for row in ws.iter_rows(min_row=HEAD_ROW + 1):
        for cell in row:
            cell.font = Font(name=THAI_FONT, size=14)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for mc in money_cols:
            row[mc - 1].number_format = "#,##0.00"
    # แถวรวม ตัวหนา
    last = ws.max_row
    for col in range(1, n + 1):
        ws.cell(row=last, column=col).font = Font(name=THAI_FONT, bold=True, size=14)

    out_dir = get_data_dir() / "documents"; out_dir.mkdir(exist_ok=True)
    path = out_dir / "ทะเบียนครุภัณฑ์.xlsx"
    wb.save(str(path))
    return str(path)


# ============================================================
# แบบฟอร์มที่ 2: ทะเบียนคุมทรัพย์สิน (การ์ดต่อ 1 ครุภัณฑ์)
# ============================================================
_FUND_OPTS = ["เงินงบประมาณ", "เงินนอกงบประมาณ", "เงินบริจาค/เงินช่วยเหลือ", "อื่น"]
_METHOD_OPTS = ["วิธีตลาดอิเล็กทรอนิกส์", "วิธีประกวดราคาอิเล็กทรอนิกส์",
                "วิธีคัดเลือก", "วิธีเฉพาะเจาะจง", "รับบริจาค"]


def _safe_sheet(name: str, used: set) -> str:
    for ch in ("[", "]", ":", "*", "?", "/", "\\"):
        name = name.replace(ch, "_")
    name = (name.strip() or "ครุภัณฑ์")[:28]
    base, n = name, 1
    while name in used:
        n += 1
        name = f"{base[:25]}_{n}"
    used.add(name)
    return name


def export_asset_cards(assets, school) -> str:
    """ออกไฟล์ Excel แบบ 'ทะเบียนคุมทรัพย์สิน' (แบบฟอร์มที่ 2) — 1 ชีตต่อ 1 ครุภัณฑ์"""
    wb = Workbook(); wb.remove(wb.active)

    def F(**kw):
        return Font(name=THAI_FONT, **kw)

    def setc(ws, coord, val, *, bold=False, size=14, align="left", wrap=False, box=False, fill=None, money=False):
        c = ws[coord]
        c.value = val
        c.font = F(bold=bold, size=size)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if box:
            c.border = _BORDER
        if fill:
            c.fill = fill
        if money:
            c.number_format = "#,##0.00"
        return c

    def chk(opt, val):
        return ("☑ " if (opt == val) else "☐ ") + opt   # ☑ / ☐

    used = set()
    for idx, a in enumerate(assets, 1):
        ws = wb.create_sheet(_safe_sheet(a.asset_code or a.name or f"ครุภัณฑ์{idx}", used))
        for i, w in enumerate([12, 11, 26, 7, 8, 13, 13, 10, 12, 13, 13, 13, 12], start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        ws.merge_cells("A1:M1"); setc(ws, "A1", "ทะเบียนคุมทรัพย์สิน", bold=True, size=18, align="center")
        ws.merge_cells("A2:M2")
        setc(ws, "A2", f"ส่วนราชการ สำนักงานคณะกรรมการการศึกษาขั้นพื้นฐาน    หน่วยงาน {school.name or ''}",
             align="center", size=13)

        def label(coord_l, lbl, coord_r_range, val):
            l = coord_l.split(":")[0]
            ws.merge_cells(coord_l); setc(ws, l, lbl, bold=True)
            r0 = coord_r_range.split(":")[0]
            ws.merge_cells(coord_r_range); setc(ws, r0, val)

        label("A4:B4", "ประเภท", "C4:F4", a.category or "")
        label("G4:H4", "หมายเลขครุภัณฑ์", "I4:M4", a.asset_code or "")
        label("A5:B5", "รายการ", "C5:F5", a.name or "")
        label("G5:H5", "ยี่ห้อ/รุ่น/ลักษณะเฉพาะ", "I5:M5", a.brand_model or "")
        label("A6:B6", "สถานที่ใช้งาน/ผู้รับผิดชอบ", "C6:M6", a.location or "")
        label("A7:B7", "ผู้ขาย/ผู้รับจ้าง/ผู้บริจาค", "C7:F7", a.vendor_name or "")
        label("G7:H7", "ที่อยู่", "I7:M7", a.vendor_address or "")
        label("A8:B8", "ประเภทเงิน", "C8:M8", "   ".join(chk(o, a.fund_type) for o in _FUND_OPTS))
        label("A9:B9", "วิธีการได้มา", "C9:M9", "   ".join(chk(o, a.acquire_method) for o in _METHOD_OPTS))

        headers = ["วัน/เดือน/ปี", "ที่เอกสาร", "รายการ", "จำนวน", "หน่วย", "ราคาต่อหน่วย",
                   "มูลค่ารวม", "อายุใช้งาน(ปี)", "อัตราค่าเสื่อม(%)", "ค่าเสื่อมประจำปี",
                   "ค่าเสื่อมสะสม", "มูลค่าสุทธิ", "หมายเหตุ"]
        hr = 11
        ws.row_dimensions[hr].height = 34
        for j, h in enumerate(headers, 1):
            setc(ws, ws.cell(row=hr, column=j).coordinate, h, bold=True, align="center",
                 wrap=True, box=True, fill=_HEAD, size=13)

        qty = a.quantity or 1
        cost = a.cost or 0
        per_unit = round(cost / qty, 2) if qty else cost
        life = a.useful_life or 0
        rate = round(100.0 / life, 2) if life else 0

        # แถวแรก = การได้มา
        r = hr + 1
        first = [thai_date(a.acquired_date) if a.acquired_date else "", a.doc_ref or "", a.name or "",
                 ("%g" % qty), a.unit or "หน่วย", per_unit, cost, ("%g" % life if life else ""),
                 ("%g" % rate if rate else ""), "", "", cost, ""]
        for j, v in enumerate(first, 1):
            money = j in (6, 7, 12)
            align = "right" if (money or j in (4, 8, 9)) else "left"
            setc(ws, ws.cell(row=r, column=j).coordinate, v, align=align, box=True, money=money)

        # แถวค่าเสื่อมรายปีงบประมาณ
        for row in depreciation_schedule(a.cost, a.salvage_value, a.useful_life, a.acquired_date):
            r += 1
            for j in range(1, 14):
                setc(ws, ws.cell(row=r, column=j).coordinate, "", box=True)
            setc(ws, ws.cell(row=r, column=1).coordinate, "30 ก.ย. %d" % row["fy"], box=True, align="center")
            setc(ws, ws.cell(row=r, column=10).coordinate, row["dep"], box=True, align="right", money=True)
            setc(ws, ws.cell(row=r, column=11).coordinate, row["acc"], box=True, align="right", money=True)
            setc(ws, ws.cell(row=r, column=12).coordinate, row["nbv"], box=True, align="right", money=True)

    if not wb.sheetnames:
        wb.create_sheet("ว่าง")["A1"] = "ยังไม่มีครุภัณฑ์"

    out_dir = get_data_dir() / "documents"; out_dir.mkdir(exist_ok=True)
    path = out_dir / "ทะเบียนคุมทรัพย์สิน(แบบ2).xlsx"
    wb.save(str(path))
    return str(path)
