"""
register_export.py
------------------
ส่งออก 'ทะเบียนคุมการจัดซื้อจัดจ้าง' เป็นไฟล์ Excel (.xlsx)
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.thai_utils import thai_date

THAI_FONT = "TH Sarabun New"


# ชื่อเล่มทะเบียนตามชนิด (kind)
_REGISTER_LABEL = {
    "buy": "ทะเบียนคุมการจัดซื้อ",
    "hire": "ทะเบียนคุมการจัดจ้าง",
    "all": "ทะเบียนคุมการจัดซื้อจัดจ้าง",
}


def export_register(procurements, fiscal_year: int, kind: str = "all") -> str:
    """สร้างไฟล์ Excel ทะเบียนคุม คืนค่าที่อยู่ไฟล์
    kind: all = รวม, buy = เฉพาะจัดซื้อ, hire = เฉพาะจัดจ้าง (แยกเล่ม)"""
    label = _REGISTER_LABEL.get(kind, _REGISTER_LABEL["all"])
    wb = Workbook()
    ws = wb.active
    ws.title = f"{label} {fiscal_year}"[:31]   # ชื่อชีต Excel จำกัด 31 ตัว

    # หัวกระดาษ (ชื่อเล่ม + ปีงบ) เหนือหัวตาราง
    ws.append([f"{label}  ประจำปีงบประมาณ {fiscal_year}"])
    ws.append([])

    headers = ["เลขที่", "วันที่", "เรื่อง", "ประเภท", "วิธี",
               "วงเงิน (บาท)", "ผู้ขาย/ผู้รับจ้าง", "หมายเหตุ"]
    ws.append(headers)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    n_col = len(headers)
    HEADER_ROW = 3   # แถวที่ 1 = ชื่อเล่ม, 2 = ว่าง, 3 = หัวตาราง, 4+ = ข้อมูล

    # ชื่อเล่ม (แถว 1) ผสานเซลล์ + จัดกึ่งกลาง
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name=THAI_FONT, bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # หัวตาราง (แถว 3)
    for col in range(1, n_col + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.font = Font(name=THAI_FONT, bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = header_fill

    # เติมข้อมูล (เลขที่ = เลขใบสั่งซื้อ/จ้าง ถ้ายังไม่มีใช้เลขบันทึก)
    for p in procurements:
        ws.append([
            (p.order_no or "").strip() or p.doc_no or "-",
            thai_date(p.order_date or p.request_date),
            p.subject,
            p.proc_type,
            p.method,
            p.total_amount or 0,
            p.vendor.name if p.vendor else "-",
            "",   # หมายเหตุ (เว้นว่างไว้ให้เขียนเอง)
        ])

    # จัดรูปทุกเซลล์ข้อมูล + กำหนดความกว้างคอลัมน์
    widths = [10, 16, 36, 10, 14, 16, 26, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        for cell in row:
            cell.font = Font(name=THAI_FONT, size=14)
            cell.border = border
            # wrap_text: ข้อความยาว (เรื่อง/ผู้ขาย) ขึ้นบรรทัดใหม่ในช่อง ไม่ถูกตัด
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[5].number_format = "#,##0.00"  # คอลัมน์วงเงิน

    out_dir = get_data_dir() / "documents"
    out_dir.mkdir(exist_ok=True)
    suffix = {"buy": "จัดซื้อ", "hire": "จัดจ้าง"}.get(kind, "จัดซื้อจัดจ้าง")
    path = out_dir / f"ทะเบียนคุม{suffix}_ปีงบ{fiscal_year}.xlsx"
    wb.save(str(path))
    return str(path)


_THAI_MONTHS = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
                "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

# วิธีจัดหาที่ใช้เกณฑ์ราคาต่ำสุด (ถ้าไม่ใช่ ใช้เหตุผล "เฉพาะเจาะจง")
_REASON_DEFAULT = "เป็นผู้มีคุณสมบัติถูกต้องครบถ้วนและเสนอราคาเหมาะสมภายในวงเงินงบประมาณ"


def export_monthly_summary(procurements, fiscal_year: int, school_name: str = "") -> str:
    """สร้างแบบ สขร.1 (สรุปผลการดำเนินการจัดซื้อจัดจ้างในรอบเดือน) แยกชีตตามเดือน
    อ้างอิงเดือนจากวันที่ใบสั่งซื้อ/จ้าง (ถ้าไม่มีใช้วันที่รายงานขอซื้อ)"""
    # จัดกลุ่มตามเดือน (ค.ศ. ปี+เดือน) โดยอิงวันที่จริงของเอกสาร
    groups: dict = {}
    for p in procurements:
        dt = p.order_date or p.request_date
        if not dt:
            continue
        groups.setdefault((dt.year, dt.month), []).append(p)

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    headers = ["ลำดับที่", "งานที่จัดซื้อจัดจ้าง", "วงเงินที่จัดซื้อจัดจ้าง (บาท)",
               "ราคากลาง (บาท)", "วิธีซื้อหรือจ้าง", "รายชื่อผู้เสนอราคาและราคาที่เสนอ",
               "ผู้ได้รับการคัดเลือกและราคาที่ตกลงซื้อหรือจ้าง", "เหตุผลที่คัดเลือกโดยสรุป",
               "เลขที่และวันที่ของสัญญาหรือข้อตกลงในการซื้อหรือจ้าง"]
    widths = [7, 30, 15, 14, 14, 26, 26, 24, 22]
    n_col = len(headers)

    if not groups:   # ไม่มีข้อมูล -> ชีตว่างพร้อมหัวตาราง
        groups = {(fiscal_year - 543, 1): []}

    for (yr, mo) in sorted(groups.keys()):
        rows = groups[(yr, mo)]
        title_month = f"{_THAI_MONTHS[mo]} พ.ศ. {yr + 543}"
        ws = wb.create_sheet(f"{_THAI_MONTHS[mo]} {yr + 543}"[:31])
        ws.append([f"แบบสรุปผลการดำเนินการจัดซื้อจัดจ้างในรอบเดือน {title_month}"])
        ws.append([f"{school_name}   (แบบ สขร.1)"])
        ws.append([])
        ws.append(headers)
        hrow = 4
        for c in range(1, n_col + 1):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_col)
            cell = ws.cell(row=hrow, column=c)
            cell.font = Font(name=THAI_FONT, bold=True, size=13)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.fill = header_fill
        for cell, sz in ((ws.cell(1, 1), 16), (ws.cell(2, 1), 14)):
            cell.font = Font(name=THAI_FONT, bold=True, size=sz)
            cell.alignment = Alignment(horizontal="center")

        for i, p in enumerate(sorted(rows, key=lambda x: (x.order_date or x.request_date)), start=1):
            amt = p.total_amount or 0
            vname = p.vendor.name if p.vendor else "-"
            offer = f"{vname}  {amt:,.2f} บาท" if p.vendor else "-"
            contract = ((p.order_no or "").strip() or p.doc_no or "-")
            cdate = p.order_date or p.request_date
            contract_txt = f"{contract}\nลว. {thai_date(cdate)}" if cdate else contract
            ws.append([i, p.subject, amt, amt, p.method, offer, offer,
                       _REASON_DEFAULT, contract_txt])

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        for row in ws.iter_rows(min_row=hrow + 1):
            for cell in row:
                cell.font = Font(name=THAI_FONT, size=13)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row[0].alignment = Alignment(horizontal="center", vertical="top")
            row[2].number_format = "#,##0.00"
            row[3].number_format = "#,##0.00"

    out_dir = get_data_dir() / "documents"
    out_dir.mkdir(exist_ok=True)
    path = out_dir / f"สขร1_สรุปผลจัดซื้อจัดจ้างรายเดือน_ปีงบ{fiscal_year}.xlsx"
    wb.save(str(path))
    return str(path)
