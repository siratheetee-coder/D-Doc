# -*- coding: utf-8 -*-
"""
bulk_io.py
----------
นำเข้าข้อมูลตั้งต้นทีละมาก ๆ จากไฟล์ Excel (.xlsx) ไฟล์เดียว หลายชีต:
  โรงเรียน / บุคลากร / ฝ่าย-งาน / ผู้ขาย

แนวคิด: โหลดเทมเพลตที่มีหัวตาราง + คำอธิบายไปกรอกใน Excel แล้วอัปโหลดกลับ
- ชีตข้อมูลหลัก (บุคลากร/ฝ่าย/ผู้ขาย): เพิ่มเข้าระบบ ข้ามชื่อซ้ำ
- ชีตโรงเรียน: อัปเดตข้อมูลโรงเรียน (มีรายการเดียว)

ทุกชีต: แถว 1 = คำอธิบาย, แถว 2 = หัวตาราง, แถว 3 เป็นต้นไป = ข้อมูล
"""
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.models import School, Person, Department, Vendor, Student
from app.thai_utils import parse_be_date

_SEX_MAP = {"ช": "M", "ชาย": "M", "m": "M", "M": "M",
            "ญ": "F", "หญิง": "F", "f": "F", "F": "F"}

THAI_FONT = "TH Sarabun New"

# ---- นิยามแต่ละชีตข้อมูลหลัก: (ชื่อชีต, [หัวคอลัมน์], คำอธิบาย, ความกว้างคอลัมน์) ----
MASTER_SHEETS = {
    "บุคลากร": {
        "headers": ["ชื่อ-นามสกุล", "ประเภท (ครู/ผู้บริหาร/ธุรการ/นักการ)", "ตำแหน่ง",
                    "วิทยฐานะ/ระดับ", "เลขบัตรประชาชน", "วันเกิด (วว/ดด/ปปปป)",
                    "วันบรรจุ (วว/ดด/ปปปป)", "เบอร์โทร", "อีเมล", "เงินเดือน"],
        "note": "กรอกรายชื่อครู/บุคลากรตั้งแต่แถวที่ 3 ลงไป (จำเป็นเฉพาะชื่อ คอลัมน์อื่นเว้นว่างได้)  เช่น  นายสมชาย ใจดี | ครู | ครู | ครู คศ.2 | 1234567890123 | 15/05/2530 | 01/05/2555 | 08x-xxxxxxx | a@b.com | 30000",
        "widths": [30, 22, 18, 18, 20, 20, 20, 16, 22, 14],
    },
    "ฝ่าย-งาน": {
        "headers": ["ชื่อฝ่าย/งาน"],
        "note": "กรอกชื่อฝ่าย/งานตั้งแต่แถวที่ 3 ลงไป  เช่น  ฝ่ายบริหารงานวิชาการ",
        "widths": [40],
    },
    "โครงการ": {
        "headers": ["ชื่อโครงการ", "ปี (พ.ศ.)", "งบประมาณ (บาท)", "ฝ่าย/ผู้รับผิดชอบ"],
        "note": "กรอกชื่อโครงการตั้งแต่แถวที่ 3 ลงไป  ปีเว้นว่างได้ (ระบบใช้ปีปัจจุบันให้)  เช่น  โครงการวันสำคัญทางวิชาการ | 2569 | 5000 | ฝ่ายบริหารงานวิชาการ",
        "widths": [40, 12, 18, 28],
    },
    "ผู้ขาย": {
        "headers": ["ชื่อร้าน/บริษัท", "ชื่อเจ้าของ/ผู้ลงนาม", "เลขประจำตัวผู้เสียภาษี",
                    "ที่อยู่", "เบอร์โทร", "เลขบัญชีธนาคาร"],
        "note": "กรอกข้อมูลผู้ขาย/ผู้รับจ้างตั้งแต่แถวที่ 3 ลงไป (จำเป็นเฉพาะชื่อร้าน คอลัมน์อื่นเว้นว่างได้)",
        "widths": [30, 24, 22, 36, 16, 22],
    },
    "นักเรียน": {
        "headers": ["ชื่อ-นามสกุล", "เพศ (ช/ญ)", "วันเกิด (วว/ดด/ปปปป)", "ระดับชั้น", "เลขประจำตัว"],
        "note": "กรอกรายชื่อนักเรียนตั้งแต่แถวที่ 3 ลงไป (จำเป็นเฉพาะชื่อ) เพศใส่ ช หรือ ญ  เช่น  เด็กชายสมชาย ใจดี | ช | 15/05/2562 | ป.1 | 10001",
        "widths": [30, 12, 22, 12, 14],
    },
}

# ---- ชีตโรงเรียน: (ป้ายกำกับ, ชื่อฟิลด์ใน School) แบบแนวตั้ง คีย์-ค่า ----
SCHOOL_FIELDS = [
    ("ชื่อโรงเรียน", "name"),
    ("ที่อยู่", "address"),
    ("อำเภอ", "district"),
    ("จังหวัด", "province"),
    ("ชื่อผู้อำนวยการ", "director_name"),
    ("ตำแหน่งผู้อำนวยการ", "director_position"),
    ("เจ้าหน้าที่พัสดุ", "officer_name"),
    ("หัวหน้าเจ้าหน้าที่พัสดุ", "head_officer_name"),
    ("อักษรนำเลขที่หนังสือ", "doc_prefix"),
]

_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
_LABEL_FILL = PatternFill("solid", fgColor="EDEDED")


def _style_header(cell):
    cell.font = Font(name=THAI_FONT, bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER
    cell.fill = _HEAD_FILL


def _style_note(cell):
    cell.font = Font(name=THAI_FONT, bold=True, size=12, color="9C6500")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.fill = _NOTE_FILL


def build_import_template() -> str:
    """สร้างไฟล์เทมเพลต Excel สำหรับกรอกข้อมูลตั้งต้น คืนค่าที่อยู่ไฟล์"""
    wb = Workbook()
    wb.remove(wb.active)   # ลบชีตเริ่มต้น

    # ===== ชีตโรงเรียน (คีย์-ค่า แนวตั้ง) =====
    ws = wb.create_sheet("โรงเรียน")
    ws.merge_cells("A1:B1")
    ws["A1"] = ("กรอกข้อมูลโรงเรียนในคอลัมน์ 'ข้อมูล' (แถวที่ 3 ลงไป) "
                "ถ้าเว้นว่างจะไม่ทับข้อมูลเดิม")
    _style_note(ws["A1"])
    for col, h in enumerate(["หัวข้อ", "ข้อมูล"], start=1):
        _style_header(ws.cell(row=2, column=col, value=h))
    for i, (label, _field) in enumerate(SCHOOL_FIELDS, start=3):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = Font(name=THAI_FONT, bold=True, size=14)
        c1.border = _BORDER
        c1.fill = _LABEL_FILL
        c2 = ws.cell(row=i, column=2)
        c2.font = Font(name=THAI_FONT, size=14)
        c2.border = _BORDER
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 48
    ws.row_dimensions[1].height = 30

    # ===== ชีตข้อมูลหลักอื่น ๆ =====
    for title, spec in MASTER_SHEETS.items():
        ws = wb.create_sheet(title)
        headers = spec["headers"]
        last_col = chr(64 + len(headers))
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = spec["note"]
        _style_note(ws["A1"])
        for col, h in enumerate(headers, start=1):
            _style_header(ws.cell(row=2, column=col, value=h))
        for i, w in enumerate(spec["widths"], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A3"   # ตรึงหัวตารางไว้

    out_dir = get_data_dir() / "documents"
    out_dir.mkdir(exist_ok=True)
    path = out_dir / "เทมเพลตนำเข้าข้อมูล.xlsx"
    wb.save(str(path))
    return str(path)


def _cell_str(v) -> str:
    """แปลงค่าเซลล์เป็นข้อความที่สะอาด (รองรับตัวเลข/None)"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _cell_float(v) -> float:
    """แปลงค่าเซลล์เป็นตัวเลข (รองรับ '5,000' / ว่าง -> 0)"""
    s = _cell_str(v).replace(",", "")
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


def import_workbook(file_bytes: bytes, db) -> dict:
    """อ่านไฟล์ Excel ที่อัปโหลด แล้วนำเข้าข้อมูล คืนสรุปผลแต่ละชีต
    คืน dict เช่น {'บุคลากร': {'added': 12, 'skipped': 2}, 'โรงเรียน': {'updated': 6}, ...}
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    summary = {}

    # ---- ชีตโรงเรียน ----
    if "โรงเรียน" in wb.sheetnames:
        ws = wb["โรงเรียน"]
        label_to_field = {lbl: f for lbl, f in SCHOOL_FIELDS}
        school = db.query(School).first()
        if school is None:
            school = School()
            db.add(school)
        updated = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 2:
                continue
            label = _cell_str(row[0])
            value = _cell_str(row[1])
            field = label_to_field.get(label)
            if field and value:
                setattr(school, field, value)
                updated += 1
        if updated:
            summary["โรงเรียน"] = {"updated": updated}

    # ---- ชีตบุคลากร ----
    if "บุคลากร" in wb.sheetnames:
        existing = {p.name for p in db.query(Person).all()}
        added = skipped = 0
        for row in wb["บุคลากร"].iter_rows(min_row=3, values_only=True):
            name = _cell_str(row[0]) if row else ""
            if not name:
                continue
            if name in existing:
                skipped += 1
                continue

            def _c(i):
                return _cell_str(row[i]) if len(row) > i else ""
            db.add(Person(
                name=name,
                person_type=_c(1) or "ครู",
                position=_c(2) or "ครู",
                rank=_c(3), id_card=_c(4),
                birthdate=parse_be_date(_c(5)), start_date=parse_be_date(_c(6)),
                phone=_c(7), email=_c(8),
                salary=_cell_float(row[9]) if len(row) > 9 else 0.0,
            ))
            existing.add(name)
            added += 1
        if added or skipped:
            summary["บุคลากร"] = {"added": added, "skipped": skipped}

    # ---- ชีตฝ่าย-งาน ----
    if "ฝ่าย-งาน" in wb.sheetnames:
        existing = {d.name for d in db.query(Department).all()}
        added = skipped = 0
        for row in wb["ฝ่าย-งาน"].iter_rows(min_row=3, values_only=True):
            name = _cell_str(row[0]) if row else ""
            if not name:
                continue
            if name in existing:
                skipped += 1
                continue
            db.add(Department(name=name))
            existing.add(name)
            added += 1
        if added or skipped:
            summary["ฝ่าย-งาน"] = {"added": added, "skipped": skipped}

    # ---- ชีตโครงการ (ชื่อ | ปี พ.ศ. | งบ | ฝ่าย) ----
    if "โครงการ" in wb.sheetnames:
        from app.models import Project
        from app.services.budget import current_plan_year
        school = db.query(School).first()
        default_py = current_plan_year(school) if school else None
        existing = {(p.name, p.plan_year) for p in db.query(Project).all()}
        added = skipped = 0
        for row in wb["โครงการ"].iter_rows(min_row=3, values_only=True):
            name = _cell_str(row[0]) if row else ""
            if not name:
                continue
            py = int(_cell_float(row[1])) if (len(row) > 1 and _cell_float(row[1])) else default_py
            if (name, py) in existing:
                skipped += 1
                continue
            db.add(Project(
                name=name, plan_year=py,
                budget=_cell_float(row[2]) if len(row) > 2 else 0.0,
                responsible=_cell_str(row[3]) if len(row) > 3 else "",
                active=True,
            ))
            existing.add((name, py))
            added += 1
        if added or skipped:
            summary["โครงการ"] = {"added": added, "skipped": skipped}

    # ---- ชีตผู้ขาย ----
    if "ผู้ขาย" in wb.sheetnames:
        existing = {v.name for v in db.query(Vendor).all()}
        added = skipped = 0
        for row in wb["ผู้ขาย"].iter_rows(min_row=3, values_only=True):
            name = _cell_str(row[0]) if row else ""
            if not name:
                continue
            if name in existing:
                skipped += 1
                continue
            db.add(Vendor(
                name=name,
                owner_name=_cell_str(row[1]) if len(row) > 1 else "",
                tax_id=_cell_str(row[2]) if len(row) > 2 else "",
                address=_cell_str(row[3]) if len(row) > 3 else "",
                phone=_cell_str(row[4]) if len(row) > 4 else "",
                bank_account=_cell_str(row[5]) if len(row) > 5 else "",
            ))
            existing.add(name)
            added += 1
        if added or skipped:
            summary["ผู้ขาย"] = {"added": added, "skipped": skipped}

    # ---- ชีตนักเรียน (ชื่อ | เพศ | วันเกิด | ชั้น | เลขประจำตัว) ----
    if "นักเรียน" in wb.sheetnames:
        existing = {s.name for s in db.query(Student).all()}
        added = skipped = 0
        for row in wb["นักเรียน"].iter_rows(min_row=3, values_only=True):
            name = _cell_str(row[0]) if row else ""
            if not name:
                continue
            if name in existing:
                skipped += 1
                continue
            db.add(Student(
                name=name,
                sex=_SEX_MAP.get(_cell_str(row[1]) if len(row) > 1 else "", ""),
                birthdate=parse_be_date(_cell_str(row[2])) if len(row) > 2 else None,
                level=_cell_str(row[3]) if len(row) > 3 else "",
                student_no=_cell_str(row[4]) if len(row) > 4 else "",
            ))
            existing.add(name)
            added += 1
        if added or skipped:
            summary["นักเรียน"] = {"added": added, "skipped": skipped}

    db.commit()
    return summary
