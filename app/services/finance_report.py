# -*- coding: utf-8 -*-
"""
finance_report.py — ส่งออกรายงานการเงินเป็น Excel (.xlsx)
2 ชีต: สรุปแยกบัญชี (ยอดยกมา/รับ/จ่าย/คงเหลือ) + สรุปรายเดือน (ตามปีงบประมาณ)
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.services.asset_utils import account_balance_year, opening_for

THAI_FONT = "TH Sarabun New"
_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD = PatternFill("solid", fgColor="DCFCE7")
_TOTAL = PatternFill("solid", fgColor="F1F5F9")

# เดือนเรียงตามปีงบประมาณไทย (เริ่ม ต.ค.)
_FY_MONTHS = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]
_MONTH_NAME = {1: "มกราคม", 2: "กุมภาพันธ์", 3: "มีนาคม", 4: "เมษายน", 5: "พฤษภาคม",
               6: "มิถุนายน", 7: "กรกฎาคม", 8: "สิงหาคม", 9: "กันยายน",
               10: "ตุลาคม", 11: "พฤศจิกายน", 12: "ธันวาคม"}


def _cell(ws, r, c, val, *, bold=False, align="left", money=False, fill=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name=THAI_FONT, bold=bold, size=14)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _BORDER
    if money:
        cell.number_format = "#,##0.00"
    if fill:
        cell.fill = fill
    return cell


def export_finance_report(accounts, txns, fiscal_year: int) -> str:
    wb = Workbook(); wb.remove(wb.active)

    # ---------- ชีต 1: สรุปแยกบัญชี ----------
    ws = wb.create_sheet("สรุปแยกบัญชี")
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value=f"รายงานสรุปการเงินแยกบัญชี  ประจำปีงบประมาณ {fiscal_year}")
    t.font = Font(name=THAI_FONT, bold=True, size=18)
    t.alignment = Alignment(horizontal="center", vertical="center")
    headers = ["บัญชี/ประเภทเงิน", "ยอดยกมา", "รวมรับ", "รวมจ่าย", "คงเหลือ"]
    for c, h in enumerate(headers, start=1):
        _cell(ws, 3, c, h, bold=True, align="center", fill=_HEAD)
    r = 4
    sum_open = sum_in = sum_out = sum_bal = 0.0
    for a in accounts:
        tin = sum(t.amount or 0 for t in a.txns if t.kind == "in" and t.fiscal_year == fiscal_year)
        tout = sum(t.amount or 0 for t in a.txns if t.kind == "out" and t.fiscal_year == fiscal_year)
        bal = account_balance_year(a, fiscal_year)
        _cell(ws, r, 1, a.name)
        _cell(ws, r, 2, opening_for(a, fiscal_year), money=True, align="right")
        _cell(ws, r, 3, tin, money=True, align="right")
        _cell(ws, r, 4, tout, money=True, align="right")
        _cell(ws, r, 5, bal, money=True, align="right")
        sum_open += opening_for(a, fiscal_year); sum_in += tin; sum_out += tout; sum_bal += bal
        r += 1
    _cell(ws, r, 1, "รวมทั้งสิ้น", bold=True, fill=_TOTAL)
    _cell(ws, r, 2, sum_open, bold=True, money=True, align="right", fill=_TOTAL)
    _cell(ws, r, 3, sum_in, bold=True, money=True, align="right", fill=_TOTAL)
    _cell(ws, r, 4, sum_out, bold=True, money=True, align="right", fill=_TOTAL)
    _cell(ws, r, 5, sum_bal, bold=True, money=True, align="right", fill=_TOTAL)
    for i, w in enumerate([34, 16, 16, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ---------- ชีต 2: สรุปรายเดือน ----------
    ws2 = wb.create_sheet("สรุปรายเดือน")
    ws2.merge_cells("A1:D1")
    t2 = ws2.cell(row=1, column=1, value=f"สรุปรับ-จ่ายรายเดือน  ประจำปีงบประมาณ {fiscal_year}")
    t2.font = Font(name=THAI_FONT, bold=True, size=18)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    for c, h in enumerate(["เดือน", "รับ", "จ่าย", "สุทธิ (รับ-จ่าย)"], start=1):
        _cell(ws2, 3, c, h, bold=True, align="center", fill=_HEAD)
    # รวมรายเดือน
    by_month = {m: [0.0, 0.0] for m in _FY_MONTHS}
    for tx in txns:
        if not tx.date:
            continue
        m = tx.date.month
        if m not in by_month:
            continue
        by_month[m][0 if tx.kind == "in" else 1] += (tx.amount or 0)
    r = 4
    tot_in = tot_out = 0.0
    for m in _FY_MONTHS:
        tin, tout = by_month[m]
        _cell(ws2, r, 1, _MONTH_NAME[m])
        _cell(ws2, r, 2, tin, money=True, align="right")
        _cell(ws2, r, 3, tout, money=True, align="right")
        _cell(ws2, r, 4, tin - tout, money=True, align="right")
        tot_in += tin; tot_out += tout
        r += 1
    _cell(ws2, r, 1, "รวมทั้งปี", bold=True, fill=_TOTAL)
    _cell(ws2, r, 2, tot_in, bold=True, money=True, align="right", fill=_TOTAL)
    _cell(ws2, r, 3, tot_out, bold=True, money=True, align="right", fill=_TOTAL)
    _cell(ws2, r, 4, tot_in - tot_out, bold=True, money=True, align="right", fill=_TOTAL)
    for i, w in enumerate([20, 18, 18, 20], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # ---------- ชีต 3: งบรายหมวด (แยกตามบัญชี) ----------
    has_items = any(it.fiscal_year == fiscal_year for a in accounts for it in a.items)
    if has_items:
        ws3 = wb.create_sheet("งบรายหมวด")
        ws3.merge_cells("A1:E1")
        t3 = ws3.cell(row=1, column=1, value=f"งบรายหมวดแยกตามบัญชี  ประจำปีงบประมาณ {fiscal_year}")
        t3.font = Font(name=THAI_FONT, bold=True, size=18)
        t3.alignment = Alignment(horizontal="center", vertical="center")
        for c, h in enumerate(["หมวด/รายการ", "งบที่ตั้งไว้", "รับเพิ่ม", "จ่ายแล้ว", "คงเหลือ"], start=1):
            _cell(ws3, 3, c, h, bold=True, align="center", fill=_HEAD)
        r = 4
        for a in accounts:
            items = [it for it in a.items if it.fiscal_year == fiscal_year]
            if not items:
                continue
            # หัวกลุ่ม = ชื่อบัญชี
            _cell(ws3, r, 1, a.name, bold=True, fill=_TOTAL)
            for c in range(2, 6):
                _cell(ws3, r, c, "", fill=_TOTAL)
            r += 1
            for it in items:
                tin = sum(t.amount or 0 for t in a.txns
                          if t.item_id == it.id and t.kind == "in" and t.fiscal_year == fiscal_year)
                tout = sum(t.amount or 0 for t in a.txns
                           if t.item_id == it.id and t.kind == "out" and t.fiscal_year == fiscal_year)
                _cell(ws3, r, 1, "    " + it.name)
                _cell(ws3, r, 2, float(it.budget or 0), money=True, align="right")
                _cell(ws3, r, 3, tin, money=True, align="right")
                _cell(ws3, r, 4, tout, money=True, align="right")
                _cell(ws3, r, 5, round((it.budget or 0) + tin - tout, 2), money=True, align="right")
                r += 1
        for i, w in enumerate([40, 16, 16, 16, 16], start=1):
            ws3.column_dimensions[chr(64 + i)].width = w

    out_dir = get_data_dir() / "documents"; out_dir.mkdir(exist_ok=True)
    path = out_dir / f"รายงานการเงิน_ปีงบ{fiscal_year}.xlsx"
    wb.save(str(path))
    return str(path)
