# -*- coding: utf-8 -*-
"""
admin_io.py — นำเข้าทะเบียนหนังสือรับ/ส่ง จาก Excel (สำหรับโรงเรียนที่มีทะเบียนเดิม)
ไฟล์เดียว 2 ชีต: หนังสือรับ / หนังสือส่ง  (แถว 1 = คำอธิบาย, แถว 2 = หัวตาราง, แถว 3+ = ข้อมูล)
"""
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_data_dir
from app.models import IncomingLetter, OutgoingLetter
from app.thai_utils import current_fiscal_year, parse_be_date

THAI_FONT = "TH Sarabun New"
_thin = Side(style="thin")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_HEAD = PatternFill("solid", fgColor="EDE9FE")     # ม่วงอ่อน (ธีมธุรการ)
_NOTE = PatternFill("solid", fgColor="FFF2CC")

SHEETS = {
    "หนังสือรับ": {
        "headers": ["เลขรับ", "วันที่รับ", "ที่หนังสือ", "ลงวันที่", "จาก", "ถึง/มอบให้", "เรื่อง", "การปฏิบัติ"],
        "note": "กรอกทะเบียนหนังสือรับตั้งแต่แถวที่ 3 (วันที่เป็น พ.ศ. เช่น 09/06/2569)",
        "widths": [10, 16, 16, 16, 24, 20, 34, 20],
    },
    "หนังสือส่ง": {
        "headers": ["เลขที่ส่ง", "ลงวันที่", "ถึง", "เรื่อง"],
        "note": "กรอกทะเบียนหนังสือส่งตั้งแต่แถวที่ 3 (เลขที่ส่ง เช่น ศธ 04123/45)",
        "widths": [22, 16, 28, 40],
    },
}


def _style_head(c):
    c.font = Font(name=THAI_FONT, bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _BORDER; c.fill = _HEAD


def build_admin_template() -> str:
    wb = Workbook(); wb.remove(wb.active)
    for title, spec in SHEETS.items():
        ws = wb.create_sheet(title)
        last = chr(64 + len(spec["headers"]))
        ws.merge_cells(f"A1:{last}1")
        ws["A1"] = spec["note"]
        ws["A1"].font = Font(name=THAI_FONT, bold=True, size=12, color="9C6500")
        ws["A1"].fill = _NOTE
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col, h in enumerate(spec["headers"], start=1):
            _style_head(ws.cell(row=2, column=col, value=h))
        for i, w in enumerate(spec["widths"], start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A3"
    out_dir = get_data_dir() / "documents"; out_dir.mkdir(exist_ok=True)
    path = out_dir / "เทมเพลตนำเข้าทะเบียนหนังสือ.xlsx"
    wb.save(str(path))
    return str(path)


def _s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _i(v) -> int:
    s = _s(v)
    try:
        return int(float(s)) if s else 0
    except ValueError:
        return 0


def import_admin_workbook(file_bytes: bytes, db) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    fy = current_fiscal_year()
    summary = {}

    if "หนังสือรับ" in wb.sheetnames:
        added = 0
        for row in wb["หนังสือรับ"].iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            subject = _s(row[6]) if len(row) > 6 else ""
            if not (subject or (len(row) > 0 and _s(row[0]))):
                continue
            db.add(IncomingLetter(
                fiscal_year=fy, recv_no=_i(row[0]),
                recv_date=parse_be_date(_s(row[1])) if len(row) > 1 else None,
                letter_no=_s(row[2]) if len(row) > 2 else "",
                letter_date=parse_be_date(_s(row[3])) if len(row) > 3 else None,
                from_org=_s(row[4]) if len(row) > 4 else "",
                to_person=_s(row[5]) if len(row) > 5 else "",
                subject=subject,
                action_note=_s(row[7]) if len(row) > 7 else "",
            ))
            added += 1
        if added:
            summary["หนังสือรับ"] = added

    if "หนังสือส่ง" in wb.sheetnames:
        import re
        added = 0
        for row in wb["หนังสือส่ง"].iter_rows(min_row=3, values_only=True):
            if not row or not any(row):
                continue
            send_no = _s(row[0])
            subject = _s(row[3]) if len(row) > 3 else ""
            if not (send_no or subject):
                continue
            mm = re.search(r"(\d+)", send_no)
            db.add(OutgoingLetter(
                fiscal_year=fy, send_no=send_no, send_seq=(int(mm.group(1)) if mm else 0),
                date=parse_be_date(_s(row[1])) if len(row) > 1 else None,
                to_org=_s(row[2]) if len(row) > 2 else "",
                subject=subject,
            ))
            added += 1
        if added:
            summary["หนังสือส่ง"] = added

    db.commit()
    return summary
