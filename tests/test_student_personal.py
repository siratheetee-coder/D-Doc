"""
ทดสอบ Phase 1: ข้อมูลส่วนตัวนักเรียนในทะเบียนกลาง
- ตัวจับคู่หัวคอลัมน์กันชน (ชื่อบิดา ไม่ไปเป็น name, โรคประจำตัว/เลขประจำตัวประชาชน ไม่ไปเป็น student_no)
- เพิ่ม + แก้ไขข้อมูลส่วนตัวรายคน อ่านกลับ (ผ่าน HTTP จริง)
- แก้ inline 6 ช่อง ไม่ลบข้อมูลส่วนตัว
- เทมเพลต Excel มีหัวคอลัมน์ครบ + นำเข้าแบบสลับตำแหน่งคอลัมน์แมปถูก
- เทมเพลตรวม (build_import_template) + import_workbook รอบเต็ม
รัน: .venv\\Scripts\\python.exe -m tests.test_student_personal
"""
import io
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.routers.auth as auth_mod
from app.tenancy import session_for
from app.models import Student
from app.main import app
from app.routers.pages import _student_col_map
from app.services.bulk_io import build_import_template, import_workbook

TID = 1                 # โรงเรียนบ้านหินลาด (active, ไม่หมดอายุ)
MARK = "ทดสอบพสด_"      # prefix เพื่อลบทิ้งใน finally


def _db():
    return session_for(TID)


def _cleanup(db):
    for s in db.query(Student).filter(Student.name.like(MARK + "%")).all():
        db.delete(s)
    db.commit()


def _login_client():
    """สร้าง TestClient ที่ล็อกอินแล้ว โดย monkeypatch authenticate"""
    auth_mod.authenticate = lambda u, p: {
        "uid": 1, "username": "tester", "role": "owner", "tenant_id": TID,
        "display_name": "ผู้ทดสอบ", "must_change": False,
    }
    c = TestClient(app)
    r = c.post("/login", data={"username": "tester", "password": "x"}, follow_redirects=False)
    assert r.status_code in (302, 303), f"login ไม่ผ่าน: {r.status_code}"
    return c


def test_col_map():
    header = ["ชื่อ-นามสกุล", "เลขประจำตัวประชาชน", "ชื่อบิดา", "ชื่อมารดา",
              "หมู่เลือด", "โรคประจำตัว", "เลขประจำตัว", "หมู่ที่", "เพศ", "ระดับชั้น", "ห้อง"]
    cm = _student_col_map(header)
    assert cm["name"] == 0, cm
    assert cm["id_card"] == 1, cm
    assert cm["father_name"] == 2, cm
    assert cm["mother_name"] == 3, cm
    assert cm["blood_group"] == 4, cm
    assert cm["congenital_disease"] == 5, cm
    assert cm["student_no"] == 6, cm
    assert cm["addr_moo"] == 7, cm
    assert cm["sex"] == 8 and cm["level"] == 9 and cm["room"] == 10, cm
    assert _student_col_map(["a", "b"]) == {}
    print("[ok] col_map กันชนถูกต้อง")


def test_http_add_detail_inline():
    c = _login_client()
    db = _db()
    _cleanup(db)
    try:
        r = c.post("/students", data={"name": MARK + "เด็กชายเอ", "sex": "ช",
                                      "level": "ป.6", "room": "9", "student_no": "60001"},
                   follow_redirects=False)
        assert r.status_code in (302, 303), r.status_code
        s = db.query(Student).filter(Student.name == MARK + "เด็กชายเอ").first()
        assert s is not None
        sid = s.id

        r = c.get(f"/students/{sid}")
        assert r.status_code == 200 and "ข้อมูลส่วนตัว" in r.text

        r = c.post(f"/students/{sid}/update", data={
            "_from": "detail", "name": MARK + "เด็กชายเอ", "sex": "ช", "level": "ป.6", "room": "9",
            "student_no": "60001", "id_card": "1449901075151", "father_name": "นายพ่อ ใจดี",
            "mother_name": "นางแม่ ใจดี", "race": "ไทย", "nationality": "ไทย", "religion": "พุทธ",
            "blood_group": "O", "congenital_disease": "ไม่มี", "addr_no": "12", "addr_moo": "3",
            "addr_tambon": "หินลาด", "addr_amphoe": "เมือง", "addr_province": "มหาสารคาม",
            "addr_zip": "44000", "phone": "0812345678", "enroll_date": "16/05/2567",
            "prev_school": "อนุบาลบ้านเดิม",
        }, follow_redirects=False)
        assert r.status_code in (302, 303)
        db.expire_all()
        s = db.get(Student, sid)
        assert s.id_card == "1449901075151", s.id_card
        assert s.father_name == "นายพ่อ ใจดี"
        assert s.blood_group == "O"
        assert s.addr_province == "มหาสารคาม"
        assert s.enroll_date is not None and s.enroll_date.year == 2024, s.enroll_date  # 2567 BE
        print("[ok] เพิ่ม+แก้ไขข้อมูลส่วนตัว อ่านกลับครบ")

        # แก้ inline 6 ช่อง (fetch) -> ข้อมูลส่วนตัวต้องไม่หาย
        r = c.post(f"/students/{sid}/update",
                   data={"name": MARK + "เด็กชายเอ", "sex": "ช", "birthdate": "",
                         "level": "ป.6", "room": "9", "student_no": "60001"},
                   headers={"X-Requested-With": "fetch"})
        assert r.status_code == 200
        db.expire_all()
        s = db.get(Student, sid)
        assert s.id_card == "1449901075151", "inline update ลบ id_card!"
        assert s.father_name == "นายพ่อ ใจดี", "inline update ลบ father_name!"
        print("[ok] แก้ inline 6 ช่อง ไม่ลบข้อมูลส่วนตัว")
    finally:
        _cleanup(db)
        db.close()


def test_template_and_import():
    c = _login_client()
    db = _db()
    _cleanup(db)
    try:
        r = c.get("/students/template.xlsx")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        hdr = [x.value for x in wb.active[1]]
        for want in ["ชื่อ-นามสกุล", "เลขประจำตัวประชาชน", "ชื่อบิดา", "หมู่เลือด", "โรคประจำตัว", "จังหวัด"]:
            assert want in hdr, f"เทมเพลตขาดหัว {want}"
        print(f"[ok] เทมเพลต /students มี {len(hdr)} คอลัมน์ ครบข้อมูลส่วนตัว")

        # นำเข้าไฟล์แบบสลับตำแหน่งคอลัมน์ (จับตามหัวคอลัมน์ ไม่ยึดตำแหน่ง)
        up = Workbook(); ws = up.active
        ws.append(["ชื่อบิดา", "ชื่อ-นามสกุล", "โรคประจำตัว", "เลขประจำตัวประชาชน", "หมู่เลือด", "เพศ", "ระดับชั้น"])
        ws.append(["นายพ่อบี ทดสอบ", MARK + "เด็กหญิงบี", "หอบหืด", "1100000000001", "AB", "ญ", "ป.6"])
        buf = io.BytesIO(); up.save(buf)
        r = c.post("/students/import",
                   files={"file": ("up.xlsx", buf.getvalue(),
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                   follow_redirects=False)
        assert r.status_code in (302, 303)
        db.expire_all()
        s = db.query(Student).filter(Student.name == MARK + "เด็กหญิงบี").first()
        assert s is not None, "นำเข้าไม่ได้ (name ถูก ชื่อบิดา แย่ง?)"
        assert s.father_name == "นายพ่อบี ทดสอบ", s.father_name
        assert s.congenital_disease == "หอบหืด", s.congenital_disease
        assert s.id_card == "1100000000001", s.id_card
        assert s.blood_group == "AB", s.blood_group
        print("[ok] นำเข้า /students สลับคอลัมน์ แมปข้อมูลส่วนตัวถูก")

        # เทมเพลตรวม + import_workbook รอบเต็ม
        _cleanup(db)
        path = build_import_template()
        wb2 = load_workbook(path)
        ws2 = wb2["นักเรียน"]
        hdr2 = [x.value for x in ws2[2]]     # หัวคอลัมน์อยู่แถว 2
        assert "ชื่อบิดา" in hdr2 and "เลขประจำตัวประชาชน" in hdr2, hdr2
        idx = {h: i for i, h in enumerate(hdr2)}
        rowvals = [""] * len(hdr2)
        rowvals[idx["ชื่อ-นามสกุล"]] = MARK + "เด็กชายซี"
        rowvals[idx["เลขประจำตัวประชาชน"]] = "1100000000002"
        rowvals[idx["ชื่อบิดา"]] = "นายพ่อซี"
        rowvals[idx["ระดับชั้น"]] = "ป.6"
        ws2.append(rowvals)
        buf2 = io.BytesIO(); wb2.save(buf2)
        summary = import_workbook(buf2.getvalue(), db)
        db.expire_all()
        s = db.query(Student).filter(Student.name == MARK + "เด็กชายซี").first()
        assert s is not None, f"import_workbook ไม่เข้า (summary={summary})"
        assert s.id_card == "1100000000002", s.id_card
        assert s.father_name == "นายพ่อซี", s.father_name
        print(f"[ok] เทมเพลตรวม + import_workbook แมปถูก (summary={summary.get('นักเรียน')})")
    finally:
        _cleanup(db)
        db.close()


def main():
    test_col_map()
    test_http_add_detail_inline()
    test_template_and_import()
    print("\nPhase 1 ผ่านทั้งหมด")


if __name__ == "__main__":
    main()
